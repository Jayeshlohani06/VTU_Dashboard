import dash
from dash import html, dcc, Input, Output, State, callback, dash_table, no_update, ALL
import dash_bootstrap_components as dbc
import pandas as pd
import re
import numpy as np
from functools import lru_cache
import ast
from io import StringIO, BytesIO
from cache_config import cache
from dash.exceptions import PreventUpdate
import json
from services.credit_service import load_credit_map, get_credit

# Register page
dash.register_page(__name__, path="/ranking", name="Ranking")

# ==================== Helpers ====================

def extract_numeric(roll):
    digits = re.findall(r'\d+', str(roll))
    return int(digits[-1]) if digits else 0

def assign_section(roll_no, section_ranges, usn_mapping=None):
    roll_no_str = str(roll_no).strip().upper()
    if usn_mapping and roll_no_str in usn_mapping:
         return usn_mapping[roll_no_str]
         
    roll_num = extract_numeric(roll_no)
    if section_ranges:
        for sec_name, (start, end) in section_ranges.items():
            start_num = extract_numeric(start)
            end_num = extract_numeric(end)
            if start_num <= roll_num <= end_num:
                return sec_name
    return "Unassigned"

def get_grade_point(percentage_score):
    score = pd.to_numeric(percentage_score, errors='coerce')
    if pd.isna(score): return 0
    score = float(score)
    if 90 <= score <= 100: return 10
    elif 80 <= score < 90: return 9
    elif 70 <= score < 80: return 8
    elif 60 <= score < 70: return 7
    elif 55 <= score < 60: return 6
    elif 50 <= score < 55: return 5
    elif 40 <= score < 50: return 4
    else: return 0

def _normalize_df(df, section_ranges, usn_mapping=None):
    if df.columns[0] != 'Student_ID':
        df = df.rename(columns={df.columns[0]: 'Student_ID'})
    df['Section'] = df['Student_ID'].apply(lambda x: assign_section(str(x), section_ranges, usn_mapping))
    
    # === ROBUST TOTAL MARKS CALCULATION ===
    # 1. Identify valid subject columns (ending in ' Total')
    subject_cols = [c for c in df.columns if c.strip().endswith(' Total')]
    
    # 2. Exclude aggregate columns like 'Grand Total'
    exclude_keywords = ['grand total', 'total marks', 'percentage', 'result']
    subject_cols = [c for c in subject_cols if c.strip().lower() not in exclude_keywords and 'total marks' not in c.lower()]

    # 3. Filter out "Phantom Columns" (columns present but with NO marks > 0 for any student)
    temp_numeric = df[subject_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
    valid_subject_cols = [c for c in subject_cols if temp_numeric[c].max() > 0]

    # 4. Strict Column Validation (Must have Internal/External sibling)
    # This prevents counting aggregate columns like 'Grand Total' or 'Semester Total' as subjects
    final_subject_cols = []
    for c in valid_subject_cols:
        base_name = c.rsplit(' Total', 1)[0].strip()
        # Check against common patterns
        # 1. Existence of sibling columns
        has_sibling = any(f"{base_name} {suffix}" in df.columns for suffix in ['Internal', 'External'])
        # 2. Or if the base_name looks like a VTU code (contains Digit)
        looks_like_code = any(char.isdigit() for char in base_name)
        
        if has_sibling or looks_like_code:
            final_subject_cols.append(c)
            
    if final_subject_cols:
        valid_subject_cols = final_subject_cols

    # Fallback to loose matching if strict matching fails
    if not valid_subject_cols:
        total_cols = [c for c in df.columns if any(k in c.lower() for k in ['total', 'marks', 'score'])]
        total_cols = [c for c in total_cols if c != 'Total_Marks' and c.lower() not in ['grand total', 'total marks']]
        valid_subject_cols = total_cols
    
    if valid_subject_cols:
        df[valid_subject_cols] = df[valid_subject_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
        df['Total_Marks'] = df[valid_subject_cols].sum(axis=1)
        df['__Num_Subjects_Calc'] = len(valid_subject_cols)
        
        # --- NEW: Calculate Total Internal and External ---
        internal_cols = []
        external_cols = []
        for c in valid_subject_cols:
            base_name = c.rsplit(' Total', 1)[0].strip()
            i_col = f"{base_name} Internal"
            e_col = f"{base_name} External"
            
            if i_col in df.columns:
                internal_cols.append(i_col)
            if e_col in df.columns:
                external_cols.append(e_col)
                
        if internal_cols:
             df[internal_cols] = df[internal_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
             df['Total_Internal'] = df[internal_cols].sum(axis=1)
        else:
             df['Total_Internal'] = 0
             
        if external_cols:
             df[external_cols] = df[external_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
             df['Total_External'] = df[external_cols].sum(axis=1)
        else:
             df['Total_External'] = 0
        # ----------------------------------------------------

        print(f"DEBUG: Calculated Subjects ({len(valid_subject_cols)}): {valid_subject_cols}")
    else:
        df['Total_Marks'] = 0
        df['Total_Internal'] = 0
        df['Total_External'] = 0
        df['__Num_Subjects_Calc'] = 0
    # ======================================

    result_cols = [c for c in df.columns if c.endswith('Result')]
    if result_cols:
        # ── Pre-convert columns once (vectorized) to avoid N×M per-cell overhead in apply ──
        _pre_converted = set()
        for _rc in result_cols:
            _bn = _rc.replace(' Result', '').replace('Result', '').strip()
            e_col = f"{_bn} External"
            if e_col in df.columns and e_col not in _pre_converted:
                df[e_col] = pd.to_numeric(df[e_col], errors='coerce').fillna(0)
                _pre_converted.add(e_col)
            i_col = f"{_bn} Internal"
            if i_col in df.columns and i_col not in _pre_converted:
                df[i_col] = pd.to_numeric(df[i_col], errors='coerce')  # Keep NaN (no fillna)
                _pre_converted.add(i_col)
            # Pre-clean Result column to uppercase string
            df[_rc] = df[_rc].astype(str).str.strip().str.upper()

        def calc_overall(row):
            subject_status = []
            failed_list = []
            for res_col in result_cols:
                base_name = res_col.replace(' Result', '').replace('Result', '').strip()

                # Values are already numeric / cleaned from pre-conversion
                i = row.get(f"{base_name} Internal", 0)
                e = row.get(f"{base_name} External", 0)
                r = row.get(res_col, '')

                if (e == 0) and (r in ['A', 'ABSENT', '']):
                    subject_status.append('A')
                elif r in ['F', 'FAIL']:
                    subject_status.append('F')
                    failed_list.append(base_name)
                else:
                    total_s = i + e
                    if r == '' and total_s < 35:
                         subject_status.append('F')
                         failed_list.append(base_name)
                    else:
                         subject_status.append('P')

            absent_count = subject_status.count('A')
            fail_count = subject_status.count('F')

            if not subject_status: res = 'P'
            elif absent_count == len(subject_status): res = 'A'
            elif fail_count > 0 or absent_count > 0: res = 'F'
            else: res = 'P'

            return pd.Series([res, absent_count, fail_count, ", ".join(failed_list)], index=['Overall_Result', 'Absent_Subjects', 'Failed_Subjects', 'Failed_Subject'])

        df[['Overall_Result', 'Absent_Subjects', 'Failed_Subjects', 'Failed_Subject']] = df.apply(calc_overall, axis=1)
    else:
        pass_mark = 18
        if total_cols:
            df['Overall_Result'] = df.apply(lambda row: 'F' if any(row[c] < pass_mark for c in total_cols) else 'P', axis=1)
        else:
            df['Overall_Result'] = 'P'
        df['Absent_Subjects'] = 0
        df['Failed_Subjects'] = 0
        df['Failed_Subject'] = ""
    if 'Name' not in df.columns: df['Name'] = ""
    return df

@lru_cache(maxsize=32)
def _prepare_base(session_id, section_key, usn_mapping_str=None):
    if not session_id: return pd.DataFrame()
    df = cache.get(session_id)
    if df is None: return pd.DataFrame()

    section_ranges = None
    if section_key not in (None, "None"):
        try: section_ranges = ast.literal_eval(section_key)
        except: section_ranges = None
        
    usn_mapping = None
    if usn_mapping_str not in (None, "None"):
        try: usn_mapping = ast.literal_eval(usn_mapping_str)
        except: usn_mapping = None
        
    df = _normalize_df(df, section_ranges, usn_mapping)
    return df

def _section_key(section_ranges):
    try: return repr(section_ranges)
    except: return "None"

def calculate_student_metrics(df):
    """Calculates percentage based on attempted subjects for each student (vectorized).
    Handles subjects with non-standard max marks (e.g. 200-mark projects)."""
    subject_total_cols = [
        c for c in df.columns
        if c.strip().endswith(' Total')
        and c not in ['Total_Marks', 'Grand Total']
        and 'grand total' not in c.lower()
    ]

    if not subject_total_cols:
        df['percentage'] = 0.0
        return df

    # Vectorized: detect actual max marks per subject from data
    numeric_totals = df[subject_total_cols].apply(pd.to_numeric, errors='coerce')
    # Per-subject max marks: ceil(column_max / 100) * 100
    # e.g. max=197 -> 200, max=95 -> 100
    per_subject_max = np.ceil(numeric_totals.max().clip(lower=1) / 100) * 100

    # For each student, sum the max marks of only subjects they attempted
    attempted_mask = numeric_totals.notna() & (numeric_totals > 0)
    max_marks_per_student = (attempted_mask * per_subject_max).sum(axis=1).clip(lower=1)

    total_marks = pd.to_numeric(df.get('Total_Marks', 0), errors='coerce').fillna(0)
    pct = ((total_marks / max_marks_per_student) * 100).round(2)
    pct[max_marks_per_student <= 0] = 0.0
    df['percentage'] = pct
    return df


# ==================== Styles ====================
import os as _os
_css_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "styles", "ranking_page.css")
with open(_css_path, "r", encoding="utf-8") as _f:
    PAGE_CSS_LIGHT = _f.read()




# ==================== Layout ====================

layout = dbc.Container([
    dcc.Markdown(f"<style>{PAGE_CSS_LIGHT}</style>", dangerously_allow_html=True),
    html.Link(rel="stylesheet", href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css"),

    # Header
    html.Div([
        html.H3(["🏆 Class & Section Ranking"], className="rnk-title text-center fw-bold mb-2"),
        html.P("Analyze performance with precision. Toggle between Marks and SGPA modes.", className="text-center text-muted mb-0")
    ], className="rnk-wrap mb-4 rnk-card p-4"),

    # Controls
    html.Div(
        dbc.Card(dbc.CardBody([
            dbc.Row([
                dbc.Col(html.Div([
                    dcc.Dropdown(
                        id="filter-dropdown",
                        options=[
                            {"label": "All Students", "value": "ALL"},
                            {"label": "Passed Students", "value": "PASS"},
                            {"label": "Failed Students", "value": "FAIL"},
                            {"label": "Absent Students", "value": "ABSENT"}
                        ],
                        value="ALL", clearable=False, className="shadow-sm",
                        style={"position": "relative", "zIndex": "1001"}
                    )
                ], style={"position": "relative", "zIndex": "1001"}), md=3, xs=12),
                dbc.Col(html.Div([
                    dcc.Dropdown(
                        id="section-dropdown",
                        options=[{"label": "All Sections", "value": "ALL"}],
                        value="ALL", clearable=False, className="shadow-sm",
                        style={"position": "relative", "zIndex": "1001"}
                    )
                ], style={"position": "relative", "zIndex": "1001"}), md=3, xs=12),
                dbc.Col(dbc.InputGroup([
                    dbc.InputGroupText(html.I(className="bi bi-search")),
                    dbc.Input(id="search-input", placeholder="Search ID / Name...", type="text")
                ], className="shadow-sm"), md=3, xs=12),
                dbc.Col(dbc.ButtonGroup([
                    dbc.Button("📖 Rules & Guidelines", id="open-legend", color="info", outline=True),
                    dbc.Button("Reset", id="reset-btn", color="secondary", outline=True),
                ], className="w-100 d-flex justify-content-end"), md=3, xs=12),
            ], className="g-2 rnk-controls mb-3"),

            dbc.Row([
                dbc.Col(html.Div([
                    # New Metric Selector
                    html.Div([
                        html.Span("Metric:", className="me-2 text-muted small fw-bold"),
                        dcc.RadioItems(
                            id='marks-metric-selector',
                            options=[
                                {'label': 'Total', 'value': 'total'},
                                {'label': 'Internal', 'value': 'internal'},
                                {'label': 'External', 'value': 'external'}
                            ],
                            value='total', inline=True,
                            labelClassName='px-2 py-1 border rounded-pill me-1 cursor-pointer small',
                            inputClassName="me-1"
                        )
                    ], id="metric-selector-container", className="d-inline-flex align-items-center me-3", style={"display": "none"}),
                    
                    dcc.RadioItems(
                        id='ranking-type',
                        options=[{'label': 'Marks Based', 'value': 'marks'}, {'label': 'SGPA Based', 'value': 'sgpa'}],
                        value='marks', inline=True, labelClassName='px-3 py-1 border rounded-pill me-2 cursor-pointer', inputClassName="me-2"
                    )
                ], className='d-flex justify-content-end align-items-center'), className='text-end')
            ])
        ]), className="rnk-card", style={"overflow": "visible"}), className="rnk-controls-wrap mb-4", style={"overflow": "visible", "zIndex": 1000, "position": "relative"}
    ),

    # SGPA Panel
    html.Div(id='sgpa-credit-panel'),

    # KPIs (Cards)
    html.Div([
        html.Div([
            html.H6([html.I(className="bi bi-activity me-2 text-primary"), "Performance Metrics"], className="fw-bold mb-0 text-primary"),
            dbc.Button(
                [html.I(className="bi bi-cloud-arrow-down-fill me-2"), "Download Performance Summary"], 
                id="export-all-kpis", size="sm", color="primary", outline=True, className="fw-bold shadow-sm"
            )
        ], className="d-flex justify-content-between align-items-center mb-3 mt-2 px-1"),
        dbc.Spinner(html.Div(id='kpi-cards'), color="primary")
    ], className="mb-4"),

    # VTU Category Breakdown Table
    dbc.Card(dbc.CardBody([
        html.Div([
            html.H6([html.I(className="bi bi-bar-chart me-2 text-success"), "VTU Category Breakdown"], className="fw-bold mb-0"),
            dbc.Button("📥 Category Report", id="download-category-report-btn", size="sm", color="success", outline=True, className="fw-bold")
        ], className="d-flex justify-content-between align-items-center mb-3"),
        dcc.Loading(
            type="circle",
            children=html.Div(id='category-breakdown-container')
        )
    ]), className="rnk-card mb-4"),

    # Overview (Explicitly Ordered: Top 5 -> Section -> Bottom 5)
    dbc.Row([
        # Top 5 Overall (Left)
        dbc.Col(dbc.Card(dbc.CardBody(html.Div([
            html.Div([html.H6("🥇 Top 5 Overall", className="fw-bold mb-3 me-auto text-dark")], className="d-flex"),
            html.Div(id="overall-top5")
        ])), className="rnk-card h-100"), md=4, xs=12),
        
        # Section-wise Toppers (Middle)
        dbc.Col(dbc.Card(dbc.CardBody([
            html.Div([html.H6("🏆 Section-wise Toppers", className="fw-bold mb-3 me-auto text-dark")], className="d-flex"),
            html.Div(id="section-toppers")
        ]), className="rnk-card h-100"), md=4, xs=12),
        
        # Bottom 5 (Right)
        dbc.Col(dbc.Card(dbc.CardBody(html.Div([
            html.Div([html.H6("⬇️ Bottom 5 (by Score)", className="fw-bold mb-3 me-auto text-dark")], className="d-flex"),
            html.Div(id="bottom-five")
        ])), className="rnk-card h-100"), md=4, xs=12),
    ], className="g-4 mb-4"),

    # Table
    dbc.Card(dbc.CardBody([
        html.Div([
            html.H6([html.I(className="bi bi-table me-2 text-primary"), "Detailed Ranking Table"], className="fw-bold mb-0"),
            dbc.Button("Download Excel", id="export-xlsx", color="success", outline=True, size="sm")
        ], className="d-flex justify-content-between align-items-center mb-3"),
        dcc.Loading(
            type="circle",
            children=dash_table.DataTable(
                id='ranking-table',
                columns=[], data=[],
                page_size=25, sort_action='native', filter_action='none',
                style_table={'height': '620px', 'overflowY': 'auto'}, fixed_rows={'headers': True},
                style_cell={'textAlign': 'center', 'fontFamily': 'Inter, sans-serif', 'fontSize': 13, 'padding': '12px', 'minWidth': '80px'},
                style_header={'backgroundColor': '#1f2937', 'color': 'white', 'fontWeight': '700', 'padding': '12px', 'border': '0'},
                style_data_conditional=[
                    {'if': {'row_index': 'odd'}, 'backgroundColor': 'rgba(0,0,0,0.02)'},
                    # Fail Styles (Check both Marks 'F' and SGPA 'Fail')
                    {'if': {'filter_query': '{Overall_Result} = "F"'}, 'backgroundColor': '#fff1f2', 'color': '#991b1b'},
                    {'if': {'filter_query': '{Result_Selected} = "Fail"'}, 'backgroundColor': '#fff1f2', 'color': '#991b1b'},

                    # Absent Styles
                    {'if': {'filter_query': '{Overall_Result} = "A"'}, 'backgroundColor': '#fffbeb', 'color': '#b45309', 'fontWeight': 'bold'},
                    {'if': {'filter_query': '{Result_Selected} = "Absent"'}, 'backgroundColor': '#fffbeb', 'color': '#b45309', 'fontWeight': 'bold'},
                    
                    # Top Ranks (Marks)
                    {'if': {'filter_query': '{Class_Rank} = 1'}, 'backgroundColor': '#fffbeb', 'color': '#92400e', 'fontWeight': 'bold'},
                    {'if': {'filter_query': '{Class_Rank} = 2'}, 'backgroundColor': '#f0f9ff', 'color': '#075985', 'fontWeight': 'bold'},
                    {'if': {'filter_query': '{Class_Rank} = 3'}, 'backgroundColor': '#fff7ed', 'color': '#9a3412', 'fontWeight': 'bold'},
                    
                    # Top Ranks (SGPA)
                    {'if': {'filter_query': '{SGPA_Class_Rank} = 1'}, 'backgroundColor': '#fffbeb', 'color': '#92400e', 'fontWeight': 'bold'},
                    {'if': {'filter_query': '{SGPA_Class_Rank} = 2'}, 'backgroundColor': '#f0f9ff', 'color': '#075985', 'fontWeight': 'bold'},
                    {'if': {'filter_query': '{SGPA_Class_Rank} = 3'}, 'backgroundColor': '#fff7ed', 'color': '#9a3412', 'fontWeight': 'bold'},
                ],
                row_selectable=False, cell_selectable=True, style_as_list_view=True
            )
        )
    ]), className="rnk-card"),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Student Profile")),
        dbc.ModalBody(id="student-modal-body"),
        dbc.ModalFooter(dbc.Button("Close", id="close-modal", className="ms-auto", color="secondary"))
    ], id="student-modal", is_open=False, size="lg", style={"zIndex": 10500}),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("📊 Ranking Page — Rules & Guidelines")),
        dbc.ModalBody(
            html.Div([
                html.H5("🚀 Getting Started", className="text-primary fw-bold mb-2"),
                html.P("This page ranks students based on marks or SGPA. Data is loaded from the Overview page — upload your result file there first.", className="text-muted small mb-3"),

                html.H6("🔀 Ranking Modes", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Marks Based: "), "Ranks students by total marks across selected subjects. You can also switch the metric to compare Internal or External marks only."]),
                    html.Li([html.Strong("SGPA Based: "), "Ranks students by SGPA (Semester Grade Point Average). Requires scheme & semester to be configured on the Overview page for automatic credit mapping."]),
                ], className="small"),
                html.Hr(),

                html.H6("📝 Pass / Fail / Absent — Per Subject", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Absent (Subject): "), "External Marks = 0 AND Result = 'A' or blank."]),
                    html.Li([html.Strong("Fail (Subject): "), "Result = 'F', or marks below pass threshold when Result is unavailable."]),
                    html.Li([html.Strong("Pass (Subject): "), "Result = 'P' or passing grade."]),
                ], className="small"),
                html.Hr(),

                html.H6("📝 Overall Student Result", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Pass: "), "Student has passed ALL enrolled subjects."]),
                    html.Li([html.Strong("Absent: "), "Student is absent in ALL subjects."]),
                    html.Li([html.Strong("Fail: "), "Student has failed or is absent in ANY subject (but appeared for others)."]),
                ], className="small"),
                html.Hr(),

                html.H6("🏆 Ranking Rules", className="fw-bold text-dark"),
                html.Ul([
                    html.Li("Rankings are calculated ONLY for students with Overall Result = 'Pass'."),
                    html.Li("Failed and absent students are excluded from the rank list."),
                    html.Li([html.Strong("Class Rank: "), "Position among all passed students across all sections."]),
                    html.Li([html.Strong("Section Rank: "), "Position among passed students within the same section."]),
                ], className="small"),
                html.Hr(),

                html.H6("📊 Percentage Calculation", className="fw-bold text-dark"),
                html.P([html.Strong("Formula: "), "(Total Marks ÷ (Number of Active Subjects × 100)) × 100"], className="small bg-light p-2 rounded border"),
                html.Ul([
                    html.Li("Active subjects = subjects where the student has marks > 0."),
                ], className="small"),
                html.Hr(),

                html.H6("🎓 VTU Class Categories", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Span("First Class Distinction (FCD):", className="fw-bold text-success"), " ≥ 70%"]),
                    html.Li([html.Span("First Class (FC):", className="fw-bold text-info"), " 60% – 69.99%"]),
                    html.Li([html.Span("Second Class (SC):", className="fw-bold text-warning"), " 50% – 59.99%"]),
                    html.Li([html.Span("Pass Class:", className="fw-bold text-danger"), " < 50%"]),
                ], className="small"),
                html.Hr(),

                html.H6("🧮 SGPA Calculation", className="fw-bold text-dark"),
                html.P([html.Strong("Formula: "), "SGPA = Σ(Grade Point × Credit) ÷ Σ(Credit)"], className="small bg-light p-2 rounded border"),
                html.Ul([
                    html.Li("Grade Points: 90-100 → 10, 80-89 → 9, 70-79 → 8, 60-69 → 7, 55-59 → 6, 50-54 → 5, 40-49 → 4, <40 → 0"),
                    html.Li("Only subjects with credit > 0 are included."),
                    html.Li("SGPA is computed only for passing students."),
                ], className="small"),
                html.Hr(),

                html.H6("🖱️ Interactive Features", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Click a performance card: "), "Opens a detailed list of students in that category (e.g., all FCD students). You can download the list as Excel."]),
                    html.Li([html.Strong("Click a table row: "), "Opens a Student Profile modal showing all subject marks, result, percentage, and rank."]),
                    html.Li([html.Strong("Filters: "), "Use the dropdowns at the top to filter by Result (Pass/Fail/Absent), Section, or search by USN/Name."]),
                    html.Li([html.Strong("Dark/Light Mode: "), "Toggle the theme switch for your preferred viewing experience."]),
                    html.Li([html.Strong("Download: "), "Export the full ranking table or category reports as Excel."]),
                ], className="small mb-0"),
            ])
        ),
        dbc.ModalFooter(dbc.Button("Got it!", id="close-legend", className="ms-auto", color="primary"))
    ], id="legend-modal", is_open=False, size="lg", style={"zIndex": 10500}),

    # --- Clickable KPI Popup Modal ---
    dbc.Modal([
        dbc.ModalHeader([
            dbc.ModalTitle(id="rnk-kpi-modal-title", className="fw-bold text-primary"),
            html.Div([
                dbc.Button("Download List (Excel)", id="rnk-kpi-modal-pdf-top", color="success", outline=True, size="sm", className="me-2"),
                dbc.Button("Close", id="rnk-kpi-modal-close-top", color="secondary", size="sm")
            ], className="ms-auto d-flex")
        ], close_button=False),
        dbc.ModalBody([
            dash_table.DataTable(
                id="rnk-kpi-modal-table",
                columns=[], data=[],
                style_table={"overflowX": "auto", "borderRadius": "8px", "border": "1px solid #d1d5db"},
                style_cell={"textAlign": "center", "padding": "12px", "fontSize": "13px"},
                style_header={"backgroundColor": "#1f2937", "color": "#ffffff", "fontWeight": "700"},
                page_action='none',
                style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': '#f3f4f6'}],
            )
        ]),
        dbc.ModalFooter([
            dbc.Button("Download List (Excel)", id="rnk-kpi-modal-pdf", color="success", outline=True),
            dbc.Button("Close", id="rnk-kpi-modal-close", className="ms-auto", color="secondary")
        ])
    ], id="rnk-kpi-modal", is_open=False, size="xl", style={"zIndex": 10550}),
    
    dcc.Download(id="rnk-kpi-excel-download"),

    dcc.Download(id="download-xlsx"),
    dcc.Download(id="download-category-report"),
    dcc.Download(id="download-all-kpis"),
], fluid=True, className="pb-5")


# ==================== Callbacks ====================

@callback(
    Output("legend-modal", "is_open"),
    [Input("open-legend", "n_clicks"), Input("close-legend", "n_clicks")],
    [State("legend-modal", "is_open")],
    prevent_initial_call=True
)
def toggle_legend(n1, n2, is_open): return not is_open if n1 or n2 else is_open


@callback(
    Output('section-dropdown', 'options'), 
    Input('stored-data', 'data'), 
    Input('section-data', 'data'),
    Input('usn-mapping-store', 'data')
)
def update_section_options(session_id, section_ranges, usn_mapping):
    if not session_id: return [{"label": "All Sections", "value": "ALL"}]
    
    # Use cached preparer
    sec_str = str(section_ranges) if section_ranges else "None"
    usn_str = str(usn_mapping) if usn_mapping else "None"
    
    df_norm = _prepare_base(session_id, sec_str, usn_str)
    if df_norm.empty: return [{"label": "All Sections", "value": "ALL"}]
    
    sections = sorted(df_norm['Section'].dropna().unique())
    return [{"label": "All Sections", "value": "ALL"}] + [{"label": s, "value": s} for s in sections]

@callback([Output('filter-dropdown', 'value'), Output('section-dropdown', 'value'), Output('search-input', 'value')], Input('reset-btn', 'n_clicks'), prevent_initial_call=True)
def reset_filters(n): return "ALL", "ALL", ""

@callback(
    Output("metric-selector-container", "style"),
    Input("ranking-type", "value")
)
def toggle_metric_selector(rank_type):
    if rank_type == 'marks':
        return {"display": "inline-flex", "alignItems": "center", "marginRight": "1rem"}
    return {"display": "none"}

# ========== Grid UI for SGPA Panel ==========
@callback(
    Output('sgpa-credit-panel', 'children'),
    Input('stored-data', 'data'),
    Input('ranking-type', 'value'),
    Input('scheme-semester-store', 'data'),
    State('section-data', 'data'),
    State('cycle-store', 'data')
)
def generate_credit_panel(session_id, ranking_type, scheme_sem_data, section_ranges, cycle_data):
    if ranking_type != 'sgpa': return html.Div()
    if not session_id: return ""
    
    # Extract scheme/sem from store if available
    scheme = "2022"
    semester = None
    if scheme_sem_data:
        scheme = scheme_sem_data.get('scheme', '2022')
        semester = scheme_sem_data.get('semester')
    
    df = cache.get(session_id)
    if df is None: return ""
    
    codes = set()
    for col in df.columns:
        m = re.match(r'^(.*?)\s+(Internal|External|Total)$', col, flags=re.IGNORECASE)
        if m: codes.add(m.group(1).strip())

    if not codes: return dbc.Alert("No recognizable subject columns found.", color='info')
    codes = sorted(codes)
    
    # Auto-detect semester from subject codes if not provided
    if semester is None:
        for code in codes:
            subj_code = code.split(" - ")[0].strip() if " - " in code else code.strip()
            sem_match = re.search(r'(\d)\d{2}', subj_code)
            if sem_match:
                semester = int(sem_match.group(1))
                print(f"[RankingCredits] Auto-detected semester={semester} from {subj_code}")
                break
        if semester is None:
            semester = 5  # ultimate fallback
    
    # Load credit map using the stored/detected scheme, semester, and cycle
    credit_map = load_credit_map(scheme, semester, cycle=cycle_data)
    
    grid_items = []
    for code in codes:
        # Extract only the subject code (everything before " - " if present)
        display_code = code.split(" - ")[0].strip() if " - " in code else code.strip()
        
        # Get credit from service, fallback to default 0 if not found
        fetched_credit = get_credit(display_code, credit_map)
        default_credit = str(fetched_credit)

        grid_items.append(dbc.Col(dbc.InputGroup([
            # InputGroupText for the label
            dbc.InputGroupText(
                html.Span(display_code, title=code),
                className="fw-bold bg-light text-dark text-truncate d-inline-block",
                style={
                    "width": "110px", 
                    "justifyContent": "center", 
                    "fontSize": "0.85rem",
                    "whiteSpace": "nowrap",
                    "overflow": "hidden",
                    "textOverflow": "ellipsis"
                }
            ),
            # Select dropdown for credits
            dbc.Select(
                id={'type': 'credit-input', 'index': code}, 
                options=[{'label': f'{i} Credits', 'value': str(i)} for i in range(10, -1, -1)], 
                value=default_credit, 
                className="form-select text-center flex-grow-1",
                style={"minHeight": "45px", "fontSize": "14px"}
            )
        ], size="sm", className="shadow-sm mb-3 d-flex flex-nowrap", style={"overflow": "hidden"}), xs=12, sm=6, md=4, lg=3))

    return dbc.Card([
        dbc.CardHeader(html.Div([html.I(className="bi bi-sliders me-2"), "SGPA Configuration"], className="fw-bold text-primary"), className="bg-white border-bottom-0 pt-3", style={"overflow": "visible"}),
        dbc.CardBody([
            html.P("Assign credits to subjects. The system will calculate SGPA based on these values.", className="text-muted small mb-4"),
            dbc.Row(grid_items, className="g-2 mb-2", style={"overflow": "visible"}),
            html.Hr(className="my-3 text-muted opacity-25"),
            html.P("✨ SGPA is computed automatically based on these values.", className="text-success small fw-bold mb-0 text-center"),
            # Status Container (starts empty)
            html.Div(id="sgpa-calc-status", className="mt-2")
        ], style={"overflow": "visible"})
    ], className="rnk-card mb-4 border-start border-4 border-primary", style={"overflow": "visible"})

# ========== SGPA Calculation (FIXED PASS/FAIL LOGIC) ==========
@callback(
    Output('sgpa-store', 'data'),
    Output('sgpa-calc-status', 'children'),
    Input({'type': 'credit-input', 'index': ALL}, 'value'),
    State('stored-data', 'data'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data'),
    State({'type': 'credit-input', 'index': ALL}, 'id'),
    prevent_initial_call=False
)
def calculate_sgpa_all(credit_vals, json_data, section_ranges, usn_mapping, credit_ids):
    if not json_data or not credit_vals: return no_update, no_update
    
    mapping_str = str(usn_mapping) if usn_mapping else "None"
    base = _prepare_base(json_data, _section_key(section_ranges), mapping_str).copy()
    credit_dict = {}
    for cid, val in zip(credit_ids, credit_vals):
        if val is not None:
            try: credit_dict[cid['index']] = int(val)
            except ValueError: credit_dict[cid['index']] = 0

    credit_dict_positive = {k: v for k, v in credit_dict.items() if v > 0}
    if not credit_dict_positive:
        return no_update, dbc.Alert("Please assign at least one credit > 0", color="warning", dismissable=True)

    # Pre-compute max marks per subject from data (handles 200-mark subjects)
    _subj_max_marks = {}
    for code in credit_dict_positive:
        total_col = f"{code} Total"
        if total_col in base.columns:
            col_max = pd.to_numeric(base[total_col], errors='coerce').max()
            _subj_max_marks[code] = int(np.ceil(max(col_max, 1) / 100) * 100) if pd.notna(col_max) else 100
        else:
            _subj_max_marks[code] = 100

    # Pre-compute per-subject participation rate (vectorized)
    # Compulsory subjects have high participation (>50%), electives have low participation.
    # This distinguishes "student didn't enroll" (elective, skip) from
    # "student didn't appear" (compulsory, count with GP=0).
    total_students = len(base)
    _subj_participation = {}
    _subj_has_data = {}  # Per-subject boolean Series: which students have data
    for code in credit_dict_positive:
        has_data = pd.Series(False, index=base.index)
        for suffix in ['Total', 'Internal', 'External']:
            col = f"{code} {suffix}"
            if col in base.columns:
                has_data |= (pd.to_numeric(base[col], errors='coerce').fillna(0) > 0)
        res_col = f"{code} Result"
        if res_col in base.columns:
            cleaned = base[res_col].astype(str).str.strip().str.upper()
            has_data |= ~cleaned.isin(['', 'NAN', 'NONE', 'NA', '-'])
        _subj_participation[code] = has_data.sum() / total_students if total_students > 0 else 0
        _subj_has_data[code] = has_data

    # Detect mutual-exclusion groups: subjects sharing the same course number
    # e.g. BKBK109 and BKSK109 both extract to "109" — students take one or the other
    from collections import defaultdict
    _course_num_to_codes = defaultdict(list)
    for code in credit_dict_positive:
        display_code = code.split(" - ")[0].strip() if " - " in code else code.strip()
        num = re.search(r'\d{3}', display_code)
        if num:
            _course_num_to_codes[num.group()].append(code)
    # Build sibling lookup: for each code, list of other codes with the same course number
    _sibling_codes = {}
    for num, codes_list in _course_num_to_codes.items():
        if len(codes_list) > 1:
            for c in codes_list:
                _sibling_codes[c] = [s for s in codes_list if s != c]

    sgpa_rows = []
    for idx, row in base.iterrows():
        total_cp, total_cre, total_marks, fail_flag = 0, 0, 0, False
        for code, credit in credit_dict_positive.items():
            # 1. Get raw scores
            i = pd.to_numeric(row.get(f"{code} Internal"), errors='coerce') or 0
            e = pd.to_numeric(row.get(f"{code} External"), errors='coerce') or 0
            
            # 2. Get Total Score
            if f"{code} Total" in base.columns:
                score = pd.to_numeric(row.get(f"{code} Total"), errors='coerce') or 0
            else:
                score = (i + e) if (i and e) else (i or e or 0)
            
            # 3. Check Result column
            res_val = str(row.get(f"{code} Result", "")).strip().upper()

            # 4. Determine if student has data for this subject
            has_marks = (i > 0) or (e > 0) or (score > 0)
            has_result = bool(res_val) and res_val not in ('', 'NAN', 'NONE', 'NA', '-')

            if not has_marks and not has_result:
                # Check mutual-exclusion: if a sibling subject (same course number)
                # has data for this student, skip — they took the other variant.
                siblings = _sibling_codes.get(code, [])
                if siblings and any(_subj_has_data[s].at[idx] for s in siblings if s in _subj_has_data):
                    continue  # Student enrolled in sibling variant, skip this one

                # No data for this student. Check if it's compulsory or elective.
                # Compulsory (>50% of students have data): count with GP=0 (didn't appear)
                # Elective  (<=50% participation):        skip (not enrolled)
                if _subj_participation.get(code, 0) > 0.5:
                    fail_flag = True
                    total_cp += 0  # GP = 0
                    total_cre += credit
                continue

            # 5. Fail logic
            if res_val in ('P', 'PASS'):
                pass
            elif res_val in ('F', 'FAIL', 'NP'):
                fail_flag = True
            elif res_val in ('A', 'AB', 'ABSENT'):
                fail_flag = True
            else:
                if score < 35: 
                    fail_flag = True

            max_m = _subj_max_marks.get(code, 100)
            pct = (score / max_m * 100) if max_m > 0 else 0
            gp = get_grade_point(pct)

            total_cp += gp * credit
            total_cre += credit
            total_marks += score
            
        sgpa = (total_cp / total_cre) if total_cre > 0 else 0.0

        ovr = str(row.get('Overall_Result', '')).strip().upper()

        if ovr in ['A', 'ABSENT']:
            res = 'Absent'
        elif ovr in ['F', 'FAIL']:
             # If the student failed overall (in the raw data), they should fail in SGPA mode too.
             # This aligns the two views and prevents "Pass" in SGPA for a failed student.
             res = 'Fail'
        elif ovr in ['P', 'PASS']:
             # If marked PASS in Marks Mode, FORCE PASS here.
             # We assume Marks Mode correctly identified passing status.
             # This resolves the discrepancy where SGPA fails students (27) while Marks fails fewer (25).
             res = 'Pass'
        else:
             # Fallback
             res = 'Pass' if (not fail_flag and total_cre > 0) else 'Fail'
        
        # Override: ONLY if fail_flag is True AND we don't have a definitive Overall Pass AND not Absent
        if fail_flag and res != 'Pass' and res != 'Absent':
            res = 'Fail'
        
        # Double Check: If Overall is Fail, SGPA must be Fail.
        if ovr in ['F', 'FAIL']:
            res = 'Fail'
            
        sgpa_rows.append({'Student_ID': row['Student_ID'], 'SGPA': round(sgpa, 2), 'Total_Marks_Selected': round(total_marks, 2), 'Result_Selected': res})

    sgpa_df = pd.DataFrame(sgpa_rows)
    sgpa_df['SGPA_Class_Rank'] = sgpa_df[sgpa_df['Result_Selected'] == 'Pass']['SGPA'].rank(method='min', ascending=False).astype('Int64')
    section_map = base.set_index('Student_ID')['Section'].to_dict()
    sgpa_df['Section'] = sgpa_df['Student_ID'].map(section_map)
    sgpa_df['SGPA_Section_Rank'] = sgpa_df.groupby('Section')['SGPA'].rank(method='min', ascending=False).astype('Int64')

    msg = dbc.Alert([html.I(className="bi bi-check-circle-fill me-2"), "Calculation Successful! Dashboard Updated."], color="success", dismissable=True, is_open=True, fade=True)
    return sgpa_df.to_json(orient='split'), msg

# ========== Main View Builder (Dynamic KPIs + Fixed Layout Order) ==========
@callback(
    Output('kpi-cards', 'children'),
    Output('overall-top5', 'children'),
    Output('section-toppers', 'children'),
    Output('bottom-five', 'children'),
    Output('ranking-table', 'columns'),
    Output('ranking-table', 'data'),
    Output('category-breakdown-container', 'children'),
    Input('filter-dropdown', 'value'),
    Input('section-dropdown', 'value'),
    Input('search-input', 'value'),
    Input('ranking-type', 'value'),
    Input('marks-metric-selector', 'value'),
    Input('sgpa-store', 'data'),
    State('stored-data', 'data'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data')
)
def build_views(filter_val, sec_val, search_val, rank_type, metric_val, sgpa_json, json_data, section_ranges, usn_mapping):
    if not json_data: return html.P("Upload data first.", className="text-center text-muted"), html.Div(), html.Div(), html.Div(), [], [], html.Div()

    mapping_str = str(usn_mapping) if usn_mapping else "None"
    base_full = _prepare_base(json_data, _section_key(section_ranges), mapping_str).copy()
    base_pre = base_full.copy()
    if sgpa_json:
        try:
            sgpa_df = pd.read_json(StringIO(sgpa_json), orient='split')
            # Drop columns from sgpa_df that already exist in base_full (except join key)
            # to avoid _x/_y suffix conflicts after merge
            overlap_cols = [c for c in sgpa_df.columns if c in base_full.columns and c != 'Student_ID']
            sgpa_df = sgpa_df.drop(columns=overlap_cols)
            base_full = base_full.merge(sgpa_df, how='left', on='Student_ID')
        except: base_full = base_pre.copy()

    scope = base_full.copy()

    # Determine Sort Column and Target Result Column
    target_res_col = "Overall_Result"
    sort_col = "Total_Marks"
    
    if rank_type == 'sgpa':
        if 'Result_Selected' in scope.columns: target_res_col = "Result_Selected"
        if 'SGPA' in scope.columns: sort_col = "SGPA"
    elif rank_type == 'marks':
        if metric_val == 'internal': sort_col = 'Total_Internal'
        elif metric_val == 'external': sort_col = 'Total_External'
        else: sort_col = 'Total_Marks'

    # Dynamic Filtering
    pass_val = ["P", "PASS"]
    fail_val = ["F", "FAIL"]
    absent_val = ["A", "ABSENT"]

    if rank_type == 'sgpa' and 'Result_Selected' in scope.columns:
        pass_val = ["PASS", "Pass"]
        fail_val = ["FAIL", "Fail"]
        absent_val = ["ABSENT", "Absent"]

    if filter_val == "PASS":
        if target_res_col in scope.columns:
            scope = scope[scope[target_res_col].astype(str).str.upper().isin([p.upper() for p in pass_val])]
    elif filter_val == "FAIL":
        if target_res_col in scope.columns:
            scope = scope[scope[target_res_col].astype(str).str.upper().isin([f.upper() for f in fail_val])]
    elif filter_val == "ABSENT":
        if target_res_col in scope.columns:
            scope = scope[scope[target_res_col].astype(str).str.upper().isin([a.upper() for a in absent_val])]
    
    if sec_val != "ALL" and 'Section' in scope.columns: scope = scope[scope["Section"] == sec_val]

    # Calculate Ranks based on Sort Column
    if sort_col in scope.columns:
        scope['Class_Rank'] = pd.NA
        # Rank only passing students
        pass_mask = scope[target_res_col].apply(lambda x: str(x).upper() in [p.upper() for p in pass_val])
        
        scope.loc[pass_mask, 'Class_Rank'] = (
            scope.loc[pass_mask, sort_col]
            .rank(method='min', ascending=False)
            .astype('Int64')
        )

    if 'Section' in scope.columns and sort_col in scope.columns:
        scope['Section_Rank'] = pd.NA
        for sec in scope['Section'].unique():
            # Check for pass in this section
            sec_mask = (scope['Section'] == sec) & (scope[target_res_col].apply(lambda x: str(x).upper() in [p.upper() for p in pass_val]))
            scope.loc[sec_mask, 'Section_Rank'] = (
                scope.loc[sec_mask, sort_col]
                .rank(method='min', ascending=False)
                .astype('Int64')
            )

    # --- KPI Logic (Dynamic Visibility) ---
    total = len(scope)
    
    # Calculate counts based on the filtered scope
    def check_res(val, allowed): return str(val).upper() in [x.upper() for x in allowed]
    
    passed = scope[target_res_col].apply(lambda x: check_res(x, pass_val)).sum() if target_res_col in scope.columns else 0
    absent = scope[target_res_col].apply(lambda x: check_res(x, absent_val)).sum() if target_res_col in scope.columns else 0
    failed = scope[target_res_col].apply(lambda x: check_res(x, fail_val)).sum() if target_res_col in scope.columns else 0
    
    # 'Appeared' Logic: conceptually Total - Absent (or Passed + Failed if filtering is weird)
    # If filtered to 'Absent', Total=Absent, Appeared=0.
    appeared = total - absent
    pass_pct = round((passed / appeared) * 100, 2) if appeared > 0 else 0
    
    # =====================================================
    # ✅ PER-STUDENT ACCURATE PERCENTAGE CALCULATION
    # =====================================================
    
    scope_calc = calculate_student_metrics(scope.copy())
    
    # Only PASSING students get a class
    pass_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, pass_val))
    
    fcd_count = ((scope_calc['percentage'] >= 70) & pass_mask).sum()
    fc_count = ((scope_calc['percentage'] >= 60) & (scope_calc['percentage'] < 70) & pass_mask).sum()
    sc_count = ((scope_calc['percentage'] >= 50) & (scope_calc['percentage'] < 60) & pass_mask).sum()
    
    # Define KPIs matching the user request style
    kpi_objs = [
        {"id": "total", "label": "Total Students", "value": total, "color": "#3b82f6", "bg": "#eff6ff"},
    ]

    # Calculate Backlog counts - Only for students with Result='F'
    # We include Absent_Subjects as backlogs for failed students
    fail_1, fail_2, fail_3plus = 0, 0, 0
    
    if 'Failed_Subjects' in scope.columns and 'Absent_Subjects' in scope.columns:
        # Filter strictly for Failed students (matches the 'Failed' KPI logic)
        is_fail_mask = scope[target_res_col].apply(lambda x: check_res(x, fail_val))
        failed_df = scope[is_fail_mask].copy()

        if not failed_df.empty:
            # Backlog count = Fails + partial Absents
            # (Note: Fully absent students have Result='A', so they are excluded here, which is correct because they aren't in 'Failed' count)
            backlogs = failed_df['Failed_Subjects'] + failed_df['Absent_Subjects']
            
            fail_1 = (backlogs == 1).sum()
            fail_2 = (backlogs == 2).sum()
            fail_3plus = (backlogs >= 3).sum()

    if filter_val == "ALL":
        kpi_objs.extend([
            {"id": "appeared", "label": "Appeared", "value": appeared, "color": "#10b981", "bg": "#ecfdf5"},
            {"id": "absent", "label": "Absent", "value": absent, "color": "#f59e0b", "bg": "#fffbeb"},
            {"id": "pass", "label": "Passed", "value": passed, "color": "#0ea5e9", "bg": "#f0f9ff"},
            {"id": "fail", "label": "Failed", "value": failed, "color": "#ef4444", "bg": "#fef2f2"},
            {"id": "rate", "label": "Pass % (Appeared)", "value": f"{pass_pct}%", "color": "#8b5cf6", "bg": "#f5f3ff"},
        ])
        # Add Backlog Breakdown for ALL view
        if failed > 0:
            kpi_objs.extend([
                {"id": "bk1", "label": "1 Subject Fail", "value": fail_1, "color": "#fb923c", "bg": "#fff7ed"},
                {"id": "bk2", "label": "2 Subject Fails", "value": fail_2, "color": "#f97316", "bg": "#fff7ed"},
                {"id": "bk3", "label": "3+ Subject Fails", "value": fail_3plus, "color": "#ef4444", "bg": "#fef2f2"},
            ])
    
    elif filter_val == "PASS":
        kpi_objs.append({"id": "pass", "label": "Passed", "value": passed, "color": "#0ea5e9", "bg": "#f0f9ff"})
    elif filter_val == "FAIL":
        kpi_objs.append({"id": "fail", "label": "Failed", "value": failed, "color": "#ef4444", "bg": "#fef2f2"})
        kpi_objs.extend([
            {"id": "bk1", "label": "1 Subject Fail", "value": fail_1, "color": "#fb923c", "bg": "#fff7ed"},
            {"id": "bk2", "label": "2 Subject Fails", "value": fail_2, "color": "#f97316", "bg": "#fff7ed"},
            {"id": "bk3", "label": "3+ Subject Fails", "value": fail_3plus, "color": "#ef4444", "bg": "#fef2f2"},
        ])
    elif filter_val == "ABSENT": 
        kpi_objs.append({"id": "absent", "label": "Absent", "value": absent, "color": "#f59e0b", "bg": "#fffbeb"})

    # Add VTU Categories (Always visible for context unless irrelevant)
    if filter_val in ["ALL", "PASS"]:
        vtu_kpis = [
            {"id": "fcd", "label": "First Class Distinction (≥70%)", "value": fcd_count, "color": "#8b5cf6", "bg": "#faf5ff"},
            {"id": "fc", "label": "First Class (60-70%)", "value": fc_count, "color": "#0ea5e9", "bg": "#f0f9ff"},
            {"id": "sc", "label": "Second Class (50-60%)", "value": sc_count, "color": "#f59e0b", "bg": "#fffbf0"},
        ]
        kpi_objs.extend(vtu_kpis)
    
    if rank_type == 'sgpa' and 'SGPA' in scope.columns:
        avg = round(scope['SGPA'].mean(), 2) if not scope['SGPA'].isna().all() else 0
        kpi_objs.insert(0, {"id": "avg", "label": "Avg SGPA", "value": avg, "color": "#7c3aed", "bg": "#f5f3ff"})

    # Dynamic column width based on count
    count = len(kpi_objs)
    # Force 6 cards row for ALL view 
    row_cls = "row-cols-2 row-cols-md-3 row-cols-lg-6 g-3" if count >= 6 else f"row-cols-2 row-cols-md-{min(count, 4)} row-cols-lg-{min(count, 4)} g-3"

    # Assign icons based on ID
    icon_map = {
        "total": "bi-people-fill",
        "appeared": "bi-person-circle",
        "absent": "bi-person-x-fill",
        "pass": "bi-check-circle-fill",
        "fail": "bi-x-circle-fill",
        "rate": "bi-percent",
        "bk1": "bi-1-circle-fill",
        "bk2": "bi-2-circle-fill",
        "bk3": "bi-3-circle-fill",
        "fcd": "bi-trophy-fill",
        "fc": "bi-award-fill",
        "sc": "bi-patch-check-fill",
        "avg": "bi-calculator-fill"
    }

    kpi_cards = dbc.Row([
        dbc.Col(
            html.Div(
                dbc.CardBody([
                    html.Div([
                        html.Div(
                            html.I(className=f"bi {icon_map.get(x['id'], 'bi-graph-up-arrow')}", style={"color": x["color"], "fontSize": "1.4rem"}),
                             className="kpi-icon-box d-flex align-items-center justify-content-center",
                             style={"minWidth": "42px", "width": "42px", "height": "42px", "borderRadius": "10px", "backgroundColor": x["bg"]}
                        ),
                        html.Div([
                            html.H6(x["label"], className="text-muted text-uppercase fw-bold mb-0", style={"fontSize": "0.7rem", "letterSpacing": "0.5px"}),
                            html.H3(str(x["value"]), className="kpi-val fw-bold mb-0", style={"color": x["color"], "fontSize": "1.6rem"})
                        ], className="ms-3")
                    ], className="d-flex align-items-center h-100"),
                    html.Div("👆 Click for details", className="kpi-hover-hint text-muted text-end mt-1", style={"fontSize": "0.6rem", "opacity": "0.8", "position": "absolute", "bottom": "8px", "right": "12px", "display": "none", "transition": "opacity 0.2s ease"}),
                ], className="p-3 position-relative"),
                className="card kpi-card shadow-sm h-100 border-0 rnk-kpi-clickable rnk-card",
                style={"borderLeft": f"4px solid {x['color']}", "transition": "transform 0.2s ease-in-out", "cursor": "pointer"},
                title=f"Click to view {x['label'].lower()} students list",
                id={'type': 'rnk-kpi-card', 'index': x['id']},
                n_clicks=0
            )
        )
        for x in kpi_objs
    ], className=row_cls)

    def make_list(df, val_col, is_asc=False):
        if df.empty: return html.P("No Data", className="text-muted small")
        sorted_df = df.sort_values(val_col, ascending=is_asc).head(5)
        items = []
        for i, (_, r) in enumerate(sorted_df.iterrows(), 1):
            val = r.get(val_col, 0)
            
            if rank_type == 'sgpa' and 'Result_Selected' in r:
                is_pass = (r['Result_Selected'] == 'Pass')
                res_txt = r['Result_Selected']
            else:
                is_pass = (r.get('Overall_Result') == 'P')
                res_txt = r.get('Overall_Result')

            res_txt = str(res_txt) if pd.notna(res_txt) else "-"
            res_cls = "badge-pass" if is_pass else "badge-fail"
            
            sec_txt = f" (Sec {r.get('Section')})" if r.get('Section') else ""
            
            items.append(html.Li([
                html.Span(f"#{i}", className="fw-bold me-2 text-muted", style={"minWidth":"25px"}),
                html.Span([html.Strong(r.get('Student_ID')), html.Span(sec_txt, className="text-muted small")]),
                html.Span("—", className="mx-2 text-muted"),
                html.Span(f"{val}", className="fw-bold me-2"),
                html.Span(res_txt[0], className=res_cls) # Just P or F
            ], className="d-flex align-items-center mb-2 small"))
        return html.Ul(items, className="list-unstyled mb-0")

    # Ensure sort_col exists (fallback)
    if sort_col not in scope.columns:
         sort_col = 'Total_Marks' if 'Total_Marks' in scope.columns else (scope.columns[0] if len(scope.columns)>0 else "")

    top5_html = make_list(scope, sort_col, False)
    bot_html = make_list(scope, sort_col, True)

    sec_cards = []
    if 'Section' in scope.columns:
        for sec, g in sorted(scope.groupby('Section')):
            if g.empty: continue
            # Handle cases where sort_col might be all NaN
            if g[sort_col].isna().all(): continue
            
            ridx = g[sort_col].idxmax()
            r = g.loc[ridx]
            
            rank_col_name = 'SGPA_Class_Rank' if rank_type=='sgpa' else 'Class_Rank'
            rank_val = r.get(rank_col_name)
            if pd.notna(rank_val) and str(rank_val).replace('.', '', 1).isdigit():
                rank_display = str(int(rank_val))
            else:
                rank_display = "-"
            
            val_label = 'Score'
            if rank_type == 'sgpa': val_label = 'SGPA'
            elif metric_val == 'internal': val_label = 'Internal'
            elif metric_val == 'external': val_label = 'External'
            else: val_label = 'Total'

            card = html.Div([
                html.H6([html.I(className="bi bi-trophy-fill me-2", style={"color":"#8b5cf6"}), f"Section {sec} Topper"], className="fw-bold mb-2", style={"color": "#4338ca", "fontSize": "0.95rem"}),
                html.Div([
                    html.Div(f"Student ID: {r.get('Student_ID')}", className="mb-1 text-muted small"),
                    html.Div([html.Span(f"{val_label}: ", className="text-muted small"), html.Span(f"{r.get(sort_col)}", className="fw-bold text-dark")], className="mb-1"),
                    html.Div([html.Span("Class Rank: ", className="text-muted small"), html.Span(rank_display, className="fw-bold text-dark")])
                ])
            ], className="p-3 mb-3 bg-white rounded shadow-sm border-start border-4 border-primary")
            sec_cards.append(card)
            
    sec_html = html.Div(sec_cards) if sec_cards else html.P("No Data", className="text-muted small")

    tdf = scope.copy()
    if search_val:
        s = str(search_val).strip().lower()
        mask = (
            tdf['Student_ID'].astype(str).str.strip().str.lower() == s
        ) | (
            tdf['Name'].astype(str).str.strip().str.lower() == s
        )
        tdf = tdf[mask]
    
    if rank_type == 'sgpa' and 'SGPA' in tdf.columns:
        tdf = tdf.sort_values('SGPA', ascending=False)
        cols = ['SGPA_Class_Rank', 'SGPA_Section_Rank', 'Student_ID', 'Name', 'Section', 'SGPA', 'Total_Marks_Selected', 'Result_Selected']
    else:
        tdf['__sort'] = tdf['Class_Rank'].fillna(9999)
        tdf = tdf.sort_values(['__sort', sort_col], ascending=[True, False])
        cols = ['Class_Rank', 'Section_Rank', 'Student_ID', 'Name', 'Section', sort_col, 'Overall_Result']
    
    # Filter to only columns that exist in data
    cols = [c for c in cols if c in tdf.columns]
    tcols = [{"name": c.replace("_", " "), "id": c} for c in cols]
    tdata = tdf[cols].to_dict('records')

    # === GENERATE VTU CATEGORY BREAKDOWN TABLE ===
    breakdown_data = []
    if 'Section' in scope.columns:
        sections = sorted(scope['Section'].unique())
    else:
        sections = ["Overall"]
    
    for section in sections:
        if section == "Overall":
            section_df = scope_calc
        else:
            section_df = scope_calc[scope_calc['Section'] == section]
        
        if section_df.empty:
            continue
        
        # Check Pass Status using result column logic
        is_pass = section_df[target_res_col].apply(lambda x: check_res(x, pass_val))
        
        fcd = ((section_df['percentage'] >= 70) & is_pass).sum()
        fc = ((section_df['percentage'] >= 60) & (section_df['percentage'] < 70) & is_pass).sum()
        sc = ((section_df['percentage'] >= 50) & (section_df['percentage'] < 60) & is_pass).sum()
        pc = ((section_df['percentage'] < 50) & is_pass).sum()
        fail = section_df[target_res_col].apply(lambda x: check_res(x, fail_val)).sum()
        absent = section_df[target_res_col].apply(lambda x: check_res(x, absent_val)).sum()
        total_sec = len(section_df)
        
        breakdown_data.append({
            'Section': section,
            'First Class Distinction (≥70%)': fcd,
            'First Class (60-70%)': fc,
            'Second Class (50-60%)': sc,
            'Pass Class (<50%)': pc,
            'Failed': fail,
            'Absent': absent,
            'Total': total_sec
        })
    
    # Add overall row
    if breakdown_data:
        overall_row = {
            'Section': 'Overall',
            'First Class Distinction (≥70%)': sum(row['First Class Distinction (≥70%)'] for row in breakdown_data),
            'First Class (60-70%)': sum(row['First Class (60-70%)'] for row in breakdown_data),
            'Second Class (50-60%)': sum(row['Second Class (50-60%)'] for row in breakdown_data),
            'Pass Class (<50%)': sum(row['Pass Class (<50%)'] for row in breakdown_data),
            'Failed': sum(row['Failed'] for row in breakdown_data),
            'Absent': sum(row['Absent'] for row in breakdown_data),
            'Total': sum(row['Total'] for row in breakdown_data)
        }
        breakdown_data.append(overall_row)
    
    # Create accordion items for each section
    accordion_items = []
    for bd_row in breakdown_data:
        section_name = bd_row['Section']
        if section_name == "Overall":
            section_df = scope_calc
        else:
            section_df = scope_calc[scope_calc['Section'] == section_name]
        
        is_pass = section_df[target_res_col].apply(lambda x: check_res(x, pass_val))
        is_fail = section_df[target_res_col].apply(lambda x: check_res(x, fail_val))
        is_absent = section_df[target_res_col].apply(lambda x: check_res(x, absent_val))
        
        # Create category breakdown rows
        category_items = []
        categories = [
            ('FCD (≥70%) - Passed Only', 'success', section_df[(section_df['percentage'] >= 70) & is_pass]),
            ('First Class (60-70%) - Passed Only', 'info', section_df[(section_df['percentage'] >= 60) & (section_df['percentage'] < 70) & is_pass]),
            ('Second Class (50-60%) - Passed Only', 'warning', section_df[(section_df['percentage'] >= 50) & (section_df['percentage'] < 60) & is_pass]),
            ('Pass Class (<50%) - Passed Only', 'primary', section_df[(section_df['percentage'] < 50) & is_pass]),
            ('Failed', 'danger', section_df[is_fail]),
            ('Absent', 'secondary', section_df[is_absent])
        ]
        
        for cat_name, cat_color, cat_df in categories:
            if len(cat_df) > 0:
                # Optimized: Use DataTable instead of HTML Loop for performance
                dt_data = cat_df.copy()
                dt_data['id'] = dt_data['Student_ID'].astype(str) # Vital for active_cell row_id
                
                # Format percentage for display
                dt_data['percentage_disp'] = dt_data['percentage'].apply(lambda x: f"{x:.2f}%")
                
                student_table = dash_table.DataTable(
                    id={'type': 'breakdown-table', 'section': section_name, 'category': cat_name},
                    columns=[
                        {'name': 'Student ID', 'id': 'Student_ID'},
                        {'name': 'Name', 'id': 'Name'},
                        {'name': 'Marks', 'id': 'Total_Marks'},
                        {'name': 'Percentage', 'id': 'percentage_disp'}
                    ],
                    data=dt_data.to_dict('records'),
                    row_selectable=False,
                    cell_selectable=True,
                    style_as_list_view=True,
                    style_table={'minWidth': '100%'},
                    style_header={
                        'backgroundColor': '#f1f5f9', 
                        'fontWeight': '700', 
                        'borderBottom': '2px solid #cbd5e1',
                        'color': '#334155',
                        'padding': '12px'
                    },
                    style_cell={
                        'textAlign': 'left', 
                        'padding': '12px', 
                        'fontFamily': 'Inter, sans-serif', 
                        'cursor': 'pointer',
                        'color': '#1f2937',
                        'fontSize': '0.9rem',
                        'backgroundColor': 'transparent'
                    },
                    style_data_conditional=[
                        # Selected Cell
                        {'if': {'state': 'active'}, 'backgroundColor': '#eff6ff', 'border': '1px solid #60a5fa'},
                        
                        # Student ID (Blue & Bold)
                        {'if': {'column_id': 'Student_ID'}, 'color': '#2563eb', 'fontWeight': 'bold'},
                        
                        # Name (Semi-Bold)
                        {'if': {'column_id': 'Name'}, 'fontWeight': '500'},
                        
                        # Stats (Bold & Darker)
                        {'if': {'column_id': 'Total_Marks'}, 'fontWeight': 'bold', 'color': '#111827'},
                        {'if': {'column_id': 'percentage_disp'}, 'fontWeight': 'bold', 'color': '#111827'},
                        
                        # Zebra Striping (Subtle)
                        {'if': {'row_index': 'odd'}, 'backgroundColor': '#f8fafc'}
                    ]
                )
                
                category_items.append(
                    dbc.AccordionItem([
                        student_table
                    ], title=f"📊 {cat_name} ({len(cat_df)} students)", className="mb-2")
                )
        
        # Main accordion item for section
        section_title = f"Section {section_name} - Total: {bd_row['Total']} | First Class Distinction: {bd_row['First Class Distinction (≥70%)']}" if section_name != "Overall" else f"📈 Overall Summary - Total: {bd_row['Total']}"
        
        accordion_items.append(
            dbc.AccordionItem(
                dbc.Accordion(category_items, always_open=False, start_collapsed=True),
                title=section_title,
                className="mb-2"
            )
        )
    
    # Create main accordion
    breakdown_container = dbc.Accordion(accordion_items, always_open=False) if accordion_items else html.P("No data available", className="text-muted")

    return kpi_cards, top5_html, sec_html, bot_html, tcols, tdata, breakdown_container

@callback(
    Output("student-modal", "is_open"), 
    Output("student-modal-body", "children"), 
    Input("ranking-table", "active_cell"), 
    Input({'type': 'breakdown-table', 'section': ALL, 'category': ALL}, 'active_cell'),
    State("ranking-table", "derived_viewport_data"), 
    State("stored-data", "data"),
    State("section-data", "data"),
    State("sgpa-store", "data"),
    State("usn-mapping-store", "data"),
    Input("close-modal", "n_clicks"), 
    prevent_initial_call=True
)
def show_modal(main_cell, bd_cells, main_data, json_data, section_data, sgpa_json, usn_mapping, close):
    trigger = dash.ctx.triggered_id
    if trigger == "close-modal" or not json_data: return False, no_update
    
    student_id = None
    
    # 1. Check Main Ranking Table
    if trigger == "ranking-table" and main_cell and main_data:
        # Safety check for row index
        if main_cell['row'] < len(main_data):
            student_id = main_data[main_cell['row']].get('Student_ID')
        
    # 2. Check Breakdown Tables
    elif isinstance(trigger, dict) and trigger.get('type') == 'breakdown-table':
        # Retrieve the relevant active_cell from the list
        # ctx.triggered consists of a list of changed properties.
        # We need to find the specific input that triggered.
        # However, dealing with ALL here is tricky to get value directly.
        # But wait! We injected 'id' into the data rows.
        # So active_cell will contain 'row_id' which IS the Student_ID!
        
        # Iterate to find the non-None cell that triggered (or use context inputs)
        # Dash context inputs_list is reliable
        for input_item in dash.ctx.inputs_list[1]: # Index 1 corresponds to breakdown-table input
            if input_item['id'] == trigger and input_item.get('value'):
                student_id = input_item['value'].get('row_id')
                break
        
    if not student_id: return no_update, no_update

    # Use Cached Data Loader for Speed
    # _prepare_base is lru_cached, so it won't re-parse JSON if string is identical
    try:
        mapping_str = str(usn_mapping) if usn_mapping else "None"
        df = _prepare_base(json_data, _section_key(section_data), mapping_str).copy()
        
        if sgpa_json:
            try:
                sgpa_df = pd.read_json(StringIO(sgpa_json), orient='split')
                overlap_cols = [c for c in sgpa_df.columns if c in df.columns and c != 'Student_ID']
                sgpa_df = sgpa_df.drop(columns=overlap_cols)
                if 'Student_ID' in df.columns and 'Student_ID' in sgpa_df.columns:
                    df = df.merge(sgpa_df, how='left', on='Student_ID')
                elif df.columns[0] == 'USN' or df.columns[0] == 'Student_ID':
                     df = df.rename(columns={df.columns[0]: 'Student_ID'})
                     if 'Student_ID' in sgpa_df.columns:
                         df = df.merge(sgpa_df, how='left', on='Student_ID')
            except Exception as e:
                pass

        if df.empty: return no_update, no_update
    except:
        return no_update, no_update

    if 'Student_ID' not in df.columns: df = df.rename(columns={df.columns[0]: 'Student_ID'})
    
    # Optimize filtering (convert to string once)
    student_row = df[df['Student_ID'].astype(str) == str(student_id)]
    
    if student_row.empty: return True, "Student data not found for ID: " + str(student_id)
    
    row = student_row.iloc[0]
    
    # Identify subject data dynamically
    processed_subjects = {}

    for col in df.columns:
        if ' ' not in col: continue
        
        prefix, suffix = col.rsplit(' ', 1)
        if suffix in ['Internal', 'External', 'Total', 'Result']:
            # Handle "Code - Name" pattern or just "Code"
            if " - " in prefix:
                code_part = prefix.split(" - ", 1)[0].strip()
                name_part = prefix.split(" - ", 1)[1].strip()
            else:
                code_part = prefix.strip()
                name_part = prefix.strip() # Fallback if no name

            if code_part not in processed_subjects:
                processed_subjects[code_part] = {"name": name_part, "prefix": prefix}

    subject_codes = sorted(list(processed_subjects.keys()))
    
    rows = []
    for code in subject_codes:
        meta = processed_subjects[code]
        prefix = meta["prefix"]
        name = meta["name"]
        
        i_val = row.get(f"{prefix} Internal")
        e_val = row.get(f"{prefix} External")
        t_val = row.get(f"{prefix} Total")
        r_val = row.get(f"{prefix} Result")

        # STRICT VALIDATION: Filter out subjects the student didn't take
        # We treat a subject as "not mapped" if:
        # 1. Total, Internal, and External are all 0 (or missing/NaN)
        # 2. AND Result is empty or missing
        
        i_num = pd.to_numeric(i_val, errors='coerce') or 0
        e_num = pd.to_numeric(e_val, errors='coerce') or 0
        t_num = pd.to_numeric(t_val, errors='coerce') or 0
        
        has_marks = (i_num > 0) or (e_num > 0) or (t_num > 0)
        
        # Check if result is meaningful (not NaN, None, empty string)
        has_result = False
        if pd.notna(r_val):
            r_str = str(r_val).strip()
            if r_str and r_str.upper() not in ['NAN', 'NONE']:
                has_result = True

        if not has_marks and not has_result:
            continue

        # Helper for display
        def fmt(v): return v if pd.notna(v) and v != 0 else ("0" if v == 0 else "-")
        # Note: We want to show 0 if it exists, but usually data has 0.
        # Let's stick to original fmt or similar.
        # Original: def fmt(v): return v if pd.notna(v) else "-"
        # Better: keep 0 as 0.
        def fmt_disp(v): 
             return v if pd.notna(v) else "-"

        res_disp = fmt_disp(r_val)    

        # Use full subject name if available
        display_subject = f"{code} - {name}" if name != code else code

        rows.append(html.Tr([
            html.Td(display_subject, className="fw-bold text-start"),
            html.Td(fmt_disp(i_val)),
            html.Td(fmt_disp(e_val)),
            html.Td(fmt_disp(t_val), className="fw-bold"),
            html.Td(res_disp, className="text-danger fw-bold" if str(res_disp).upper() in ['F', 'FAIL'] else "text-success fw-bold")
        ]))

    sgpa_val = row.get('SGPA', 'N/A')
    if pd.notna(sgpa_val) and str(sgpa_val).strip() != 'N/A':
        try:
            sgpa_val = f"{float(sgpa_val):.2f}"
        except ValueError:
            pass  # keep original if it fails to convert

    body = html.Div([
        dbc.Row([
            dbc.Col([
                html.H4(row.get('Name', 'Unknown'), className="fw-bold text-primary"),
                html.H6(f"USN: {student_id}", className="text-muted")
            ], width=8),
            dbc.Col([
                dbc.Badge(f"Total: {row.get('Total_Marks', 0)}", color="info", className="p-2 fs-6 me-2"),
                dbc.Badge(f"SGPA: {sgpa_val}", color="success", className="p-2 fs-6")
            ], width=4, className="text-end align-self-center")
        ], className="mb-3 border-bottom pb-3"),
        
        dbc.Table([
            html.Thead(html.Tr([html.Th("Subject"), html.Th("Internal"), html.Th("External"), html.Th("Total"), html.Th("Result")])),
            html.Tbody(rows)
        ], bordered=True, hover=True, striped=True, responsive=True, className="text-center small")
    ])
    
    return True, body

@callback(
    Output("download-xlsx", "data"), 
    Input("export-xlsx", "n_clicks"), 
    State('ranking-table', 'data'), 
    State('ranking-table', 'columns'), 
    prevent_initial_call=True
)
def exp_xlsx(n, d, cols): 
    if not d: return no_update
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    df = pd.DataFrame(d)
    
    # Filter and rename according to displayed columns
    if cols:
        col_ids = [c['id'] for c in cols if c['id'] in df.columns]
        df = df[col_ids]
        df.columns = [c['name'] for c in cols if c['id'] in df.columns]
        
    out = BytesIO()
    writer = pd.ExcelWriter(out, engine='openpyxl')
    df.to_excel(writer, sheet_name="Ranking", index=False)
    
    worksheet = writer.sheets["Ranking"]
    
    # Define Styles
    rank1_style = {"bg": "FFFFFBEB", "font": "FF92400E"}
    rank2_style = {"bg": "FFF0F9FF", "font": "FF075985"}
    rank3_style = {"bg": "FFFFF7ED", "font": "FF9A3412"}
    fail_style = {"bg": "FFFFF1F2", "font": "FF991B1B"}
    absent_style = {"bg": "FFFFFBEB", "font": "FFB45309"}
    odd_row_bg = "FFFAFAFA"
    
    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    
    # Apply column widths and header styles
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 5, 12)
        if "Name" in col_name:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 30
            
    # Apply row styles based on conditions (ignoring original index, checking dataframe data)
    # We need to map dataframe column names back to what is being evaluated (overall_result, class_rank, etc)
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        row_dict = dict(zip(df.columns, row_data))
        
        # Check conditions
        is_fail = str(row_dict.get("Overall Result", "")).upper() == "F" or str(row_dict.get("Result Selected", "")).upper() == "FAIL"
        is_absent = str(row_dict.get("Overall Result", "")).upper() == "A" or str(row_dict.get("Result Selected", "")).upper() == "ABSENT"
        
        rank_val = row_dict.get("Class Rank", row_dict.get("SGPA Class Rank", ""))
        try:
            rank_val = float(rank_val) if pd.notna(rank_val) else 999
        except (ValueError, TypeError):
            rank_val = 999
            
        # Determine row style
        bg_color = None
        font_color = "FF000000"
        is_bold = False
        
        if is_fail:
            bg_color, font_color = fail_style["bg"], fail_style["font"]
        elif is_absent:
            bg_color, font_color = absent_style["bg"], absent_style["font"]
            is_bold = True
        elif rank_val == 1:
            bg_color, font_color = rank1_style["bg"], rank1_style["font"]
            is_bold = True
        elif rank_val == 2:
            bg_color, font_color = rank2_style["bg"], rank2_style["font"]
            is_bold = True
        elif rank_val == 3:
            bg_color, font_color = rank3_style["bg"], rank3_style["font"]
            is_bold = True
        elif row_idx % 2 == 1:
            bg_color = odd_row_bg
            
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            if font_color != "FF000000" or is_bold:
                cell.font = Font(color=font_color, bold=is_bold)

    writer.close()
    out.seek(0)
    return dcc.send_bytes(out.read(), "ranking_table.xlsx")

@callback(
    Output("download-all-kpis", "data"),
    Input("export-all-kpis", "n_clicks"),
    State('filter-dropdown', 'value'),
    State('section-dropdown', 'value'),
    State('ranking-type', 'value'),
    State('stored-data', 'data'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data'),
    State('sgpa-store', 'data'),
    prevent_initial_call=True
)
def export_all_kpis_report(n_clicks, filter_val, sec_val, rank_type, json_data, sec_ranges, usn_mapping, sgpa_json):
    if not json_data: return no_update

    mapping_str = str(usn_mapping) if usn_mapping else "None"
    base_full = _prepare_base(json_data, _section_key(sec_ranges), mapping_str).copy()
    
    if sgpa_json:
        try:
            sgpa_df = pd.read_json(StringIO(sgpa_json), orient='split')
            overlap_cols = [c for c in sgpa_df.columns if c in base_full.columns and c != 'Student_ID']
            sgpa_df = sgpa_df.drop(columns=overlap_cols)
            base_full = base_full.merge(sgpa_df, how='left', on='Student_ID')
        except: pass

    scope = base_full.copy()
    target_res_col = "Result_Selected" if rank_type == 'sgpa' and 'Result_Selected' in scope.columns else "Overall_Result"
    
    pass_val = ["PASS", "Pass"] if rank_type == 'sgpa' else ["P", "PASS"]
    fail_val = ["FAIL", "Fail"] if rank_type == 'sgpa' else ["F", "FAIL"]
    absent_val = ["ABSENT", "Absent"] if rank_type == 'sgpa' else ["A", "ABSENT"]

    def check_res(val, allowed): return str(val).upper() in [x.upper() for x in allowed]

    if filter_val == "PASS": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, pass_val))]
    elif filter_val == "FAIL": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, fail_val))]
    elif filter_val == "ABSENT": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, absent_val))]

    if sec_val != "ALL" and 'Section' in scope.columns: scope = scope[scope["Section"] == sec_val]

    scope_calc = calculate_student_metrics(scope.copy())

    is_pass_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, pass_val))
    is_fail_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, fail_val))
    is_absent_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, absent_val))

    backlogs = pd.Series(0, index=scope_calc.index)
    if 'Failed_Subjects' in scope_calc.columns and 'Absent_Subjects' in scope_calc.columns:
        backlogs = scope_calc['Failed_Subjects'] + scope_calc['Absent_Subjects']

    kpi_definitions = [
        ('Total Students', scope_calc),
        ('Appeared', scope_calc[~is_absent_mask]),
        ('Absent', scope_calc[is_absent_mask]),
        ('Passed', scope_calc[is_pass_mask]),
        ('Failed', scope_calc[is_fail_mask]),
        ('1 Subject Fail', scope_calc[is_fail_mask & (backlogs == 1)]),
        ('2 Subject Fails', scope_calc[is_fail_mask & (backlogs == 2)]),
        ('3+ Subject Fails', scope_calc[is_fail_mask & (backlogs >= 3)]),
        ('First Class Distinction', scope_calc[is_pass_mask & (scope_calc['percentage'] >= 70)]),
        ('First Class', scope_calc[is_pass_mask & (scope_calc['percentage'] >= 60) & (scope_calc['percentage'] < 70)]),
        ('Second Class', scope_calc[is_pass_mask & (scope_calc['percentage'] >= 50) & (scope_calc['percentage'] < 60)])
    ]
    
    # Overview Calculation
    overview_data = []
    for sheet_name, df_kpi in kpi_definitions:
        overview_data.append({"Metric": sheet_name, "Count": len(df_kpi)})
        
    total_app = len(scope_calc[~is_absent_mask])
    pass_cnt = len(scope_calc[is_pass_mask])
    pass_perc = round((pass_cnt / total_app * 100) if total_app > 0 else 0, 2)
    overview_data.insert(5, {"Metric": "Pass % (Appeared)", "Count": f"{pass_perc}%"})
    
    overview_df = pd.DataFrame(overview_data)

    out = BytesIO()
    writer = pd.ExcelWriter(out, engine='openpyxl')
    
    # Write Overview as the first sheet
    overview_df.to_excel(writer, sheet_name="Performance Overview", index=False)
    
    has_sheets = True
    for sheet_name, df in kpi_definitions:
        if df.empty:
            continue
            
        display_cols = ['Student_ID', 'Name', 'Section', 'Total_Marks', 'percentage']
        if rank_type == 'sgpa' and 'SGPA' in df.columns: display_cols.append('SGPA')
        if 'Failed_Subject' in df.columns and 'Fail' in sheet_name: display_cols.append('Failed_Subject')
        
        final_cols = [c for c in display_cols if c in df.columns]
        export_df = df[final_cols].copy()
        
        if 'percentage' in export_df.columns:
            export_df['percentage'] = export_df['percentage'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")
            
        export_df.rename(columns={'Total_Marks': 'Marks', 'percentage': 'Percentage (%)', 'Failed_Subject': 'Failed Subject(s)'}, inplace=True)
        
        valid_sheet_name = sheet_name[:31]
        export_df.to_excel(writer, sheet_name=valid_sheet_name, index=False)
        has_sheets = True
        
    if not has_sheets:
        empty_df = pd.DataFrame({"Message": ["No data available"]})
        empty_df.to_excel(writer, sheet_name="No Data", index=False)
        
    writer.close()
    out.seek(0)
    
    return dcc.send_bytes(out.read(), "Consolidated_Performance_Report.xlsx")

# ==================== Download Reports ====================

@callback(
    Output("download-category-report", "data"),
    Input("download-category-report-btn", "n_clicks"),
    State("stored-data", "data"),
    State("section-data", "data"),
    State("usn-mapping-store", "data"),
    prevent_initial_call=True
)
def download_category_report(n_clicks, json_data, section_data, usn_mapping):
    if not json_data: return no_update
    
    mapping_str = str(usn_mapping) if usn_mapping else "None"
    # Load Full Data (All Sections) - Use copy to avoid modifying cached data
    df = _prepare_base(json_data, _section_key(section_data), mapping_str).copy()
    
    # Apply Metrics (Percentage)
    df = calculate_student_metrics(df)
    
    # Create Excel Buffer
    out = BytesIO()
    writer = pd.ExcelWriter(out, engine='openpyxl')
    
    # Define Export Columns
    desired_cols = ['Student_ID', 'Name', 'Section', 'Total_Marks', 'percentage', 'Overall_Result', 'Failed_Subjects', 'Absent_Subjects']
    export_cols = [c for c in desired_cols if c in df.columns]
    
    # Define Export Columns for Failed Sheet
    failed_desired_cols = desired_cols + ['Failed_Subject']
    failed_export_cols = [c for c in failed_desired_cols if c in df.columns]
    
    # Filter Pass/Fail/Absent
    pass_complete = df[df['Overall_Result'] == 'P'].copy()
    fail_students = df[df['Overall_Result'] == 'F'].copy()
    absent_students = df[df['Overall_Result'] == 'A'].copy()
    
    # 1. First Class Distinction (>= 70%)
    fcd = pass_complete[pass_complete['percentage'] >= 70]
    if not fcd.empty: fcd[export_cols].to_excel(writer, sheet_name='FCD (Distinction)', index=False)
    
    # 2. First Class (60% - 69.99%)
    fc = pass_complete[(pass_complete['percentage'] >= 60) & (pass_complete['percentage'] < 70)]
    if not fc.empty: fc[export_cols].to_excel(writer, sheet_name='First Class', index=False)
    
    # 3. Second Class (50% - 59.99%)
    sc = pass_complete[(pass_complete['percentage'] >= 50) & (pass_complete['percentage'] < 60)]
    if not sc.empty: sc[export_cols].to_excel(writer, sheet_name='Second Class', index=False)
    
    # 4. Pass Class (< 50%)
    pc = pass_complete[pass_complete['percentage'] < 50]
    if not pc.empty: pc[export_cols].to_excel(writer, sheet_name='Pass Class', index=False)
    
    # 5. Failed
    if not fail_students.empty: 
        fail_students[failed_export_cols].to_excel(writer, sheet_name='Failed', index=False)
        
    # 6. Absent
    if not absent_students.empty:
        absent_students[export_cols].to_excel(writer, sheet_name='Absent', index=False)
        
    # Save
    writer.close()
    out.seek(0)
    
    return dcc.send_bytes(out.read(), "VTU_Category_Report.xlsx")

# ==================== KPI Modal & Export Logic ====================

@callback(
    Output("rnk-kpi-modal", "is_open"),
    Output("rnk-kpi-modal-title", "children"),
    Output("rnk-kpi-modal-table", "data"),
    Output("rnk-kpi-modal-table", "columns"),
    Input({"type": "rnk-kpi-card", "index": ALL}, "n_clicks"),
    Input("rnk-kpi-modal-close", "n_clicks"),
    Input("rnk-kpi-modal-close-top", "n_clicks"),
    State('filter-dropdown', 'value'),
    State('section-dropdown', 'value'),
    State('ranking-type', 'value'),
    State('stored-data', 'data'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data'),
    State('sgpa-store', 'data'),
    prevent_initial_call=True
)
def handle_rnk_kpi_click(kpi_clicks, c1, c2, filter_val, sec_val, rank_type, json_data, sec_ranges, usn_mapping, sgpa_json):
    if not dash.ctx.triggered: 
        raise PreventUpdate

    trigger = dash.ctx.triggered_id

    # Handle Modal Close
    if trigger in ["rnk-kpi-modal-close", "rnk-kpi-modal-close-top"]:
        return False, dash.no_update, dash.no_update, dash.no_update

    # Handle Dictionary ID matching
    if not isinstance(trigger, dict) or trigger.get("type") != "rnk-kpi-card":
        raise PreventUpdate
        
    # Prevent triggering on render when n_clicks are 0
    if all(c == 0 or c is None for c in kpi_clicks): 
        raise PreventUpdate
        
    kpi_type = trigger.get("index")

    if not json_data: return True, "No Data Available", [], []

    mapping_str = str(usn_mapping) if usn_mapping else "None"
    base_full = _prepare_base(json_data, _section_key(sec_ranges), mapping_str).copy()
    
    if sgpa_json:
        try:
            sgpa_df = pd.read_json(StringIO(sgpa_json), orient='split')
            overlap_cols = [c for c in sgpa_df.columns if c in base_full.columns and c != 'Student_ID']
            sgpa_df = sgpa_df.drop(columns=overlap_cols)
            base_full = base_full.merge(sgpa_df, how='left', on='Student_ID')
        except: pass

    scope = base_full.copy()
    target_res_col = "Result_Selected" if rank_type == 'sgpa' and 'Result_Selected' in scope.columns else "Overall_Result"
    
    pass_val = ["PASS", "Pass"] if rank_type == 'sgpa' else ["P", "PASS"]
    fail_val = ["FAIL", "Fail"] if rank_type == 'sgpa' else ["F", "FAIL"]
    absent_val = ["ABSENT", "Absent"] if rank_type == 'sgpa' else ["A", "ABSENT"]

    def check_res(val, allowed): return str(val).upper() in [x.upper() for x in allowed]

    if filter_val == "PASS": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, pass_val))]
    elif filter_val == "FAIL": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, fail_val))]
    elif filter_val == "ABSENT": scope = scope[scope[target_res_col].apply(lambda x: check_res(x, absent_val))]

    if sec_val != "ALL" and 'Section' in scope.columns: scope = scope[scope["Section"] == sec_val]

    scope_calc = calculate_student_metrics(scope.copy())

    is_pass_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, pass_val))
    is_fail_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, fail_val))
    is_absent_mask = scope_calc[target_res_col].apply(lambda x: check_res(x, absent_val))

    backlogs = pd.Series(0, index=scope_calc.index)
    if 'Failed_Subjects' in scope_calc.columns and 'Absent_Subjects' in scope_calc.columns:
        backlogs = scope_calc['Failed_Subjects'] + scope_calc['Absent_Subjects']

    # Filter Logic
    if kpi_type == 'total': res_df = scope_calc
    elif kpi_type == 'appeared': res_df = scope_calc[~is_absent_mask]
    elif kpi_type == 'absent': res_df = scope_calc[is_absent_mask]
    elif kpi_type == 'pass' or kpi_type == 'rate': res_df = scope_calc[is_pass_mask]
    elif kpi_type == 'fail': res_df = scope_calc[is_fail_mask]
    elif kpi_type == 'bk1': res_df = scope_calc[is_fail_mask & (backlogs == 1)]
    elif kpi_type == 'bk2': res_df = scope_calc[is_fail_mask & (backlogs == 2)]
    elif kpi_type == 'bk3': res_df = scope_calc[is_fail_mask & (backlogs >= 3)]
    elif kpi_type == 'fcd': res_df = scope_calc[is_pass_mask & (scope_calc['percentage'] >= 70)]
    elif kpi_type == 'fc': res_df = scope_calc[is_pass_mask & (scope_calc['percentage'] >= 60) & (scope_calc['percentage'] < 70)]
    elif kpi_type == 'sc': res_df = scope_calc[is_pass_mask & (scope_calc['percentage'] >= 50) & (scope_calc['percentage'] < 60)]
    elif kpi_type == 'avg': return True, "Average SGPA represents the full view calculated in Ranking Mode.", [], []
    else: res_df = scope_calc

    # KPI Names mapping
    kpi_labels = {
        'total': 'Total Students',
        'appeared': 'Appeared',
        'absent': 'Absent',
        'pass': 'Passed',
        'fail': 'Failed',
        'rate': 'Pass Percentage',
        'bk1': '1 Subject Fail',
        'bk2': '2 Subject Fails',
        'bk3': '3+ Subject Fails',
        'fcd': 'First Class Distinction',
        'fc': 'First Class',
        'sc': 'Second Class',
        'avg': 'Average SGPA'
    }
    
    kpi_display_name = kpi_labels.get(kpi_type, str(kpi_type).upper())

    if res_df.empty:
        return True, f"Student List: {kpi_display_name} (0 Students)", [], []

    display_cols = ['Student_ID', 'Name', 'Section', 'Total_Marks', 'percentage']
    if rank_type == 'sgpa' and 'SGPA' in res_df.columns: display_cols.append('SGPA')
    if 'Failed_Subject' in res_df.columns and kpi_type in ['fail', 'bk1', 'bk2', 'bk3']: display_cols.append('Failed_Subject')

    final_cols = [c for c in display_cols if c in res_df.columns]
    col_map = {'Total_Marks': 'Marks', 'percentage': 'Percentage (%)', 'Failed_Subject': 'Failed Subject(s)'}
    tcols = [{"name": col_map.get(c, c.replace('_', ' ')), "id": c} for c in final_cols]

    if 'percentage' in res_df.columns:
        res_df['percentage'] = res_df['percentage'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

    tdata = res_df[final_cols].to_dict('records')
    title = f"📃 Detail List: {kpi_display_name} ({len(res_df)} Students)"

    return True, title, tdata, tcols

# ...existing code...
@callback(
    Output("rnk-kpi-excel-download", "data"),
    Input("rnk-kpi-modal-pdf", "n_clicks"),
    Input("rnk-kpi-modal-pdf-top", "n_clicks"),
    State("rnk-kpi-modal-table", "data"),
    State("rnk-kpi-modal-table", "columns"),
    State("rnk-kpi-modal-title", "children"),
    prevent_initial_call=True
)
def download_modal_excel(n_clicks_bottom, n_clicks_top, table_data, table_cols, title):
    if not dash.ctx.triggered or not table_data:
        raise PreventUpdate
        
    df = pd.DataFrame(table_data)
    
    # Reorder and rename columns precisely to match the visible modal table map
    if table_cols:
        col_map = {col['id']: col['name'] for col in table_cols if 'name' in col and 'id' in col}
        cols_to_keep = [col['id'] for col in table_cols if col['id'] in df.columns]
        if cols_to_keep:
            df = df[cols_to_keep]
            df.rename(columns=col_map, inplace=True)
    
    # Create a clean filename from the modal title
    safe_title = "Student_List"
    if title and isinstance(title, str):
        # Extract just the category part before the parentheses
        clean_str = title.split('(')[0].replace('📃', '').replace('Detail List:', '').strip()
        safe_title = re.sub(r'[^A-Za-z0-9_]', '_', clean_str)
        
    # Standard excel output for simple tables
    if isinstance(df.columns, pd.MultiIndex):
        df.index = [""] * len(df)
        df.index.name = None
        return dcc.send_data_frame(df.to_excel, f"{safe_title}_Report.xlsx", index=True)
    else:
        return dcc.send_data_frame(df.to_excel, f"{safe_title}_Report.xlsx", index=False)
    

dash.clientside_callback(
    """
    function(n_clicks_bottom, n_clicks_top) {
        if (!n_clicks_bottom && !n_clicks_top) return window.dash_clientside.no_update;
        setTimeout(function () { window.print(); }, 150);
        return "";
    }
    """,
    Output("rnk-kpi-pdf-trigger-hidden", "children"),
    Input("rnk-kpi-modal-pdf", "n_clicks"),
    Input("rnk-kpi-modal-pdf-top", "n_clicks"),
    prevent_initial_call=True
)