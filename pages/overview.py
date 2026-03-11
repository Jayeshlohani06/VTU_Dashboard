import dash
from dash import html, dcc, Input, Output, State, callback, ALL, no_update, ctx, dash_table
import dash_bootstrap_components as dbc
import pandas as pd
import base64
import io
import re
import uuid
from cache_config import cache
from dash.exceptions import PreventUpdate

dash.register_page(__name__, path='/', name="Overview")

# ==================== Styles ====================

PAGE_CSS_LIGHT = r"""
:root{
  --bg: #f5f7fb;
  --card: #ffffff;
  --text: #1f2937;
  --muted:#6b7280;
  --shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
  --shadow-hover: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
  --k1:#fffbeb; --k2:#eff6ff; --k3:#fff7ed; --k45:#f8fafc;
  --pass-bg:#ecfdf5; --pass-text:#065f46;
  --fail-bg:#fef2f2; --fail-text:#991b1b;
}
.rnk-wrap{ background: var(--bg); padding: 20px; border-radius: 16px; }
.rnk-card{
  background: var(--card); border: 0 !important; border-radius: 12px !important;
  box-shadow: var(--shadow); transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
}
.rnk-card:hover{ transform: translateY(-2px); box-shadow: var(--shadow-hover); }
.kpi-card{ border-left: 4px solid transparent; height: 100%; display: flex; flex-direction: column; justify-content: center; }
.kpi-label{ color: var(--muted); font-size: 0.85rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-value{ font-weight: 800; font-size: 2.2rem; line-height: 1.2; }
.rank-chip{ display:inline-flex; align-items:center; justify-content:center; width:28px; height:28px; border-radius:50%; font-weight:700; font-size:0.9rem; margin-right:8px; }
.rank-1{ background:var(--k1); color:#b45309; border:1px solid #fcd34d; }
.rank-2{ background:var(--k2); color:#1e40af; border:1px solid #93c5fd; }
.rank-3{ background:var(--k3); color:#9a3412; border:1px solid #fdba74; }
.rank-4,.rank-5{ background:var(--k45); color:#cbd5e1; border:1px solid #475569; }
.badge-pass{ background:var(--pass-bg); color:var(--pass-text); padding:2px 8px; border-radius:12px; font-size:0.75rem; font-weight:700; }
.badge-fail{ background:var(--fail-bg); color:var(--fail-text); padding:2px 8px; border-radius:12px; font-size:0.75rem; font-weight:700; }
.overview-kpi-clickable:hover { transform: translateY(-4px); box-shadow: 0 12px 28px rgba(0,0,0,0.12) !important; cursor: pointer; }
.overview-kpi-clickable:hover .kpi-hover-hint { display: block !important; }
"""



# ---------- HELPER FUNCTIONS ----------

def process_uploaded_excel(contents):
    """Processes Excel with Multi-Index headers and cleans column names."""
    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        
        # Determine header depth
        # Read first few rows as generic
        df_preview = pd.read_excel(io.BytesIO(decoded), header=None, nrows=10)
        
        header_row_count = 2 # Default
        for i, row in df_preview.iterrows():
            row_str = row.astype(str).str.lower().tolist()
            has_internal = any("internal" in x for x in row_str)
            has_external = any("external" in x for x in row_str)
            
            if has_internal and has_external:
                # Detected the component row. Its index + 1 is the header count.
                # e.g. if internal is at index 1 -> header rows are 0,1 (count 2)
                # e.g. if internal is at index 2 -> header rows are 0,1,2 (count 3)
                header_row_count = i + 1
                break
        
        header_indices = list(range(header_row_count))
        df_raw = pd.read_excel(io.BytesIO(decoded), header=header_indices)

        fixed_cols = []
        last_valid_code = None

        def is_empty(h): return str(h).lower() == "nan" or str(h).startswith("Unnamed:")
        
        # Normalize to tuple access regardless of depth
        cols = df_raw.columns
        for col_tuple in cols:
            # Pad with empty strings if not 3 items (unlikely if header=list)
            # Actually pandas MultiIndex will have tuples of length == len(header_indices)
            
            # Map to H1 (Code), H2 (Name or Component), H3 (Component or Empty)
            if header_row_count == 3:
                h1 = str(col_tuple[0]).strip() # Code
                h2 = str(col_tuple[1]).strip() # Name
                h3 = str(col_tuple[2]).strip() # Component
                component = h3
            else:
                h1 = str(col_tuple[0]).strip() # Code
                h2 = str(col_tuple[1]).strip() # Component
                component = h2
            
            # Forward Fill Subject Code (from H1)
            # Only fill if we're in a subject block (where component is present)
            if not is_empty(h1):
                last_valid_code = h1
            elif last_valid_code:
                # If h1 is empty, use the last valid code
                h1 = last_valid_code

            # Determine Column Name
            if is_empty(component):
                # Identity Column (Name, USN)
                # It might have been preserved in H1 or H2 in the merged cells
                val = h1
                if is_empty(val) and header_row_count == 3: val = h2
                
                # Normalize common names
                v_lower = val.lower()
                if "name" in v_lower and "code" not in v_lower:
                    fixed_cols.append("Name")
                elif any(x in v_lower for x in ["seat", "usn", "number"]):
                    fixed_cols.append("University Seat Number")
                else:
                    fixed_cols.append(val)
            else:
                # Normalized Component Name
                comp_clean = str(component).strip()
                comp_lower = comp_clean.lower()
                
                if comp_lower in ['ia', 'internal', 'cie', 'test', 'internal assessment', 'int']:
                    comp_clean = "Internal"
                elif comp_lower in ['ea', 'external', 'see', 'final', 'exam', 'sem end exam', 'ext']:
                    comp_clean = "External"
                elif comp_lower in ['tot', 'total', 'grand total']:
                    comp_clean = "Total"
                elif comp_lower in ['res', 'result', 'grade']:
                    comp_clean = "Result"
                
                # Subject Column: "Code Name Component" or "Code Component"
                if header_row_count == 3 and not is_empty(h2):
                     # Clean up name to avoid very long headers?
                     # For now, append it. Format: "Code - Name Component"
                     # We use " - " as separator to easily split later if needed
                     fixed_cols.append(f"{h1} - {h2} {comp_clean}")
                else:
                     fixed_cols.append(f"{h1} {comp_clean}")

        df_raw.columns = fixed_cols
        # Remove empty columns
        df = df_raw.loc[:, ~df_raw.columns.str.contains('^Unnamed')]
        df = df.loc[:, df.columns.str.strip() != ""]
        return df
    except Exception as e:
        print(f"Error processing excel: {e}")
        return pd.DataFrame()

def get_subject_codes(df):
    """Extracts unique subject codes using strict VTU format."""
    subject_codes = set()
    for col in df.columns:
        col = col.strip()

        # Handle "Code - Name Component" format
        if " - " in col:
            # We assume format: "Code - Name Component"
            parts = col.split(" - ", 1)
            code = parts[0].strip()
            # Verify code format
            if re.fullmatch(r"[A-Z]{2,}\d{3}[A-Z]?", code):
                # Verify component at the end
                if any(col.endswith(f" {s}") for s in ["Internal", "External", "Total", "Result"]):
                    subject_codes.add(code)
            continue
            
        if " " not in col:
            continue
            
        prefix, suffix = col.rsplit(" ", 1)
        if suffix in ["Internal", "External", "Total", "Result"]:
            if re.fullmatch(r"[A-Z]{2,}\d{3}[A-Z]?", prefix):
                subject_codes.add(prefix)
    return sorted(list(subject_codes))

def extract_numeric(roll):
    """Extracts the numeric part of a USN/Roll Number safely."""
    digits = re.findall(r'\d+', str(roll))
    return int(digits[-1]) if digits else 0

def assign_section(roll_no, section_ranges, usn_mapping=None):
    """Assigns sections based on either specific mapping or numeric roll number ranges."""
    roll_no_str = str(roll_no).strip().upper()
    
    # Check direct mapping first
    if usn_mapping:
         # Ensure usn_mapping keys are all upper/stripped just in case
         # (Though we do this at upload time, safe to cover bases)
         if roll_no_str in usn_mapping:
             return usn_mapping[roll_no_str]
    
    # Then check ranges if mapping not found
    roll_num = extract_numeric(roll_no)
    if section_ranges:
        for sec_name, (start, end) in section_ranges.items():
            start_num = extract_numeric(start)
            end_num = extract_numeric(end)
            if start_num <= roll_num <= end_num:
                return sec_name
    return "Unassigned"

def process_usn_mapping_file(contents, filename, section_name=None):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        if 'csv' in filename:
            df = pd.read_csv(io.StringIO(decoded.decode('utf-8')))
        else:
            df = pd.read_excel(io.BytesIO(decoded))
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        # Scenario 1: Section name provided (Single section upload)
        if section_name:
            # Look for USN column only
            usn_col = next((c for c in df.columns if 'usn' in c), None)
            if usn_col:
                 # Map all USNs in this file to the provided section_name
                 return {usn: section_name for usn in df[usn_col].astype(str).str.strip().str.upper()}
            return {}

        # Scenario 2: No section name (Global mapping file)
        # Find USN and Section columns
        usn_col = next((c for c in df.columns if 'usn' in c), None)
        section_col = next((c for c in df.columns if 'section' in c or 'sec' in c), None)
        
        if usn_col and section_col:
            # Create mapping: USN -> Section
            return dict(zip(df[usn_col].astype(str).str.strip().str.upper(), df[section_col].astype(str).str.strip()))
        return {}
    except Exception as e:
        print(f"Error processing USN file: {e}")
        return {}

# ---------- UI COMPONENTS ----------

def kpi_card(title, value, id_val, icon, color, bg_color):
    return html.Div(
        dbc.CardBody([
            html.Div([
                # Icon Box
                html.Div(
                    html.I(className=f"bi {icon}", style={"color": color, "fontSize": "1.4rem"}),
                    className="kpi-icon-box d-flex align-items-center justify-content-center",
                    style={
                        "minWidth": "44px", "width": "44px", "height": "44px", 
                        "borderRadius": "10px", "backgroundColor": bg_color
                    }
                ),
                # Text Content
                html.Div([
                    html.H6(title, className="text-muted text-uppercase fw-bold mb-0 text-truncate", style={"fontSize": "0.7rem", "letterSpacing": "0.5px", "maxWidth": "100px"}),
                    html.H3(children=value, id=f"{id_val}-text", className="kpi-val fw-bold mb-0", style={"color": color, "fontSize": "1.6rem"})
                ], className="ms-2")
            ], className="d-flex align-items-center h-100"),
            # Text Cue for Clickability
            html.Div("👆 Click for details", className="kpi-hover-hint text-muted text-end mt-1", style={"fontSize": "0.6rem", "opacity": "0.8", "position": "absolute", "bottom": "8px", "right": "12px", "display": "none", "transition": "opacity 0.2s ease"}),
        ], className="p-2 position-relative"),
        className="card kpi-card shadow-sm h-100 border-0 overflow-hidden overview-kpi-clickable rnk-card",
        style={"borderLeft": f"4px solid {color} !important", "transition": "transform 0.2s ease-in-out", "cursor": "pointer"},
        title=f"Click to view {title.lower()} students list",
        id={'type': 'overview-kpi-card', 'index': id_val},
        n_clicks=0
    )

# ---------- LAYOUT ----------

layout = dbc.Container([
    # Add Bootstrap Icons stylesheet
    html.Link(rel="stylesheet", href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css"),
    dcc.Markdown(f"<style>{PAGE_CSS_LIGHT}</style>", dangerously_allow_html=True),
    # Hero Header
    html.Div([
        html.H2("Student Performance Dashboard", className="fw-bold text-white mb-1"),
        html.P("Analyze university results with custom section filtering", className="text-white-50 mb-0"),
        dbc.Button("📖 Rules & Guidelines", id="open-legend-overview", color="light", size="sm", className="mt-3 fw-bold", outline=True)
    ], style={
        "background": "linear-gradient(135deg, #2c3e50 0%, #4ca1af 100%)", 
        "padding": "2.0rem 1rem", 
        "borderRadius": "0 0 15px 15px", 
        "textAlign": "center", 
        "marginBottom": "2rem"
    }),

    dbc.Row([
        # Left Sidebar: Inputs
        dbc.Col([
            dbc.Card([
                dbc.CardHeader([
                    html.Span("1. Data Intake", className="fw-bold"),
                    dbc.Button("View Sample", id="btn-sample-format", color="link", size="sm", className="float-end p-0 text-decoration-none")
                ], className="bg-light d-flex justify-content-between align-items-center"),
                dbc.CardBody([
                    dcc.Upload(
                        id='upload-data',
                        children=html.Div(["Drop Excel File or ", html.B("Click to Upload")]),
                        style={
                            'width': '100%', 'height': '70px', 'lineHeight': '70px', 
                            'borderWidth': '2px', 'borderStyle': 'dashed', 'borderRadius': '10px', 
                            'textAlign': 'center', 'backgroundColor': '#fbfcfc', 'cursor': 'pointer'
                        }
                    ),
                    html.Div(id='upload-status-feedback'),
                ], style={"overflow": "visible"}),
            ], className="mb-4 border-0 shadow-sm", style={"overflow": "visible"}),

            dbc.Card([
                dbc.CardHeader("2. Configuration", className="fw-bold bg-light", style={"overflow": "visible"}),
                dbc.CardBody([
                    
                    # ⚠️ LIVE MOVING ALERT
                    dbc.Alert(
                        html.Marquee("🚨 Action Required: Please select your Scheme and Semester and click 'Submit Mapping' to enable automatic SGPA calculation across the dashboard! 🚨",
                                     style={"fontWeight": "bold", "fontSize": "1.1rem"}
                        ),
                        id="scheme-moving-alert",
                        color="danger",
                        is_open=True,
                        className="p-1 mb-3 shadow-sm border-danger"
                    ),

                    html.Label("Scheme & Semester", className="small fw-bold mb-1"),
                    dbc.Row([
                        dbc.Col([
                            dcc.Dropdown(
                                id='scheme-selector',
                                options=[{'label': f'{yr} Scheme', 'value': str(yr)} for yr in [2022, 2021, 2018, 2025]],
                                value=None,
                                clearable=False,
                                className="mb-3",
                                placeholder="Select Scheme..."
                            )
                        ], width=6),
                        dbc.Col([
                            dcc.Dropdown(
                                id='semester-selector',
                                options=[{'label': f'Sem {i}', 'value': i} for i in range(1, 9)],
                                value=None,
                                clearable=False,
                                className="mb-3",
                                placeholder="Select Semester..."
                            )
                        ], width=6),
                    ]),

                    dbc.Button(
                        [html.I(className="bi bi-check-circle-fill me-2"), "Submit Mapping for SGPA"],
                        id="submit-scheme-btn",
                        color="success",
                        className="w-100 mb-3 fw-bold shadow-sm"
                    ),

                    html.Label("Filter Subjects", className="small fw-bold mb-1"),
                    html.Div([
                        dbc.ButtonGroup([
                            dbc.Button("Select All", id="subject-select-all-btn", color="primary", size="sm", outline=True, className="fw-bold"),
                            dbc.Button("Remove All", id="subject-remove-all-btn", color="danger", size="sm", outline=True, className="fw-bold"),
                        ], size="sm", className="mb-2 w-100"),
                    ]),
                    html.Div([
                        dcc.Dropdown(
                            id='subject-selector', 
                            multi=True, 
                            className="mb-3 custom-dropdown",
                            optionHeight=50,
                            maxHeight=300,
                            style={
                                "position": "relative", 
                                "zIndex": "1000",
                                "minHeight": "45px"
                            }
                        )
                    ], style={"overflow": "visible", "position": "relative", "zIndex": "1000"}),
                    html.Div(style={"height": "10px"}),
                    
                    html.Label("Section Config Mode", className="small fw-bold mb-2"),
                    dbc.RadioItems(
                        id="config-mode-selector",
                        options=[
                            {"label": "Manual Ranges", "value": "manual"},
                            {"label": "Upload Files", "value": "upload"},
                        ],
                        value="manual",
                        inline=True,
                        className="mb-3 small",
                        inputClassName="me-1",
                        labelClassName="me-3"
                    ),

                    # --- MANUAL MODE ---
                    html.Div([
                        html.Label("Define Ranges", className="small fw-bold mb-1"),
                        dbc.InputGroup([
                            dbc.Input(id='num-sections', type='number', value=1, min=1, max=10),
                            dbc.Button("Generate", id='generate-sections-btn', color="secondary"),
                        ], size="sm", className="mb-3"),
                        
                        html.Div(id='section-input-container'),
                        dbc.Button("Apply Ranges", id='submit-sections-btn', color='info', className='w-100 mt-2 fw-bold text-white'),
                    ], id="manual-section-container"),

                    # --- UPLOAD MODE ---
                    html.Div([
                        html.Div([
                            html.Label("Upload per Section", className="small fw-bold mb-1"),
                            dbc.Button("View Format", id="open-section-format", size="sm", color="link", className="text-decoration-none p-0 small")
                        ], className="d-flex justify-content-between align-items-center"),

                        dbc.InputGroup([
                            dbc.Input(id='num-upload-sections', type='number', value=1, min=1, max=10),
                            dbc.Button("Generate", id='generate-upload-sections-btn', color="secondary"),
                        ], size="sm", className="mb-3"),
                        
                        html.Div(id='upload-sections-container'),
                    ], id="upload-section-container", style={"display": "none"}),

                    html.Div(id='usn-upload-status', className="small text-muted mt-2 fw-bold"),

                    html.Hr(className="my-3"),
                    html.Div([
                        dbc.Button(
                            [
                                html.I(className="bi bi-download me-2", style={"fontSize": "1.1rem"}),
                                html.Span("Download Complete Report", style={"verticalAlign": "middle"}),
                            ],
                            id="universal-download-btn",
                            className="w-100 fw-bold download-report-btn",
                            size="lg",
                        ),
                        html.Div([
                            html.I(className="bi bi-file-earmark-excel me-1"),
                            html.Span("Overview · Ranking · Subject Analysis · Category Breakdown"),
                        ], className="text-center mt-2", style={"fontSize": "0.7rem", "color": "#6b7280", "letterSpacing": "0.02em"}),
                    ]),
                    dcc.Download(id="universal-download-excel"),
                ], style={"overflow": "visible", "position": "relative"}),
            ], className="border-0 shadow-sm", style={"overflow": "visible"})
        ], lg=4, md=5, style={"overflow": "visible"}),

        # Right Main: Analytics
        dbc.Col([
            # KPI Cards
            dbc.Row([
                dbc.Col(kpi_card("Total", "0", "total-students", "bi-people-fill", "#3b82f6", "#eff6ff"), className="d-flex"),
                dbc.Col(kpi_card("Appeared", "0", "present-students", "bi-person-circle", "#10b981", "#ecfdf5"), className="d-flex"),
                dbc.Col(kpi_card("Passed", "0", "passed-students", "bi-check-circle-fill", "#0ea5e9", "#f0f9ff"), className="d-flex"),
                dbc.Col(kpi_card("Failed", "0", "failed-students", "bi-x-circle-fill", "#ef4444", "#fef2f2"), className="d-flex"),
                dbc.Col(kpi_card("Absent", "0", "absent-students", "bi-person-x-fill", "#f59e0b", "#fffbeb"), className="d-flex"),
                dbc.Col(kpi_card("Pass %", "0%", "result-percent", "bi-percent", "#8b5cf6", "#f5f3ff"), className="d-flex"),
            ], className="row-cols-2 row-cols-md-3 row-cols-lg-6 g-2 mb-4"),

            # Table Card
            dbc.Card([
                dbc.CardHeader([
                    html.H6("Top 10 Data Preview", className="mb-0 fw-bold d-inline-block"),
                    html.Small(" (Filtered by Selection)", className="text-muted ms-2")
                ], className="bg-white py-3"),
                dbc.CardBody([
                    dcc.Loading(
                        id="loading-table", 
                        children=html.Div(id='data-preview'), 
                        type="default", 
                        color="#3498db"
                    )
                ], className="p-0")
            ], className="border-0 shadow-sm overflow-hidden")
        ], lg=8, md=7)
    ], style={"overflow": "visible"}),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("📊 Dashboard Rules & Guidelines")),
        dbc.ModalBody(
            html.Div([
                # --- Getting Started ---
                html.H5("🚀 Getting Started", className="text-primary fw-bold mb-2"),
                html.P("Welcome! This dashboard helps you analyze VTU semester results with section-wise filtering, subject-wise analytics, rankings, and more. Follow these steps to get started:", className="text-muted small mb-3"),

                # --- Step 1: Upload ---
                html.H6("📥 Step 1 — Upload Result File", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Accepted Format: "), ".xlsx or .xls (standard VTU result Excel sheet)."]),
                    html.Li([html.Strong("Header Structure: "), "The file must have multi-row headers — either 2-row (Subject Code → Component) or 3-row (Subject Code → Subject Name → Component). Both formats are auto-detected."]),
                    html.Li([html.Strong("Required Columns: "), "Must contain a 'USN' (or 'University Seat Number') column and a 'Name' column."]),
                    html.Li([html.Strong("Subject Columns: "), "Each subject should have 'Internal', 'External', 'Total', and 'Result' sub-columns."]),
                    html.Li([html.Strong("Tip: "), "Click 'View Sample' above the upload box to see the expected format."]),
                ], className="small"),
                html.Hr(),

                # --- Step 2: Scheme & Semester ---
                html.H6("📐 Step 2 — Select Scheme & Semester", className="fw-bold text-dark"),
                html.Ul([
                    html.Li("Select the correct Scheme Year (2018 / 2021 / 2022 / 2025) and Semester from the dropdowns."),
                    html.Li([html.Strong("Why? "), "This maps subject codes to their credit values, enabling automatic SGPA calculation across all pages."]),
                    html.Li("Click 'Submit Mapping for SGPA' to confirm. The red alert banner will disappear once configured."),
                    html.Li([html.Strong("Note: "), "If scheme/semester is not set, SGPA-related features will not be available."]),
                ], className="small"),
                html.Hr(),

                # --- Step 3: Subject Filter ---
                html.H6("🎯 Step 3 — Filter Subjects", className="fw-bold text-dark"),
                html.Ul([
                    html.Li("After upload, all detected subject codes appear in the 'Filter Subjects' dropdown."),
                    html.Li("By default, all subjects are selected. You can deselect subjects to exclude them from analysis."),
                    html.Li([html.Strong("Elective Handling: "), "If a student has not taken a particular subject (all marks blank), that subject is automatically skipped in their pass/fail calculation — it will NOT count as a fail or absent."]),
                ], className="small"),
                html.Hr(),

                # --- Step 4: Sections ---
                html.H6("⚙️ Step 4 — Configure Sections", className="fw-bold text-dark"),
                html.P("Map students to their classroom sections using one of two methods:", className="text-muted small mb-2"),
                html.Ul([
                    html.Li([html.Strong("Manual Ranges: "), "For sequential USNs. Enter section name + start/end USN (e.g., Section A: 001 to 060). Click 'Apply Ranges'."]),
                    html.Li([html.Strong("Upload Mapping Files: "), "For non-sequential USN lists. Upload a CSV/Excel file per section with a 'USN' column listing students in that section."]),
                    html.Li([html.Strong("Note: "), "Students not matching any section range or mapping will appear as 'Unassigned'."]),
                    html.Li([html.Strong("Tip: "), "If USNs from the mapping file are not found in the result data, a warning will list the missing USNs."]),
                ], className="small"),
                html.Hr(),

                # --- KPIs ---
                html.H6("📊 Understanding the Performance Cards", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Total: "), "Total number of students in the uploaded result file."]),
                    html.Li([html.Strong("Appeared: "), "Students who appeared for at least one selected subject (Total minus fully Absent)."]),
                    html.Li([html.Strong("Passed: "), "Students who passed ALL selected subjects they were enrolled in."]),
                    html.Li([html.Strong("Failed: "), "Students who failed or were absent in one or more selected subjects."]),
                    html.Li([html.Strong("Absent: "), "Students marked absent in ALL selected subjects (External = 0 and Result is blank/A)."]),
                    html.Li([html.Strong("Pass %: "), "Passed ÷ Appeared × 100 (fully absent students are excluded from this denominator)."]),
                ], className="small"),
                dbc.Alert([
                    html.I(className="bi bi-hand-index-thumb me-2"),
                    html.Strong("Clickable! "),
                    "Click any performance card to see the detailed student list for that category. You can also download the list as Excel."
                ], color="info", className="small py-2 mb-2"),
                html.Hr(),

                # --- Pass/Fail Logic ---
                html.H6("✅ Pass / Fail / Absent — Classification Rules", className="fw-bold text-dark"),
                html.P("For each selected subject, a student is classified per-subject first, then an overall result is computed:", className="text-muted small mb-2"),
                html.Ul([
                    html.Li([html.Strong("Pass (P): "), "Result column contains 'P' or a passing grade."]),
                    html.Li([html.Strong("Fail (F): "), "Result column contains 'F' or 'FAIL', OR total marks < 35 with no result grade."]),
                    html.Li([html.Strong("Absent (A): "), "External marks = 0 AND Result is blank, 'A', or 'ABSENT'."]),
                    html.Li([html.Strong("Skipped (Elective): "), "If Internal, External, Total are all blank AND Result is blank — the subject is completely ignored for that student."]),
                ], className="small"),
                html.P([
                    html.Strong("Overall Result: "),
                    "A student Passes overall only if they pass every enrolled subject. If even one subject is Failed or Absent, the overall result is Fail."
                ], className="small bg-light p-2 rounded border"),
                html.Hr(),

                # --- Dashboard Pages ---
                html.H6("🗂️ Dashboard Pages Overview", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Overview (this page): "), "Upload data, configure sections, view performance summary and data preview."]),
                    html.Li([html.Strong("Ranking: "), "View student rankings with SGPA, class rank, and performance tiers."]),
                    html.Li([html.Strong("Branch Analysis: "), "Analyze branch-level performance — upload a single branch result file for deep analytics."]),
                    html.Li([html.Strong("Subject Analysis: "), "Drill down into individual subject performance — pass rates, score distributions, and comparisons."]),
                    html.Li([html.Strong("Student Detail: "), "Look up individual student performance across all subjects with detailed breakdowns."]),
                    html.Li([html.Strong("Branch Intelligence: "), "Multi-branch comparative analysis and insights."]),
                ], className="small"),
                html.Hr(),

                # --- Important Notes ---
                html.H6("⚠️ Important Notes & Tips", className="fw-bold text-dark"),
                html.Ul([
                    html.Li("Data uploaded on the Overview page is shared across all other pages — no need to re-upload."),
                    html.Li("If the session expires (e.g., server restart), you will need to re-upload the file."),
                    html.Li("The dashboard auto-detects VTU subject codes in the format: 2+ letters followed by 3 digits and an optional letter (e.g., BCS501, BAIL504A)."),
                    html.Li("For best results, use the original unmodified VTU result Excel file — avoid renaming columns or restructuring the sheet."),
                    html.Li("Section configuration and scheme/semester selection are preserved as you navigate between pages."),
                ], className="small mb-0"),
            ])
        ),
        dbc.ModalFooter(dbc.Button("Got it!", id="close-legend-overview", className="ms-auto", color="primary"))
    ], id="legend-modal-overview", is_open=False, size="lg", style={"zIndex": 10500}),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("📅 Sample Excel Format")),
        dbc.ModalBody([
            html.P("Your uploaded Excel file must follow this structure:", className="text-muted small"),
            dbc.Table([
                html.Thead([
                    # Row 1: Subject Codes
                    html.Tr([
                        html.Th("University Seat Number", rowSpan=3, className="align-middle border-bottom"), 
                        html.Th("Name", rowSpan=3, className="align-middle border-bottom"), 
                        html.Th("BAIL504", colSpan=4, className="text-center border-start border-dark bg-light"), 
                        html.Th("BCS501", colSpan=4, className="text-center border-start border-dark bg-light")
                    ]),
                    # Row 2: Subject Names (Optional/New Format)
                    html.Tr([
                        html.Th("Artificial Intelligence Lab", colSpan=4, className="text-center border-start border-dark text-muted small"),
                        html.Th("Computer Networks", colSpan=4, className="text-center border-start border-dark text-muted small")
                    ]),
                    # Row 3: Components
                    html.Tr([
                        html.Th("IA / Internal", className="border-start border-dark"), html.Th("EA / External"), html.Th("Total"), html.Th("Result"),
                        html.Th("IA / Internal", className="border-start border-dark"), html.Th("EA / External"), html.Th("Total"), html.Th("Result")
                    ], className="small text-muted")
                ]),
                html.Tbody([
                    html.Tr([
                        html.Td("1XX23CSXXX"), html.Td("Bob"), 
                        html.Td("46", className="border-start border-dark"), html.Td("49"), html.Td("95"), html.Td("P", className="text-success fw-bold"),
                        html.Td("44", className="border-start border-dark"), html.Td("37"), html.Td("81"), html.Td("P", className="text-success fw-bold")
                    ]),
                    html.Tr([
                        html.Td("1XX23CSXXX"), html.Td("Alice"), 
                        html.Td("47", className="border-start border-dark"), html.Td("49"), html.Td("96"), html.Td("P", className="text-success fw-bold"),
                        html.Td("37", className="border-start border-dark"), html.Td("20"), html.Td("57"), html.Td("P", className="text-success fw-bold")
                    ]),
                ])
            ], bordered=True, responsive=True, className="mb-0")
        ]),
    ], id="modal-sample-format", size="lg", is_open=False),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("📅 Sample Section File Format")),
        dbc.ModalBody([
            html.P("For each section, upload a file containing a list of USNs belonging to that section.", className="text-muted small"),
            html.P("The file should have a column header named 'USN' or 'Student ID'.", className="fw-bold small"),
            dbc.Table([
                html.Thead(html.Tr(html.Th("USN"))),
                html.Tbody([
                    html.Tr(html.Td("1XX20CS001")),
                    html.Tr(html.Td("1XX20CS005")),
                    html.Tr(html.Td("1XX20CS012")),
                    html.Tr(html.Td("...")),
                ])
            ], bordered=True, striped=True, className="mb-0", style={"maxWidth": "200px"})
        ]),
    ], id="modal-section-format", size="sm", is_open=False),

    # --- Clickable KPI Popup Modal ---
    dbc.Modal([
        dbc.ModalHeader([
            dbc.ModalTitle(id="overview-kpi-modal-title", className="fw-bold text-primary"),
            html.Div([
                dbc.Button("Download List (Excel)", id="overview-kpi-modal-excel-top", color="success", outline=True, size="sm", className="me-2"),
                dbc.Button("Close", id="overview-kpi-modal-close-top", color="secondary", size="sm")
            ], className="ms-auto d-flex")
        ], close_button=False),
        dbc.ModalBody([
            dash_table.DataTable(
                id="overview-kpi-modal-table",
                columns=[], data=[],
                style_table={"overflowX": "auto", "borderRadius": "8px", "border": "1px solid #d1d5db"},
                style_cell={"textAlign": "center", "padding": "12px", "fontSize": "13px"},
                style_header={"backgroundColor": "#1f2937", "color": "#ffffff", "fontWeight": "700"},
                page_action='none',
                style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': '#f3f4f6'}],
            )
        ]),
        dbc.ModalFooter([
            dbc.Button("Download List (Excel)", id="overview-kpi-modal-excel", color="success", outline=True),
            dbc.Button("Close", id="overview-kpi-modal-close", className="ms-auto", color="secondary")
        ])
    ], id="overview-kpi-modal", is_open=False, size="xl", style={"zIndex": 10550}),
    
    dcc.Download(id="overview-kpi-excel-download"),

    # STORES REMOVED FROM HERE TO APP.PY TO ENSURE PERSISTENCE
    
], fluid=True, className="pb-5 bg-light", style={"minHeight": "100vh"})

# ---------- CALLBACKS ----------

@callback(
    Output("legend-modal-overview", "is_open"),
    [Input("open-legend-overview", "n_clicks"), Input("close-legend-overview", "n_clicks")],
    [State("legend-modal-overview", "is_open")],
    prevent_initial_call=True
)
def toggle_legend_overview(n1, n2, is_open): return not is_open if n1 or n2 else is_open

@callback(
    Output("modal-sample-format", "is_open"),
    [Input("btn-sample-format", "n_clicks")],
    [State("modal-sample-format", "is_open")],
    prevent_initial_call=True
)
def toggle_sample_format(n, is_open):
    return not is_open if n else is_open

@callback(
    Output("modal-section-format", "is_open"),
    [Input("open-section-format", "n_clicks")],
    [State("modal-section-format", "is_open")],
    prevent_initial_call=True
)
def toggle_section_format(n, is_open):
    return not is_open if n else is_open

@callback(
    [Output("manual-section-container", "style"),
     Output("upload-section-container", "style")],
    Input("config-mode-selector", "value")
)
def toggle_config_mode(mode):
    if mode == "upload":
        return {"display": "none"}, {"display": "block"}
    # Default manual
    return {"display": "block"}, {"display": "none"}

@callback(
    Output('subject-selector', 'options'),
    Output('subject-selector', 'value'),
    Output('stored-data', 'data'),
    Output('overview-selected-subjects', 'data'),
    Output('subject-options-store', 'data'),
    Output('usn-mapping-store', 'data', allow_duplicate=True),
    Output('section-data', 'data', allow_duplicate=True),
    Output('upload-status-feedback', 'children'),
    Input('upload-data', 'contents'),
    Input('upload-data', 'filename'),
    Input('subject-options-store', 'data'),
    Input('url', 'pathname'),
    State('overview-selected-subjects', 'data'),
    prevent_initial_call='initial_duplicate'
)
def manage_subjects(upload_contents, upload_filename, stored_options, pathname, stored_subjects):
    if pathname != "/" and pathname is not None:
        return no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update

    ctx_id = ctx.triggered_id

    # 1️⃣ If new file uploaded (Explicit User Action)
    if ctx_id == 'upload-data' and upload_contents:
        df = process_uploaded_excel(upload_contents)
        if df.empty:
            err_msg = dbc.Alert(
                [html.I(className="bi bi-exclamation-triangle-fill me-2"), "Upload failed — file is empty or invalid."],
                color="danger", className="mt-2 mb-0 py-2 px-3 small fw-bold", dismissable=True,
            )
            return [], [], None, None, None, None, None, err_msg

        subjects = get_subject_codes(df)
        options = [{'label': s, 'value': s} for s in subjects]
        
        # Save to Server Cache instead of JSON string
        session_id = str(uuid.uuid4())
        cache.set(session_id, df)
        
        # Build success feedback
        fname = upload_filename or "file"
        success_msg = dbc.Alert([
            html.Div([
                html.I(className="bi bi-check-circle-fill me-2", style={"fontSize": "1.1rem"}),
                html.Span("Data uploaded successfully!", style={"fontWeight": "700"}),
            ]),
            html.Div([
                html.Span(f"📄 {fname}", className="me-3"),
                html.Span(f"👥 {len(df)} students", className="me-3"),
                html.Span(f"📚 {len(subjects)} subjects"),
            ], className="mt-1", style={"fontSize": "0.8rem"}),
        ], color="success", className="mt-2 mb-0 py-2 px-3 small", dismissable=True)
        
        # Clear section and usn mappings because it is a new upload
        return options, subjects, session_id, subjects, options, {}, {}, success_msg

    # 2️⃣ If data already exists in session (Navigation / Restore)
    if stored_options:
        safe_subjects = stored_subjects if isinstance(stored_subjects, list) else []
        return stored_options, safe_subjects, no_update, no_update, no_update, no_update, no_update, no_update

    # 3️⃣ Default empty state
    return [], [], no_update, no_update, no_update, no_update, no_update, no_update

@callback(
    Output('scheme-selector', 'value'),
    Output('semester-selector', 'value'),
    Input('scheme-semester-store', 'data'),
    prevent_initial_call=False
)
def load_scheme_semester_ui(store_data):
    if store_data:
        return store_data.get('scheme', None), store_data.get('semester', None)
    return None, None

@callback(
    Output('scheme-semester-store', 'data', allow_duplicate=True),
    Output('scheme-moving-alert', 'is_open'),
    Input('submit-scheme-btn', 'n_clicks'),
    State('scheme-selector', 'value'),
    State('semester-selector', 'value'),
    State('scheme-semester-store', 'data'),
    prevent_initial_call=True
)
def update_scheme_semester_store(n_clicks, scheme, semester, store_data):
    if not n_clicks:
        # If no button click yet, keep alert open if nothing in store
        return no_update, True if not store_data else False
        
    if not scheme or not semester:
        # If they clicked but didn't select both, keep alert open and don't store
        return no_update, True

    # Valid submission: update store, hide alert
    return {"scheme": scheme, "semester": semester}, False

# Add a load callback to hide alert initially if already configured
@callback(
    Output('scheme-moving-alert', 'is_open', allow_duplicate=True),
    Input('scheme-semester-store', 'data'),
    prevent_initial_call='initial_duplicate'
)
def hide_alert_on_load(store_data):
    if store_data and store_data.get('scheme') and store_data.get('semester'):
        return False
    return True

@callback(
    Output('subject-selector', 'value', allow_duplicate=True),
    Input('subject-select-all-btn', 'n_clicks'),
    Input('subject-remove-all-btn', 'n_clicks'),
    State('subject-selector', 'options'),
    prevent_initial_call=True
)
def select_or_remove_all_subjects(select_clicks, remove_clicks, options):
    trigger = ctx.triggered_id
    if trigger == 'subject-select-all-btn' and options:
        return [o['value'] if isinstance(o, dict) else o for o in options]
    if trigger == 'subject-remove-all-btn':
        return []
    return no_update

@callback(
    Output('overview-selected-subjects', 'data', allow_duplicate=True),
    Input('subject-selector', 'value'),
    prevent_initial_call=True
)
def update_selected_subjects_store(selected_values):
    return selected_values

@callback(
    Output('section-input-container', 'children'),
    Input('generate-sections-btn', 'n_clicks'),
    Input('section-data', 'data'), # Listen to store changes or initial load
    State('num-sections', 'value'),
    prevent_initial_call=False
)
def render_section_fields(n_clicks, stored_sections, num_sections):
    ctx_id = ctx.triggered_id
    
    # 1. Button Click - Generate New Empty Fields
    if ctx_id == 'generate-sections-btn' and n_clicks:
        count = num_sections if num_sections else 1
        return [
            dbc.Row([
                dbc.Col(dbc.Input(id={'type': 'sec-n', 'index': i}, placeholder="Name", size="sm"), width=3),
                dbc.Col(dbc.Input(id={'type': 'sec-s', 'index': i}, placeholder="Start USN", size="sm"), width=4),
                dbc.Col(dbc.Input(id={'type': 'sec-e', 'index': i}, placeholder="End USN", size="sm"), width=4),
            ], className="g-2 mb-2") for i in range(1, count + 1)
        ]

    # 2. Restore from Store (Initial Load or Store Update)
    if stored_sections and isinstance(stored_sections, dict):
        rows = []
        for i, (name, (start, end)) in enumerate(stored_sections.items()):
            rows.append(dbc.Row([
                dbc.Col(dbc.Input(id={'type': 'sec-n', 'index': i+1}, value=name, placeholder="Name", size="sm"), width=3),
                dbc.Col(dbc.Input(id={'type': 'sec-s', 'index': i+1}, value=start, placeholder="Start USN", size="sm"), width=4),
                dbc.Col(dbc.Input(id={'type': 'sec-e', 'index': i+1}, value=end, placeholder="End USN", size="sm"), width=4),
            ], className="g-2 mb-2"))
        if rows:
            return rows

    # Default empty
    return []


@callback(
    Output('upload-sections-container', 'children'),
    Input('generate-upload-sections-btn', 'n_clicks'),
    State('num-upload-sections', 'value'),
    prevent_initial_call=True
)
def render_upload_section_fields(n, num):
    if not n or not num: return no_update
    return [
        dbc.Row([
            dbc.Col(dbc.Input(id={'type': 'usec-n', 'index': i}, placeholder=f"Sec {chr(65 + i)} Name", size="sm"), width=4),
            dbc.Col(
                dcc.Upload(
                    id={'type': 'usec-u', 'index': i},
                    children=html.Div(['📂 Upload USN List'], style={'fontSize': '0.8rem', 'fontWeight': 'bold'}),
                    style={
                        'width': '100%', 'height': '31px', 'lineHeight': '31px', 
                        'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '4px',
                        'textAlign': 'center', 'backgroundColor': '#f0f2f5', 'cursor': 'pointer',
                        'color': '#495057'
                    },
                    multiple=False
                ), width=8
            ),
        ], className="g-2 mb-2 align-items-center") for i in range(0, num)
    ]


@callback(
    Output('section-data', 'data'),
    Output('submit-sections-btn', 'children'), # Visual feedback on button
    Input('submit-sections-btn', 'n_clicks'),
    [State({'type': 'sec-n', 'index': ALL}, 'value'),
     State({'type': 'sec-s', 'index': ALL}, 'value'),
     State({'type': 'sec-e', 'index': ALL}, 'value'),
     State('section-data', 'data')],
    prevent_initial_call=True
)
def save_sections(n, names, starts, ends, current_data):
    if not n: return no_update, "Apply Config & Refresh"
    
    # Create new dict from inputs
    new_section_dict = {str(n).strip(): (str(s).strip(), str(e).strip()) for n, s, e in zip(names, starts, ends) if n and s and e}
    
    # If inputs are empty (user cleared them), we should probably clear the store too?
    # Or keep the old store? Given "Apply" button intent, if you see empty fields and click Apply, you expect clear.
    # However, since fields might not have rendered yet (if manage_subjects is slow?), we should be careful.
    # But save_sections is triggered by CLICK, so fields must exist.
    
    return new_section_dict, "✅ Config Applied"

@callback(
    Output('usn-mapping-store', 'data'),
    Output('usn-upload-status', 'children'),
    [Input({'type': 'usec-u', 'index': ALL}, 'contents')],
    [State({'type': 'usec-u', 'index': ALL}, 'filename'),
     State({'type': 'usec-n', 'index': ALL}, 'value'),
     State('usn-mapping-store', 'data')],
    prevent_initial_call=True
)
def process_multi_usn_upload(all_contents, all_filenames, all_names, current_mapping):
    # Ensure all_contents is a list, otherwise return
    if not isinstance(all_contents, list) or not any(all_contents):
        return no_update, ""
    
    # Initialize or copy existing mapping
    mapping = current_mapping.copy() if current_mapping else {}
    duplicates = []
    
    # Iterate through all upload components
    for i, content in enumerate(all_contents):
        if content: # If this specific upload has content
             name = all_names[i] if i < len(all_names) else None
             filename = all_filenames[i] if i < len(all_filenames) else ""
             
             # Determine section name (User input > Default A, B, C...)
             sec_name = name.strip() if name and name.strip() else f"Section {chr(65+i)}"
             
             # Process the file for this specific section mapping
             file_mapping = process_usn_mapping_file(content, filename, sec_name)
             
             if file_mapping:
                 # Check for conflicts
                 for usn, section in file_mapping.items():
                     if usn in mapping and mapping[usn] != section:
                         duplicates.append(f"{usn} (in {mapping[usn]} & {section})")
                 
                 mapping.update(file_mapping)
    
    total_entries = len(mapping)
    status_msg = f"✅ Total {total_entries} USNs mapped"
    
    if duplicates:
        count = len(duplicates)
        examples = ", ".join(duplicates[:2])
        status_msg = f"⚠️ {count} Duplicates found: {examples}..."
    
    if total_entries > 0:
        return mapping, status_msg
    
    return no_update, "ℹ️ No valid USNs found in uploaded files"

@callback(
    [Output('total-students-text', 'children'),
     Output('present-students-text', 'children'),
     Output('passed-students-text', 'children'),
     Output('failed-students-text', 'children'),
     Output('absent-students-text', 'children'),
     Output('result-percent-text', 'children'),
     Output('data-preview', 'children')],
    [Input('stored-data', 'data'),
     Input('subject-selector', 'value'),
     Input('section-data', 'data'),
     Input('usn-mapping-store', 'data')]
)
def update_dashboard(session_id, selected_subjects, section_ranges, usn_mapping):
    if not session_id or not selected_subjects:
        return "0", "0", "0", "0", "0", "0%", html.Div("Upload data and select subjects to view analytics.", className="p-4 text-center text-muted")
    
    # Retrieve from cache
    df = cache.get(session_id)
    if df is None:
        # Session expired or invalid
        return "0", "0", "0", "0", "0", "0%", html.Div("Session expired. Please re-upload data.", className="text-danger p-4 text-center")
    
    # df d pd.read_json(data, orient='split') <-- OLD
    # meta_col = df.columns[0]

    # Detect Meta Column (USN)
    meta_col = 'University Seat Number'
    if meta_col not in df.columns:
        # Fallback: check columns containing 'USN' or just take the first column
        usn_candidates = [c for c in df.columns if 'USN' in str(c).upper()]
        if usn_candidates:
            meta_col = usn_candidates[0]
        else:
            meta_col = df.columns[0]

    # 1. Filter relevant relevant
    all_subject_codes = get_subject_codes(df)
    
    # Start with just info columns
    info_cols = [c for c in df.columns if not any(s in c for s in all_subject_codes)]
    df_filtered = df[info_cols].copy()
    
    # Add relevant subject columns
    # New logic: columns must START with the code OR match "Code - Name" pattern
    # But simplified: just check if the code is present at the start followed by space or " - "
    subject_data_cols = []
    for c in df.columns:
        for s in selected_subjects:
            # Check for standard "CODE Component"
            if c.startswith(f"{s} "):
                 subject_data_cols.append(c)
                 break
            # Check for "CODE - Name Component"
            if c.startswith(f"{s} - "):
                 subject_data_cols.append(c)
                 break
    
    df_filtered = pd.concat([df_filtered, df[subject_data_cols]], axis=1)

    # 2. Convert Mark columns to Numeric (keep NaN for subjects student didn't take)
    for c in subject_data_cols:
        if any(k in c for k in ['Internal', 'External', 'Total']):
            df_filtered[c] = pd.to_numeric(df_filtered[c], errors='coerce')

    # 3. Robust Pass Logic (Matching Ranking Page Logic)
    res_cols = [c for c in subject_data_cols if "Result" in c]
    
    if res_cols:
        def calc_overall(row):
            subject_status = []
            for res_col in res_cols:
                base_name = res_col.rsplit(' Result', 1)[0].rsplit('Result', 1)[0].strip()
                
                i_col = f"{base_name} Internal"
                e_col = f"{base_name} External"
                t_col = f"{base_name} Total"

                i_raw = row.get(i_col) if i_col in df_filtered.columns else None
                e_raw = row.get(e_col) if e_col in df_filtered.columns else None
                t_raw = row.get(t_col) if t_col in df_filtered.columns else None
                r_raw = row.get(res_col, None)

                # Skip subject entirely if student has no data for it (elective not taken)
                i_na = pd.isna(i_raw)
                e_na = pd.isna(e_raw)
                t_na = pd.isna(t_raw)
                r_empty = pd.isna(r_raw) or str(r_raw).strip() == ''
                if i_na and e_na and t_na and r_empty:
                    continue

                try: i = float(i_raw) if not i_na else 0
                except: i = 0
                try: e = float(e_raw) if not e_na else 0
                except: e = 0
                
                r = str(r_raw).strip().upper() if not pd.isna(r_raw) else ''

                if (e == 0) and (r in ['A', 'ABSENT', '']):
                    subject_status.append('A')
                elif r in ['F', 'FAIL']:
                    subject_status.append('F')
                else:
                    total_s = i + e
                    if r == '' and total_s < 35:
                         subject_status.append('F')
                    else:
                         subject_status.append('P')

            if not subject_status: return 'P'
            elif subject_status.count('A') == len(subject_status): return 'A'
            elif subject_status.count('F') > 0 or subject_status.count('A') > 0: return 'F'
            return 'P'

        df_filtered['Overall_Result'] = df_filtered.apply(calc_overall, axis=1)
    else:
        df_filtered['Overall_Result'] = 'P'

    # 4. Metrics calculation
    total = len(df_filtered)
    
    passed_count = (df_filtered['Overall_Result'] == 'P').sum()
    absent_count = (df_filtered['Overall_Result'] == 'A').sum()
    failed_count = (df_filtered['Overall_Result'] == 'F').sum()
    
    # Present is Total - Absent (Absent means absent in ALL selected subjects)
    present_count = total - absent_count
    
    # Pass Percentage (Pass Rate) here:
    # Use Present count as denominator instead of Total
    passed_rate_val = (passed_count / present_count) * 100 if present_count > 0 else 0
    rate = f"{round(passed_rate_val, 2)}%"

    # 5. Section Assignment
    if section_ranges or usn_mapping:
        df_filtered['Section'] = df_filtered[meta_col].apply(lambda x: assign_section(x, section_ranges, usn_mapping))

    # 6. USN Validation (Check for Mismatched USNs)
    alert_msg = None
    if usn_mapping:
        result_usns = set(df_filtered[meta_col].astype(str).str.strip().str.upper())
        mapping_usns = set(k.strip().upper() for k in usn_mapping.keys())
        missing_usns = mapping_usns - result_usns
        
        if missing_usns:
            count = len(missing_usns)
            sorted_missing = sorted(list(missing_usns))
            
            # Format with Section for better details
            sorted_missing_display = [f"{u} ({usn_mapping.get(u, 'Unknown')})" for u in sorted_missing]
            
            if count <= 5:
                # Show all if few
                display_content = html.Div(f"Missing: {', '.join(sorted_missing_display)}", className="small mt-1")
            else:
                # Show summary + expander for many
                display_content = html.Div([
                    html.Div(f"Missing first 5: {', '.join(sorted_missing_display[:5])}...", className="small mt-1"),
                    html.Details([
                        html.Summary(f"Click to see all {count} missing USNs", style={"cursor": "pointer"}, className="small fw-bold text text mt-1"),
                        html.Div(
                            ", ".join(sorted_missing_display), 
                            className="small p-2 bg-light text-dark border rounded mt-1 text-break", 
                            style={"maxHeight": "150px", "overflowY": "auto"}
                        )
                    ])
                ])

            alert_msg = dbc.Alert(
                [
                    html.Div([
                        html.I(className="bi bi-exclamation-triangle-fill me-2"),
                        html.Strong(f"Warning: {count} USN(s) in Section Mapping NOT found in Result Data."),
                    ]),
                    display_content,
                    html.Div("These students will simply be ignored assigned to 'Unassigned'.", className="small text-muted mt-1")
                ],
                color="warning",
                className="mb-3 border-warning"
            )

    # 7. Generate Table UI with Complex Headers (Grouped Subjects)
    # Convert 'Code - Name Component' or 'Code Component' to multi-index
    cols_def = []
    
    # Identify non-subject columns first (Identity, Section, Overall Result)
    # We want them to span 2 rows in the header vertically
    # But dash_table handles this via empty strings in the first row r we want grouping
    
    preview_df = df_filtered.head(10)
    
    for c in preview_df.columns:
        # Check if it's a valid subject column
        is_subject_col = False
        col_header = ["" , c] # Default [Top, Bottom]

        # Is this a subject column?
        # Check if it ends with a known component
        # Need to iterate carefully to handle " - "
        for s in ["Internal", "External", "Total", "Result"]:
            # Check for EXACT ending match
            if c.endswith(f" {s}"):
                # Yes. Now extract the base name.
                # E.g. "BCS504 - DATA VIS LAB"
                base = c[:-len(s)].strip()
                col_header = [base, s]
                is_subject_col = True
                break
        
        # Identity columns should span nicely
        if not is_subject_col:
             # Use empty string for the second row to avoid duplicate text display
             # This usually creates a cleaner look, though vertical merging might not be perfect in all versions.
             # Alternatively, ["", c] puts the label at the bottom which lines up with components.
             col_header = ["", c]

        cols_def.append({"name": col_header, "id": c})

    # Use dash_table instead of dbc.Table for multi-header support
    # Ensure columns are sorted so that components (Int, Ext, Total, Result) appear next to each other
    # This is critical for the header merging to work visually
    
    # We rely on get_subject_codes logic which might give us subjects.
    # But preview_df columns might be in any order.
    # Let's enforce an order: Identity Cols -> (Sub1 Int, Sub1 Ext...) -> (Sub2 Int...) -> Result Cols
    
    # Simple check: If our processing keeps "Code - Name Int", "Code - Name Ext" adjacent, merging works.
    # process_uploaded_excel seems to append them sequentially, so order should be preserved per subject.
    
    table = dash_table.DataTable(
        data=preview_df.to_dict('records'),
        columns=cols_def,
        style_table={'overflowX': 'auto', 'minWidth': '100%'},
        style_header={
            'backgroundColor': '#f8fafc',
            'fontWeight': 'bold',
            'border': '1px solid #e2e8f0',
            'ground': 'center',
            'whiteSpace': 'normal',
            'height': 'auto',
        },
        style_cell={
            'textAlign': 'center',
            'padding': '12px',
            'fontSize': '0.85rem',
            'fontFamily': 'var(--bs-body-font-family)',
            'border': '1px solid #f1f5f9'
        },
        style_data_conditional=[
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': '#f9fafb'
            },
            # Style for different components
            # Cannot a lambda in style_data_conditional
            # We must use filter_query or specify column_id
            # Since we have dynamic columns, we can add this rule per-column or omit it for now
            # Alternative: Add 'Result' to cell style based on column name MATCH logic?
            # Dash doesn't support regex on column_id in styles easily without loop.
        ],
        merge_duplicate_headers=True, # This enables the grouping!
        page_size=10
    )
    
    # Dynamic styling for Result columns
    for col in preview_df.columns:
        if str(col).endswith(' Result'):
            table.style_data_conditional.append({
                'if': {'column_id': col},
                'fontWeight': 'bold',
                'color': '#2563eb'
            })
    
    final_output = html.Div([alert_msg, table]) if alert_msg else table
    
    final_output = html.Div([alert_msg, table]) if alert_msg else table

    return str(total), str(present_count), str(passed_count), str(failed_count), str(absent_count), rate, final_output

# ==================== KPI Modal & Export Logic ====================

@callback(
    Output("overview-kpi-modal", "is_open"),
    Output("overview-kpi-modal-title", "children"),
    Output("overview-kpi-modal-table", "data"),
    Output("overview-kpi-modal-table", "columns"),
    Input({"type": "overview-kpi-card", "index": ALL}, "n_clicks"),
    Input("overview-kpi-modal-close", "n_clicks"),
    Input("overview-kpi-modal-close-top", "n_clicks"),
    State('stored-data', 'data'),
    State('subject-selector', 'value'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data'),
    prevent_initial_call=True
)
def handle_overview_kpi_click(kpi_clicks, c1, c2, session_id, selected_subjects, section_ranges, usn_mapping):
    if not dash.ctx.triggered: 
        raise PreventUpdate

    trigger = dash.ctx.triggered_id

    # Handle Modal Close
    if trigger in ["overview-kpi-modal-close", "overview-kpi-modal-close-top"]:
        return False, dash.no_update, dash.no_update, dash.no_update

    # Handle Dictionary ID matching
    if not isinstance(trigger, dict) or trigger.get("type") != "overview-kpi-card":
        raise PreventUpdate
        
    if all(c == 0 or c is None for c in kpi_clicks): 
        raise PreventUpdate
        
    kpi_type = trigger.get("index")

    if not session_id or not selected_subjects:
        return True, "No Data Available", [], []

    df = cache.get(session_id)
    if df is None:
        return True, "Session Expired", [], []

    # Detect Meta Column (USN)
    meta_col = 'University Seat Number'
    if meta_col not in df.columns:
        usn_candidates = [c for c in df.columns if 'USN' in str(c).upper()]
        meta_col = usn_candidates[0] if usn_candidates else df.columns[0]

    # Replication of overview passing logic
    all_subject_codes = get_subject_codes(df)
    info_cols = [c for c in df.columns if not any(s in c for s in all_subject_codes)]
    df_filtered = df[info_cols].copy()
    
    subject_data_cols = []
    for c in df.columns:
        for s in selected_subjects:
            if c.startswith(f"{s} ") or c.startswith(f"{s} - "):
                 subject_data_cols.append(c)
                 break
                 
    df_filtered = pd.concat([df_filtered, df[subject_data_cols]], axis=1)

    for c in subject_data_cols:
        if any(k in c for k in ['Internal', 'External', 'Total']):
            df_filtered[c] = pd.to_numeric(df_filtered[c], errors='coerce')

    res_cols = [c for c in subject_data_cols if "Result" in c]
    
    if res_cols:
        def calc_overall(row):
            subject_status = []
            for res_col in res_cols:
                base_name = res_col.rsplit(' Result', 1)[0].rsplit('Result', 1)[0].strip()
                i_col = f"{base_name} Internal"
                e_col = f"{base_name} External"
                t_col = f"{base_name} Total"

                i_raw = row.get(i_col) if i_col in df_filtered.columns else None
                e_raw = row.get(e_col) if e_col in df_filtered.columns else None
                t_raw = row.get(t_col) if t_col in df_filtered.columns else None
                r_raw = row.get(res_col, None)

                # Skip subject entirely if student has no data for it
                i_na = pd.isna(i_raw)
                e_na = pd.isna(e_raw)
                t_na = pd.isna(t_raw)
                r_empty = pd.isna(r_raw) or str(r_raw).strip() == ''
                if i_na and e_na and t_na and r_empty:
                    continue

                try: i = float(i_raw) if not i_na else 0
                except: i = 0
                try: e = float(e_raw) if not e_na else 0
                except: e = 0
                r = str(r_raw).strip().upper() if not pd.isna(r_raw) else ''

                if (e == 0) and (r in ['A', 'ABSENT', '']): subject_status.append('A')
                elif r in ['F', 'FAIL']: subject_status.append('F')
                else: subject_status.append('F' if r == '' and (i + e) < 35 else 'P')

            if not subject_status: return 'P'
            elif subject_status.count('A') == len(subject_status): return 'A'
            elif subject_status.count('F') > 0 or subject_status.count('A') > 0: return 'F'
            return 'P'

        df_filtered['Overall_Result'] = df_filtered.apply(calc_overall, axis=1)
    else:
        df_filtered['Overall_Result'] = 'P'

    # Assign Section
    if section_ranges or usn_mapping:
        df_filtered['Section'] = df_filtered[meta_col].apply(lambda x: assign_section(x, section_ranges, usn_mapping))
    else:
        df_filtered['Section'] = "Unassigned"

    # Filter Logic based on clicked card
    if kpi_type == 'total-students': res_df = df_filtered
    elif kpi_type == 'present-students': res_df = df_filtered[df_filtered['Overall_Result'] != 'A']
    elif kpi_type == 'absent-students': res_df = df_filtered[df_filtered['Overall_Result'] == 'A']
    elif kpi_type in ['passed-students', 'result-percent']: res_df = df_filtered[df_filtered['Overall_Result'] == 'P']
    elif kpi_type == 'failed-students': res_df = df_filtered[df_filtered['Overall_Result'] == 'F']
    else: res_df = df_filtered

    if res_df.empty:
        return True, f"Student List: {str(kpi_type).upper()} (0 Students)", [], []

    # Display specific clean columns
    possible_name_cols = [c for c in res_df.columns if 'name' in c.lower()]
    name_col = possible_name_cols[0] if possible_name_cols else None
    
    display_cols = []
    if meta_col in res_df.columns: display_cols.append(meta_col)
    if name_col and name_col in res_df.columns: display_cols.append(name_col)
    display_cols.extend(['Section', 'Overall_Result'])

    tcols = [{"name": c.replace("_", " "), "id": c} for c in display_cols]
    tdata = res_df[display_cols].to_dict('records')
    title = f"📃 Detail List: {str(kpi_type).replace('-', ' ').title()} ({len(res_df)} Students)"

    return True, title, tdata, tcols

@callback(
    Output("overview-kpi-excel-download", "data"),
    Input("overview-kpi-modal-excel", "n_clicks"),
    Input("overview-kpi-modal-excel-top", "n_clicks"),
    State("overview-kpi-modal-table", "data"),
    State("overview-kpi-modal-title", "children"),
    prevent_initial_call=True
)
def download_overview_modal_excel(n_clicks_bottom, n_clicks_top, table_data, title):
    if not dash.ctx.triggered or not table_data:
        raise PreventUpdate
        
    df = pd.DataFrame(table_data)
    
    safe_title = "Overview_Student_List"
    if title and isinstance(title, str):
        clean_str = title.split('(')[0].replace('📃', '').replace('Detail List:', '').strip()
        safe_title = re.sub(r'[^A-Za-z0-9_]', '_', clean_str)
        
    return dcc.send_data_frame(df.to_excel, f"{safe_title}_Report.xlsx", index=False)


# ==================== UNIVERSAL DOWNLOAD ====================

def _get_selected_subject_cols(df, selected_subjects):
    """Get columns belonging to selected subjects only."""
    cols = []
    for c in df.columns:
        for s in selected_subjects:
            if c.startswith(f"{s} ") or c.startswith(f"{s} - "):
                cols.append(c)
                break
    return cols


def _calc_overall_result(row, res_cols, df_ref):
    """Shared pass/fail/absent logic matching the dashboard exactly."""
    subject_status = []
    for res_col in res_cols:
        base_name = res_col.rsplit(' Result', 1)[0].rsplit('Result', 1)[0].strip()
        e_col = f"{base_name} External"
        i_col = f"{base_name} Internal"
        t_col = f"{base_name} Total"
        e_raw = row.get(e_col) if e_col in df_ref.columns else None
        i_raw = row.get(i_col) if i_col in df_ref.columns else None
        t_raw = row.get(t_col) if t_col in df_ref.columns else None
        r_raw = row.get(res_col, None)
        # Skip electives not taken
        if all(pd.isna(x) or str(x).strip() == '' for x in [i_raw, e_raw, t_raw, r_raw]):
            continue
        try: e = float(e_raw) if not pd.isna(e_raw) else 0
        except: e = 0
        try: i = float(i_raw) if not pd.isna(i_raw) else 0
        except: i = 0
        r = str(r_raw).strip().upper() if not pd.isna(r_raw) else ''
        if (e == 0) and (r in ['A', 'ABSENT', '']):
            subject_status.append('A')
        elif r in ['F', 'FAIL']:
            subject_status.append('F')
        else:
            if r == '' and (i + e) < 35:
                subject_status.append('F')
            else:
                subject_status.append('P')
    if not subject_status: return 'P'
    elif subject_status.count('A') == len(subject_status): return 'A'
    elif subject_status.count('F') > 0 or subject_status.count('A') > 0: return 'F'
    return 'P'


def _build_overview_sheet(df, selected_subjects, section_ranges, usn_mapping):
    """Build the Overview sheet data."""
    meta_col = 'University Seat Number'
    if meta_col not in df.columns:
        usn_candidates = [c for c in df.columns if 'USN' in str(c).upper()]
        meta_col = usn_candidates[0] if usn_candidates else df.columns[0]

    all_subject_codes = get_subject_codes(df)
    info_cols = [c for c in df.columns if not any(s in c for s in all_subject_codes)]
    df_filtered = df[info_cols].copy()

    subject_data_cols = _get_selected_subject_cols(df, selected_subjects)
    df_filtered = pd.concat([df_filtered, df[subject_data_cols]], axis=1)

    for c in subject_data_cols:
        if any(k in c for k in ['Internal', 'External', 'Total']):
            df_filtered[c] = pd.to_numeric(df_filtered[c], errors='coerce')

    res_cols = [c for c in subject_data_cols if 'Result' in c]
    if res_cols:
        df_filtered['Overall_Result'] = df_filtered.apply(
            lambda row: _calc_overall_result(row, res_cols, df_filtered), axis=1)
    else:
        df_filtered['Overall_Result'] = 'P'

    if section_ranges or usn_mapping:
        df_filtered['Section'] = df_filtered[meta_col].apply(lambda x: assign_section(x, section_ranges, usn_mapping))

    return df_filtered


def _build_ranking_sheet(df, selected_subjects, section_ranges, usn_mapping):
    """Build Marks-based Ranking sheet — uses the SAME logic as ranking.py dashboard."""
    from pages.ranking import _normalize_df, calculate_student_metrics

    df = _normalize_df(df.copy(), section_ranges, usn_mapping)
    df = calculate_student_metrics(df)

    # Rank only passing students (same as build_views in ranking.py)
    pass_mask = df['Overall_Result'] == 'P'
    df['Class_Rank'] = pd.NA
    df.loc[pass_mask, 'Class_Rank'] = df.loc[pass_mask, 'Total_Marks'].rank(method='min', ascending=False).astype('Int64')

    display_cols = ['Student_ID']
    if 'Name' in df.columns:
        display_cols.append('Name')
    display_cols += ['Section', 'Total_Marks', 'percentage', 'Overall_Result', 'Class_Rank']
    display_cols = [c for c in display_cols if c in df.columns]
    out = df[display_cols].sort_values('Class_Rank', na_position='last')
    out = out.rename(columns={'percentage': 'Percentage'})
    return out


def _build_subject_analysis_sheet(df, selected_subjects, section_ranges, usn_mapping):
    """Build per-subject pass/fail/absent stats."""
    first_col = df.columns[0]
    rows = []
    for subj in selected_subjects:
        subj_cols = [c for c in df.columns if c.startswith(f"{subj} ") or c.startswith(f"{subj} - ")]
        res_col = next((c for c in subj_cols if 'Result' in c), None)
        ext_col = next((c for c in subj_cols if 'External' in c), None)
        if not res_col:
            continue
        valid = df[df[res_col].notna() & (df[res_col].astype(str).str.strip() != '')].copy()
        if valid.empty:
            rows.append({'Subject': subj, 'Total': 0, 'Appeared': 0, 'Absent': 0, 'Passed': 0, 'Failed': 0, 'Pass %': 0})
            continue
        valid[res_col] = valid[res_col].astype(str).str.strip().str.upper()
        if ext_col:
            valid[ext_col] = pd.to_numeric(valid[ext_col], errors='coerce').fillna(0)
        def status(row):
            r = row[res_col]
            e = row[ext_col] if ext_col else 0
            try: e = float(e)
            except: e = 0
            if r in ['A', 'ABSENT'] and e == 0: return 'Absent'
            elif r in ['F', 'FAIL']: return 'Fail'
            elif r in ['P', 'PASS']: return 'Pass'
            return 'Ignore'
        valid['_status'] = valid.apply(status, axis=1)
        valid = valid[valid['_status'] != 'Ignore']
        total = len(valid)
        absent = (valid['_status'] == 'Absent').sum()
        appeared = total - absent
        passed = (valid['_status'] == 'Pass').sum()
        failed = (valid['_status'] == 'Fail').sum()
        pass_pct = round((passed / appeared) * 100, 2) if appeared > 0 else 0
        rows.append({'Subject': subj, 'Total': total, 'Appeared': appeared, 'Absent': absent, 'Passed': passed, 'Failed': failed, 'Pass %': pass_pct})
    return pd.DataFrame(rows)


def _build_category_sheet(df, selected_subjects):
    """Build FCD/FC/SC/Pass Class category breakdown."""
    first_col = df.columns[0]
    all_subj_total_cols = []
    for subj in selected_subjects:
        subj_cols = [c for c in df.columns if (c.startswith(f"{subj} ") or c.startswith(f"{subj} - ")) and c.strip().endswith(' Total')]
        all_subj_total_cols.extend(subj_cols)

    res_cols = []
    for subj in selected_subjects:
        rc = [c for c in df.columns if (c.startswith(f"{subj} ") or c.startswith(f"{subj} - ")) and 'Result' in c]
        res_cols.extend(rc)

    if not res_cols:
        return pd.DataFrame()

    df_cat = df.copy()
    df_cat['_Result'] = df_cat.apply(
        lambda row: _calc_overall_result(row, res_cols, df_cat), axis=1)
    pass_df = df_cat[df_cat['_Result'] == 'P'].copy()

    if pass_df.empty or not all_subj_total_cols:
        return pd.DataFrame(columns=[first_col, 'Name', 'Total_Marks', 'Percentage', 'Category'])

    for c in all_subj_total_cols:
        pass_df[c] = pd.to_numeric(pass_df[c], errors='coerce').fillna(0)
    pass_df['_Total'] = pass_df[all_subj_total_cols].sum(axis=1)
    # Detect actual max marks per subject: ceil(column_max / 100) * 100
    # Handles 200-mark subjects (e.g. projects) correctly
    import numpy as _np
    _per_subj_max = _np.ceil(pass_df[all_subj_total_cols].max().clip(lower=1) / 100) * 100
    # Per-student: only count max marks for subjects the student actually attempted
    _attempted = pass_df[all_subj_total_cols] > 0
    _student_total_max = (_attempted * _per_subj_max.values).sum(axis=1).clip(lower=1)
    pass_df['_Pct'] = ((pass_df['_Total'] / _student_total_max) * 100).round(2)

    def category(pct):
        if pct >= 70: return 'FCD (First Class Distinction)'
        elif pct >= 60: return 'First Class'
        elif pct >= 50: return 'Second Class'
        return 'Pass Class'

    pass_df['Category'] = pass_df['_Pct'].apply(category)
    out_cols = [first_col]
    if 'Name' in pass_df.columns:
        out_cols.append('Name')
    out_cols += ['_Total', '_Pct', 'Category']
    result = pass_df[out_cols].rename(columns={'_Total': 'Total_Marks', '_Pct': 'Percentage'})
    return result.sort_values('Percentage', ascending=False)


def _get_grade_point(score):
    """VTU grade-point mapping (same as ranking.py)."""
    s = pd.to_numeric(score, errors='coerce')
    if pd.isna(s): return 0
    s = float(s)
    if 90 <= s <= 100: return 10
    elif 80 <= s < 90: return 9
    elif 70 <= s < 80: return 8
    elif 60 <= s < 70: return 7
    elif 55 <= s < 60: return 6
    elif 50 <= s < 55: return 5
    elif 40 <= s < 50: return 4
    return 0


def _auto_compute_sgpa(df, selected_subjects, section_ranges, usn_mapping, scheme_sem_data):
    """Auto-compute SGPA using ranking.py's exact same logic — reuses calculate_sgpa_all pipeline."""
    from services.credit_service import extract_course_number, load_credit_map
    from pages.ranking import _normalize_df, get_grade_point as rnk_grade_point
    import json as _json
    import os as _os

    scheme = None
    semester = None
    if scheme_sem_data:
        scheme = scheme_sem_data.get('scheme')
        semester = scheme_sem_data.get('semester')

    # Auto-detect semester from subject codes if not provided
    if not semester and selected_subjects:
        for code in selected_subjects:
            num = extract_course_number(code)
            if num and len(num) == 3:
                semester = int(num[0])
                break
    if not semester:
        semester = 5

    if not scheme:
        scheme = '2022'

    print(f"[SGPA] Auto-detected scheme={scheme}, semester={semester}")

    credit_path = _os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        'utils', 'credit_database', f'{scheme}_scheme', f'sem{semester}.json')
    if not _os.path.exists(credit_path):
        print(f"[SGPA] Credit file not found!")
        return None, False
    with open(credit_path, 'r') as _f:
        credit_map = _json.load(_f)

    # Build credit dict for selected subjects
    credit_dict = {}
    for code in selected_subjects:
        num = extract_course_number(code)
        cr = credit_map.get(num, 0) if num else 0
        if cr > 0:
            credit_dict[code] = cr

    if not credit_dict:
        return None, False

    # Use ranking.py's _normalize_df to get the exact same base data as the dashboard
    base = _normalize_df(df.copy(), section_ranges, usn_mapping)

    # Resolve subject codes to actual column prefixes (e.g., "BPEK359" -> "BPEK359 - PHYSICAL EDUCATION")
    # This matches exactly how ranking.py's generate_credit_panel extracts codes from columns.
    col_prefixes = set()
    for col in base.columns:
        m = re.match(r'^(.*?)\s+(Internal|External|Total|Result)$', col, flags=re.IGNORECASE)
        if m:
            col_prefixes.add(m.group(1).strip())

    # Map each selected subject code to its full column prefix
    resolved_credits = {}
    for code, credit in credit_dict.items():
        # Try exact match first (e.g., "BPEK359 Internal" exists)
        if f"{code} Internal" in base.columns or f"{code} Total" in base.columns:
            resolved_credits[code] = credit
        else:
            # Look for "CODE - NAME" prefix pattern
            for prefix in col_prefixes:
                if prefix.startswith(f"{code} - ") or prefix == code:
                    resolved_credits[prefix] = credit
                    break

    if not resolved_credits:
        return None, False

    # Pre-compute max marks per subject from data (handles 200-mark subjects)
    import numpy as np
    _subj_max_marks = {}
    for code in resolved_credits:
        total_col = f"{code} Total"
        if total_col in base.columns:
            col_max = pd.to_numeric(base[total_col], errors='coerce').max()
            _subj_max_marks[code] = int(np.ceil(max(col_max, 1) / 100) * 100) if pd.notna(col_max) else 100
        else:
            _subj_max_marks[code] = 100

    # Now compute SGPA using ranking.py's exact same per-student loop (from calculate_sgpa_all)
    sgpa_rows = []
    for _, row in base.iterrows():
        total_cp, total_cre, total_marks, fail_flag = 0, 0, 0, False
        for code, credit in resolved_credits.items():
            i = pd.to_numeric(row.get(f"{code} Internal"), errors='coerce') or 0
            e = pd.to_numeric(row.get(f"{code} External"), errors='coerce') or 0

            if f"{code} Total" in base.columns:
                score = pd.to_numeric(row.get(f"{code} Total"), errors='coerce') or 0
            else:
                score = (i + e) if (i and e) else (i or e or 0)

            res_val = str(row.get(f"{code} Result", "")).strip().upper()

            # Skip subjects this student did NOT take
            has_marks = (i > 0) or (e > 0) or (score > 0)
            has_result = res_val in ('P', 'F')
            if not has_marks and not has_result:
                continue

            if res_val == 'P':
                pass
            elif res_val == 'F':
                fail_flag = True
            elif res_val == 'A':
                fail_flag = True
            else:
                if score < 35:
                    fail_flag = True

            max_m = _subj_max_marks.get(code, 100)
            pct = (score / max_m * 100) if max_m > 0 else 0
            total_cp += rnk_grade_point(pct) * credit
            total_cre += credit
            total_marks += score

        sgpa = (total_cp / total_cre) if total_cre > 0 else 0.0

        ovr = str(row.get('Overall_Result', '')).strip().upper()
        if ovr in ['A', 'ABSENT']:
            res = 'Absent'
        elif ovr in ['F', 'FAIL']:
            res = 'Fail'
        elif ovr in ['P', 'PASS']:
            res = 'Pass'
        else:
            res = 'Pass' if (not fail_flag and total_cre > 0) else 'Fail'

        if fail_flag and res != 'Pass' and res != 'Absent':
            res = 'Fail'
        if ovr in ['F', 'FAIL']:
            res = 'Fail'

        sgpa_rows.append({
            'Student_ID': row['Student_ID'],
            'Name': row.get('Name', ''),
            'Section': row.get('Section', ''),
            'SGPA': round(sgpa, 2),
            'Total_Marks': round(total_marks, 2),
            'Overall_Result': res,
        })

    sgpa_df = pd.DataFrame(sgpa_rows)
    pass_mask = sgpa_df['Overall_Result'] == 'Pass'
    sgpa_df['SGPA_Rank'] = pd.NA
    sgpa_df.loc[pass_mask, 'SGPA_Rank'] = sgpa_df.loc[pass_mask, 'SGPA'].rank(
        method='min', ascending=False).astype('Int64')
    return sgpa_df.sort_values('SGPA_Rank', na_position='last'), True


def _build_sgpa_sheet(sgpa_json_str, df=None, selected_subjects=None,
                      section_ranges=None, usn_mapping=None, scheme_sem_data=None):
    """Build SGPA-based ranking sheet. Uses stored data if available, otherwise auto-computes."""
    # Try stored SGPA data first
    if sgpa_json_str:
        try:
            from io import StringIO
            sgpa_df = pd.read_json(StringIO(sgpa_json_str), orient='split')
            if not sgpa_df.empty and 'SGPA' in sgpa_df.columns:
                # Merge Name from base data if not present in sgpa-store
                if 'Name' not in sgpa_df.columns and df is not None:
                    meta_col = df.columns[0]
                    name_map = df.set_index(meta_col)['Name'].to_dict() if 'Name' in df.columns else {}
                    sgpa_df['Name'] = sgpa_df['Student_ID'].map(name_map).fillna('')
                display_cols = ['Student_ID']
                if 'Name' in sgpa_df.columns:
                    display_cols.append('Name')
                if 'Section' in sgpa_df.columns:
                    display_cols.append('Section')
                display_cols += ['SGPA', 'Total_Marks_Selected', 'Result_Selected', 'SGPA_Class_Rank']
                display_cols = [c for c in display_cols if c in sgpa_df.columns]
                out = sgpa_df[display_cols].copy()
                out = out.rename(columns={
                    'Total_Marks_Selected': 'Total_Marks',
                    'Result_Selected': 'Overall_Result',
                    'SGPA_Class_Rank': 'SGPA_Rank',
                })
                return out.sort_values('SGPA_Rank', na_position='last'), True
        except Exception:
            pass

    # Auto-compute SGPA if we have the data
    if df is not None and selected_subjects:
        return _auto_compute_sgpa(df.copy(), selected_subjects, section_ranges, usn_mapping, scheme_sem_data)

    return None, False


@callback(
    Output('universal-download-excel', 'data'),
    Input('universal-download-btn', 'n_clicks'),
    State('stored-data', 'data'),
    State('subject-selector', 'value'),
    State('section-data', 'data'),
    State('usn-mapping-store', 'data'),
    State('sgpa-store', 'data'),
    State('scheme-semester-store', 'data'),
    prevent_initial_call=True
)
def universal_download(n_clicks, session_id, selected_subjects, section_ranges, usn_mapping, sgpa_json, scheme_sem_data):
    print(f"[DOWNLOAD] ENTERED. n_clicks={n_clicks}, session_id={session_id}, subjects_len={len(selected_subjects) if selected_subjects else 0}")
    if not n_clicks or not session_id or not selected_subjects:
        print(f"[DOWNLOAD] PreventUpdate: n_clicks={bool(n_clicks)}, session_id={bool(session_id)}, subjects={bool(selected_subjects)}")
        raise PreventUpdate

    print(f"[DOWNLOAD] Triggered. session_id={session_id}, subjects={selected_subjects}")

    df = cache.get(session_id)
    if df is None:
        print("[DOWNLOAD] Cache miss - session expired")
        raise PreventUpdate

    from io import BytesIO
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    try:
        # Each builder creates its own internal copies — no need to copy here
        overview_df = _build_overview_sheet(df, selected_subjects, section_ranges, usn_mapping)
        print(f"[DOWNLOAD] Overview: {len(overview_df)} rows")
        ranking_df = _build_ranking_sheet(df, selected_subjects, section_ranges, usn_mapping)
        print(f"[DOWNLOAD] Ranking: {len(ranking_df)} rows")
        subject_df = _build_subject_analysis_sheet(df, selected_subjects, section_ranges, usn_mapping)
        print(f"[DOWNLOAD] Subject Analysis: {len(subject_df)} rows")
        category_df = _build_category_sheet(df, selected_subjects)
        print(f"[DOWNLOAD] Category: {len(category_df)} rows")
        sgpa_df, sgpa_computed = _build_sgpa_sheet(
            sgpa_json, df, selected_subjects,
            section_ranges, usn_mapping, scheme_sem_data)
        print(f"[DOWNLOAD] SGPA computed: {sgpa_computed}")
    except Exception as e:
        print(f"[DOWNLOAD ERROR] Building sheets failed: {e}")
        import traceback; traceback.print_exc()
        raise PreventUpdate

    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        # ── Compute KPI metrics from overview_df ──
        _kpi_total = len(overview_df)
        _kpi_absent = int((overview_df['Overall_Result'] == 'A').sum()) if 'Overall_Result' in overview_df.columns else 0
        _kpi_appeared = _kpi_total - _kpi_absent
        _kpi_passed = int((overview_df['Overall_Result'] == 'P').sum()) if 'Overall_Result' in overview_df.columns else 0
        _kpi_failed = int((overview_df['Overall_Result'] == 'F').sum()) if 'Overall_Result' in overview_df.columns else 0
        _kpi_pass_pct = round((_kpi_passed / _kpi_appeared) * 100, 2) if _kpi_appeared > 0 else 0

        # ── Write Summary (KPI) sheet as first sheet ──
        kpi_data = {
            'Metric': ['Total', 'Appeared', 'Passed', 'Failed', 'Absent', 'Pass %'],
            'Value': [_kpi_total, _kpi_appeared, _kpi_passed, _kpi_failed, _kpi_absent, f"{_kpi_pass_pct}%"]
        }
        kpi_df = pd.DataFrame(kpi_data)
        kpi_df.to_excel(writer, sheet_name='Summary', index=False)

        ws_kpi = writer.sheets['Summary']
        # Style header row
        kpi_header_fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
        kpi_header_font = Font(bold=True, color='FFFFFFFF', size=12)
        for col_idx in range(1, 3):
            cell = ws_kpi.cell(row=1, column=col_idx)
            cell.fill = kpi_header_fill
            cell.font = kpi_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_kpi.column_dimensions['A'].width = 20
        ws_kpi.column_dimensions['B'].width = 20

        # Style data rows with distinct colors per metric
        kpi_styles = {
            'Total':    (PatternFill(start_color='FFE0E7FF', end_color='FFE0E7FF', fill_type='solid'), Font(color='FF3730A3', bold=True, size=11)),
            'Appeared': (PatternFill(start_color='FFDBEAFE', end_color='FFDBEAFE', fill_type='solid'), Font(color='FF1E40AF', bold=True, size=11)),
            'Passed':   (PatternFill(start_color='FFD1FAE5', end_color='FFD1FAE5', fill_type='solid'), Font(color='FF065F46', bold=True, size=11)),
            'Failed':   (PatternFill(start_color='FFFEE2E2', end_color='FFFEE2E2', fill_type='solid'), Font(color='FF991B1B', bold=True, size=11)),
            'Absent':   (PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid'), Font(color='FF92400E', bold=True, size=11)),
            'Pass %':   (PatternFill(start_color='FFF5F3FF', end_color='FFF5F3FF', fill_type='solid'), Font(color='FF7C3AED', bold=True, size=11)),
        }
        for row_idx in range(2, 8):
            metric = str(ws_kpi.cell(row=row_idx, column=1).value or '')
            if metric in kpi_styles:
                fill, font = kpi_styles[metric]
                for col_idx in range(1, 3):
                    ws_kpi.cell(row=row_idx, column=col_idx).fill = fill
                    ws_kpi.cell(row=row_idx, column=col_idx).font = font
                    ws_kpi.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # ── Write Overview sheet with grouped 2-row headers (matching dashboard) ──
        from openpyxl.utils import get_column_letter as _gcl
        overview_df.to_excel(writer, sheet_name='Overview', index=False, startrow=1)
        ws_ov = writer.sheets['Overview']

        # Build grouped headers: subject cols get [Subject Name, Component], others get ["", colname]
        components = ['Internal', 'External', 'Total', 'Result']
        for col_idx, col_name in enumerate(overview_df.columns, 1):
            top_val = ""
            bot_val = col_name
            for comp in components:
                if col_name.endswith(f" {comp}"):
                    top_val = col_name[:-len(comp)].strip()
                    bot_val = comp
                    break
            ws_ov.cell(row=1, column=col_idx, value=top_val)
            ws_ov.cell(row=2, column=col_idx, value=bot_val)

        # Merge adjacent cells in row 1 that have the same subject group name
        merge_start = 1
        for col_idx in range(2, ws_ov.max_column + 2):
            prev_val = str(ws_ov.cell(row=1, column=merge_start).value or '')
            curr_val = str(ws_ov.cell(row=1, column=col_idx).value or '') if col_idx <= ws_ov.max_column else ''
            if curr_val != prev_val or col_idx > ws_ov.max_column:
                if merge_start < col_idx - 1 and prev_val:
                    ws_ov.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col_idx - 1)
                elif not prev_val:
                    # Non-subject cols: merge row 1 and row 2 vertically
                    for c in range(merge_start, col_idx):
                        ws_ov.cell(row=1, column=c, value=ws_ov.cell(row=2, column=c).value)
                        ws_ov.cell(row=2, column=c, value='')
                        ws_ov.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
                merge_start = col_idx

        ranking_df.to_excel(writer, sheet_name='Ranking (Marks)', index=False)
        if sgpa_computed and sgpa_df is not None:
            sgpa_df.to_excel(writer, sheet_name='Ranking (SGPA)', index=False)
        else:
            # Write a message sheet explaining SGPA is not calculated
            msg_df = pd.DataFrame({'Note': [
                'SGPA Ranking could not be calculated.',
                'Possible reasons:',
                '1. No credit mapping found for the selected scheme/semester',
                '2. Scheme and semester were not selected on the overview page',
                'Ensure you select the correct Scheme and Semester before downloading.',
            ]})
            msg_df.to_excel(writer, sheet_name='Ranking (SGPA)', index=False)
        if not subject_df.empty:
            subject_df.to_excel(writer, sheet_name='Subject Analysis', index=False)
        if not category_df.empty:
            category_df.to_excel(writer, sheet_name='Category Breakdown', index=False)

        # --- Color definitions ---
        header_fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFFFF')
        pass_fill = PatternFill(start_color='FFECFDF5', end_color='FFECFDF5', fill_type='solid')
        pass_font = Font(color='FF065F46')
        fail_fill = PatternFill(start_color='FFFEF2F2', end_color='FFFEF2F2', fill_type='solid')
        fail_font = Font(color='FF991B1B')
        absent_fill = PatternFill(start_color='FFFFFBEB', end_color='FFFFFBEB', fill_type='solid')
        absent_font = Font(color='FFB45309', bold=True)
        fcd_fill = PatternFill(start_color='FFF5F3FF', end_color='FFF5F3FF', fill_type='solid')
        fcd_font = Font(color='FF7C3AED', bold=True)
        fc_fill = PatternFill(start_color='FFF0F9FF', end_color='FFF0F9FF', fill_type='solid')
        fc_font = Font(color='FF075985')
        sc_fill = PatternFill(start_color='FFFFFBF0', end_color='FFFFFBF0', fill_type='solid')
        sc_font = Font(color='FFB45309')
        rank1_fill = PatternFill(start_color='FFFFFBEB', end_color='FFFFFBEB', fill_type='solid')
        rank1_font = Font(color='FF92400E', bold=True)
        rank2_fill = PatternFill(start_color='FFF0F9FF', end_color='FFF0F9FF', fill_type='solid')
        rank2_font = Font(color='FF075985', bold=True)
        rank3_fill = PatternFill(start_color='FFFFF7ED', end_color='FFFFF7ED', fill_type='solid')
        rank3_font = Font(color='FF9A3412', bold=True)
        odd_fill = PatternFill(start_color='FFF9FAFB', end_color='FFF9FAFB', fill_type='solid')

        def _style_headers(ws):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                max_len = max(len(str(cell.value or '')), 10)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

        def _find_col(ws, name):
            for col_idx in range(1, ws.max_column + 1):
                if str(ws.cell(row=1, column=col_idx).value or '').strip() == name:
                    return col_idx
            return None

        def _apply_row_fill(ws, row_idx, fill, font=None):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
                if font:
                    ws.cell(row=row_idx, column=col_idx).font = font

        # ── Style Overview sheet (2-row grouped header, data starts row 3) ──
        ws_ov = writer.sheets['Overview']
        # Style both header rows
        for hdr_row in [1, 2]:
            for col_idx in range(1, ws_ov.max_column + 1):
                cell = ws_ov.cell(row=hdr_row, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                max_len = max(len(str(cell.value or '')), 10)
                ws_ov.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)
        # Find Overall_Result column (could be in row 1 for merged non-subject cols, or row 2 for subject cols)
        res_ci = None
        for col_idx in range(1, ws_ov.max_column + 1):
            for hdr_row in [1, 2]:
                if str(ws_ov.cell(row=hdr_row, column=col_idx).value or '').strip() == 'Overall_Result':
                    res_ci = col_idx
                    break
            if res_ci:
                break
        for row_idx in range(3, ws_ov.max_row + 1):
            val = str(ws_ov.cell(row=row_idx, column=res_ci).value or '').strip().upper() if res_ci else ''
            if val in ['F', 'FAIL']:
                _apply_row_fill(ws_ov, row_idx, fail_fill, fail_font)
            elif val in ['A', 'ABSENT']:
                _apply_row_fill(ws_ov, row_idx, absent_fill, absent_font)
            elif val in ['P', 'PASS']:
                _apply_row_fill(ws_ov, row_idx, pass_fill, pass_font)
            elif row_idx % 2 == 1:
                _apply_row_fill(ws_ov, row_idx, odd_fill)

        # ── Style Ranking (Marks) sheet ──
        ws_rk = writer.sheets['Ranking (Marks)']
        _style_headers(ws_rk)
        res_ci = _find_col(ws_rk, 'Overall_Result')
        rank_ci = _find_col(ws_rk, 'Class_Rank')
        for row_idx in range(2, ws_rk.max_row + 1):
            val = str(ws_rk.cell(row=row_idx, column=res_ci).value or '').strip().upper() if res_ci else ''
            rk = None
            if rank_ci:
                try: rk = int(ws_rk.cell(row=row_idx, column=rank_ci).value)
                except: rk = None
            if val in ['F', 'FAIL']:
                _apply_row_fill(ws_rk, row_idx, fail_fill, fail_font)
            elif val in ['A', 'ABSENT']:
                _apply_row_fill(ws_rk, row_idx, absent_fill, absent_font)
            elif rk == 1:
                _apply_row_fill(ws_rk, row_idx, rank1_fill, rank1_font)
            elif rk == 2:
                _apply_row_fill(ws_rk, row_idx, rank2_fill, rank2_font)
            elif rk == 3:
                _apply_row_fill(ws_rk, row_idx, rank3_fill, rank3_font)
            elif val in ['P', 'PASS']:
                _apply_row_fill(ws_rk, row_idx, pass_fill, pass_font)
            elif row_idx % 2 == 1:
                _apply_row_fill(ws_rk, row_idx, odd_fill)

        # ── Style Ranking (SGPA) sheet ──
        ws_sgpa = writer.sheets['Ranking (SGPA)']
        _style_headers(ws_sgpa)
        if sgpa_computed and sgpa_df is not None:
            sgpa_res_ci = _find_col(ws_sgpa, 'Overall_Result')
            sgpa_rank_ci = _find_col(ws_sgpa, 'SGPA_Rank')
            for row_idx in range(2, ws_sgpa.max_row + 1):
                val = str(ws_sgpa.cell(row=row_idx, column=sgpa_res_ci).value or '').strip().upper() if sgpa_res_ci else ''
                rk = None
                if sgpa_rank_ci:
                    try: rk = int(ws_sgpa.cell(row=row_idx, column=sgpa_rank_ci).value)
                    except: rk = None
                if val in ['FAIL']:
                    _apply_row_fill(ws_sgpa, row_idx, fail_fill, fail_font)
                elif val in ['ABSENT']:
                    _apply_row_fill(ws_sgpa, row_idx, absent_fill, absent_font)
                elif rk == 1:
                    _apply_row_fill(ws_sgpa, row_idx, rank1_fill, rank1_font)
                elif rk == 2:
                    _apply_row_fill(ws_sgpa, row_idx, rank2_fill, rank2_font)
                elif rk == 3:
                    _apply_row_fill(ws_sgpa, row_idx, rank3_fill, rank3_font)
                elif val in ['PASS']:
                    _apply_row_fill(ws_sgpa, row_idx, pass_fill, pass_font)
                elif row_idx % 2 == 1:
                    _apply_row_fill(ws_sgpa, row_idx, odd_fill)
        else:
            # Style the info message sheet
            info_fill = PatternFill(start_color='FFFFF7ED', end_color='FFFFF7ED', fill_type='solid')
            info_font = Font(color='FF9A3412', italic=True)
            for row_idx in range(2, ws_sgpa.max_row + 1):
                _apply_row_fill(ws_sgpa, row_idx, info_fill, info_font)

        # ── Style Subject Analysis sheet ──
        if 'Subject Analysis' in writer.sheets:
            ws_sa = writer.sheets['Subject Analysis']
            _style_headers(ws_sa)
            # Find column indices for cell-level coloring
            sa_passed_ci = _find_col(ws_sa, 'Passed')
            sa_failed_ci = _find_col(ws_sa, 'Failed')
            sa_absent_ci = _find_col(ws_sa, 'Absent')
            sa_appeared_ci = _find_col(ws_sa, 'Appeared')
            sa_total_ci = _find_col(ws_sa, 'Total')
            sa_pp_ci = _find_col(ws_sa, 'Pass %')

            # Cell-level fills for value columns
            passed_cell_fill = PatternFill(start_color='FFD1FAE5', end_color='FFD1FAE5', fill_type='solid')
            passed_cell_font = Font(color='FF065F46', bold=True)
            failed_cell_fill = PatternFill(start_color='FFFEE2E2', end_color='FFFEE2E2', fill_type='solid')
            failed_cell_font = Font(color='FF991B1B', bold=True)
            absent_cell_fill = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
            absent_cell_font = Font(color='FF92400E', bold=True)
            appeared_cell_fill = PatternFill(start_color='FFDBEAFE', end_color='FFDBEAFE', fill_type='solid')
            appeared_cell_font = Font(color='FF1E40AF', bold=True)
            total_cell_fill = PatternFill(start_color='FFE0E7FF', end_color='FFE0E7FF', fill_type='solid')
            total_cell_font = Font(color='FF3730A3', bold=True)
            pp_high_fill = PatternFill(start_color='FFA7F3D0', end_color='FFA7F3D0', fill_type='solid')
            pp_high_font = Font(color='FF065F46', bold=True)
            pp_mid_fill = PatternFill(start_color='FFFEF9C3', end_color='FFFEF9C3', fill_type='solid')
            pp_mid_font = Font(color='FF854D0E', bold=True)
            pp_low_fill = PatternFill(start_color='FFFECACA', end_color='FFFECACA', fill_type='solid')
            pp_low_font = Font(color='FF991B1B', bold=True)

            for row_idx in range(2, ws_sa.max_row + 1):
                if row_idx % 2 == 1:
                    _apply_row_fill(ws_sa, row_idx, odd_fill)
                # Color individual value cells by column
                if sa_passed_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_passed_ci)
                    c.fill = passed_cell_fill
                    c.font = passed_cell_font
                if sa_failed_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_failed_ci)
                    c.fill = failed_cell_fill
                    c.font = failed_cell_font
                if sa_absent_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_absent_ci)
                    c.fill = absent_cell_fill
                    c.font = absent_cell_font
                if sa_appeared_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_appeared_ci)
                    c.fill = appeared_cell_fill
                    c.font = appeared_cell_font
                if sa_total_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_total_ci)
                    c.fill = total_cell_fill
                    c.font = total_cell_font
                if sa_pp_ci:
                    c = ws_sa.cell(row=row_idx, column=sa_pp_ci)
                    try: pct = float(c.value)
                    except: pct = 0
                    if pct >= 70:
                        c.fill = pp_high_fill
                        c.font = pp_high_font
                    elif pct >= 50:
                        c.fill = pp_mid_fill
                        c.font = pp_mid_font
                    else:
                        c.fill = pp_low_fill
                        c.font = pp_low_font

        # ── Style Category Breakdown sheet ──
        if 'Category Breakdown' in writer.sheets:
            ws_cb = writer.sheets['Category Breakdown']
            _style_headers(ws_cb)
            cat_ci = _find_col(ws_cb, 'Category')
            # Distinct colors per category
            pc_fill = PatternFill(start_color='FFECFDF5', end_color='FFECFDF5', fill_type='solid')
            pc_font = Font(color='FF065F46')
            fcd_row_fill = PatternFill(start_color='FFEDE9FE', end_color='FFEDE9FE', fill_type='solid')
            fcd_row_font = Font(color='FF6D28D9', bold=True)
            fc_row_fill = PatternFill(start_color='FFDBEAFE', end_color='FFDBEAFE', fill_type='solid')
            fc_row_font = Font(color='FF1E40AF', bold=True)
            sc_row_fill = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
            sc_row_font = Font(color='FF92400E', bold=True)
            for row_idx in range(2, ws_cb.max_row + 1):
                cat = str(ws_cb.cell(row=row_idx, column=cat_ci).value or '').strip() if cat_ci else ''
                if 'FCD' in cat or 'Distinction' in cat:
                    _apply_row_fill(ws_cb, row_idx, fcd_row_fill, fcd_row_font)
                elif 'First' in cat:
                    _apply_row_fill(ws_cb, row_idx, fc_row_fill, fc_row_font)
                elif 'Second' in cat:
                    _apply_row_fill(ws_cb, row_idx, sc_row_fill, sc_row_font)
                elif 'Pass' in cat:
                    _apply_row_fill(ws_cb, row_idx, pc_fill, pc_font)
                elif row_idx % 2 == 1:
                    _apply_row_fill(ws_cb, row_idx, odd_fill)

    return dcc.send_bytes(out.getvalue(), 'Complete_Report.xlsx')