# pages/subject_analysis.py
# Final stable version — Fixed DuplicateCallback error with 'initial_duplicate'
# Updated: Custom Graph Tooltip + StringIO Fix

import dash
from dash import html, dcc, Input, Output, State, callback, dash_table, no_update, ALL
import dash_bootstrap_components as dbc
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dash.exceptions import PreventUpdate
from io import StringIO  # <--- Added for stability
from cache_config import cache
import json
import re  # Added for section extraction

dash.register_page(__name__, path="/subject_analysis", name="Subject Analysis")

# ---------- Helper Functions ----------
def extract_numeric(roll):
    digits = re.findall(r'\d+', str(roll))
    return int(digits[-1]) if digits else 0

def sa_assign_section(roll_no, section_ranges=None, usn_mapping=None):
    roll_str = str(roll_no).strip().upper()
    if usn_mapping and roll_str in usn_mapping:
         return usn_mapping[roll_str]

    roll_num = extract_numeric(roll_no)
    if section_ranges:
        for sec_name, (start, end) in section_ranges.items():
            start_num = extract_numeric(start)
            end_num = extract_numeric(end)
            if start_num <= roll_num <= end_num:
                return sec_name
    return "Not Assigned"

# ==================== Global Styles ====================
import os as _os
_css_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "styles", "subject_analysis_page.css")
with open(_css_path, "r", encoding="utf-8") as _f:
    PAGE_CSS = _f.read()

# ==================== Layout ====================
layout = dbc.Container([
    dcc.Markdown(f"<style>{PAGE_CSS}</style>", dangerously_allow_html=True),

    html.Div([
        html.H3("📊 Subject-wise Performance Analysis", 
                className="text-center fw-bold mb-2 sa-title"),
        html.P("Analyze class performance across multiple subjects with interactive visuals.", 
               className="text-center text-muted mb-4")
    ], className="sa-wrap sa-card p-3 mb-3"),

    # --- Controls (with Exports included) ---
    html.Div(
        dbc.Card(dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H6("Select Subjects", className="fw-bold text-muted mb-1"),
                    html.Div([
                        dcc.Dropdown(
                            id="sa-subject-checklist",
                            options=[], value=[], multi=True,
                            placeholder="Select subjects to analyze...",
                            className="custom-dropdown",
                            searchable=True,
                            clearable=True,
                            optionHeight=50,
                            maxHeight=300,
                            style={
                                "position": "relative", 
                                "zIndex": "1000",
                                "minHeight": "45px"
                            }
                        ),
                        html.Div(style={"height": "10px"})
                    ], style={"position": "relative", "zIndex": "1000"}),
                ], xs=12, lg=4, style={"position": "relative", "zIndex": "1060", "overflow": "visible"}),
                
                dbc.Col([
                    html.H6("Select Section", className="fw-bold text-muted mb-1"),
                    html.Div([
                        dcc.Dropdown(
                            id="sa-section-filter",
                            options=[{"label": "All Sections", "value": "ALL"}],
                            value="ALL", clearable=False, className="custom-dropdown",
                            searchable=True,
                            optionHeight=50,
                            maxHeight=300,
                            style={
                                "position": "relative", 
                                "zIndex": "1000",
                                "minHeight": "45px"
                            }
                        ),
                        html.Div(style={"height": "10px"})
                    ], style={"position": "relative", "zIndex": "1000"}),
                ], xs=12, lg=2, style={"position": "relative", "zIndex": "1050", "overflow": "visible"}),

                dbc.Col([
                    html.H6("Filter by Result", className="fw-bold text-muted mb-1"),
                    html.Div([
                        dcc.Dropdown(
                            id="sa-result-filter",
                            options=[
                                {"label": "All Students", "value": "ALL"},
                                {"label": "Passed Only", "value": "PASS"},
                                {"label": "Failed Only", "value": "FAIL"},
                                {"label": "Absent Only", "value": "ABSENT"},
                            ],
                            value="ALL", clearable=False, className="custom-dropdown",
                            searchable=True,
                            optionHeight=50,
                            maxHeight=300,
                            style={
                                "position": "relative", 
                                "zIndex": "1000",
                                "minHeight": "45px"
                            }
                        ),
                        html.Div(style={"height": "10px"})
                    ], style={"position": "relative", "zIndex": "1000"}),
                ], xs=12, lg=3, style={"position": "relative", "zIndex": "1050", "overflow": "visible"}),

                dbc.Col([
                    html.H6("Actions", className="fw-bold text-muted mb-1"),
                    dbc.ButtonGroup([
                        dbc.Button("CSV", id="sa-export-csv", color="primary", outline=True, className="me-1"),
                        dbc.Button("Excel", id="sa-export-xlsx", color="success", outline=True, className="me-1"),
                        dbc.Button("PDF", id="sa-export-pdf", color="danger", outline=True, className="me-1"),
                        dbc.Button("📖", id="sa-open-legend", color="info", outline=True, title="Rules & Guidelines"),
                    ], className="w-100", style={"height": "45px"}),
                ], xs=12, lg=3), 
            ], className="g-3 align-items-start"),
            dbc.Row([
                dbc.Col(
                    dbc.Spinner(
                        html.Div(id="sa-selected-count", className="mt-2 small text-muted"),
                        size="sm",
                        color="primary"
                    )
                )
            ], className="mt-1")
        ], style={"overflow": "visible", "position": "relative"}), className="sa-card", style={"overflow": "visible", "position": "relative"}),
        className="mb-4", style={"overflow": "visible", "position": "relative", "zIndex": "1050"}
    ),

    # --- KPIs ---
    # Wrapped in its own Loading component
    dcc.Loading(type="default", children=[
        dbc.Card(dbc.CardBody([
            html.Div([
                html.H6([html.I(className="bi bi-activity me-2 text-primary"), "Performance Metrics"], className="fw-bold mb-0 text-primary"),
                dbc.Button(
                    [html.I(className="bi bi-cloud-arrow-down-fill me-2"), "Download Performance Summary"], 
                    id="sa-export-all-kpis", size="sm", color="primary", outline=True, className="fw-bold shadow-sm"
                )
            ], className="d-flex justify-content-between align-items-center mb-3"),
            html.Div(id="sa-kpi-cards")
        ]), className="sa-card mb-4")
    ]),

    # --- Table ---
    # Wrapped in its own Loading component
    dcc.Loading(type="default", children=[
        dbc.Card(dbc.CardBody([
            html.Div([
                html.H5("📋 Detailed Subject Breakdown", className="fw-bold mb-0 text-center"),
                html.Div([
                    dbc.Button("Download CSV", id="sa-export-detailed-csv", color="primary", outline=True, size="sm", className="me-2"),
                    dbc.Button("Download Excel", id="sa-export-detailed-xlsx", color="success", outline=True, size="sm"),
                ])
            ], className="d-flex justify-content-between align-items-center mb-3"),
            dash_table.DataTable(
                id="sa-subject-table",
                columns=[], data=[],
                style_table={
                    "overflowX": "auto", 
                    "borderRadius": "8px", 
                    "border": "1px solid #d1d5db",
                    "boxShadow": "0 4px 6px -1px rgba(0, 0, 0, 0.1)"
                },
                style_cell={
                    "textAlign": "center", 
                    "padding": "12px",
                    "fontFamily": "Inter, Segoe UI, system-ui, -apple-system, Arial",
                    "fontSize": "13px",
                    "color": "#1f2937",
                    "border": "1px solid #e5e7eb"
                },
                style_header={
                    "backgroundColor": "#1f2937", 
                    "color": "#ffffff",
                    "fontWeight": "700",
                    "textTransform": "uppercase",
                    "fontSize": "12px",
                    "letterSpacing": "0.5px",
                    "borderBottom": "2px solid #111827"
                },
                style_data={
                    "whiteSpace": "normal",
                    "height": "auto",
                    "backgroundColor": "#ffffff"
                },
                style_data_conditional=[
                    {'if': {'row_index': 'odd'}, 'backgroundColor': '#f3f4f6'},
                    {"if": {"state": "selected"}, "backgroundColor": "rgba(59, 130, 246, 0.1)", "border": "1px solid #3b82f6"},
                    
                    # Result coloring
                    {"if": {"filter_query": "{Overall_Result} = 'Fail'"}, "backgroundColor": "#fef2f2", "color": "#dc2626", "fontWeight": "700"},
                    {"if": {"filter_query": "{Overall_Result} = 'Pass'"}, "backgroundColor": "#ecfdf5", "color": "#059669", "fontWeight": "700"},
                    {"if": {"filter_query": "{Overall_Result} = 'FCD'"}, "backgroundColor": "#ecfdf5", "color": "#059669", "fontWeight": "700"},
                    {"if": {"filter_query": "{Overall_Result} = 'FC'"}, "backgroundColor": "#ecfdf5", "color": "#10b981", "fontWeight": "700"},
                    {"if": {"filter_query": "{Overall_Result} = 'SC'"}, "backgroundColor": "#ecfdf5", "color": "#3b82f6", "fontWeight": "700"},
                    {"if": {"filter_query": "{Overall_Result} = 'Absent'"}, "backgroundColor": "#fff7ed", "color": "#d97706", "fontWeight": "700"},
                ],
                merge_duplicate_headers=True,
                page_size=10,
                sort_action="native",
                filter_action="native",
            ),
        ]), className="sa-card mb-4"),
    ]),

    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("📊 Subject Analysis — Rules & Guidelines")),
        dbc.ModalBody(
            html.Div([
                html.H5("🚀 Getting Started", className="text-primary fw-bold mb-2"),
                html.P("This page lets you analyze performance at the subject level — see pass/fail distributions, average marks, and drill down into individual subjects. Data is loaded from the Overview page.", className="text-muted small mb-3"),

                html.H6("🎯 Subject Selection", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Multi-select: "), "Choose one or more subjects from the dropdown to analyze."]),
                    html.Li([html.Strong("Select All: "), "Use the 'Select All' option to include every detected subject."]),
                    html.Li([html.Strong("Remove All: "), "Quickly clear your selection to start fresh."]),
                    html.Li("Subject codes and full names (when available) are displayed for easy identification."),
                ], className="small"),
                html.Hr(),

                html.H6("🔍 Filters", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Section Filter: "), "Narrow results to a specific section or view all sections."]),
                    html.Li([html.Strong("Result Filter: "), "View All Students, or filter to show only Passed / Failed / Absent students."]),
                ], className="small"),
                html.Hr(),

                html.H6("📝 Per-Subject Status Logic", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Pass: "), "Result column = 'P' or passing grade."]),
                    html.Li([html.Strong("Fail: "), "Result column = 'F' or marks below passing threshold."]),
                    html.Li([html.Strong("Absent: "), "External Marks = 0 AND Result = 'A' or blank."]),
                ], className="small"),
                html.Hr(),

                html.H6("📝 Overall Student Status (across selected subjects)", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Pass: "), "Passed in ALL selected subjects."]),
                    html.Li([html.Strong("Fail: "), "Failed or absent in AT LEAST ONE selected subject."]),
                    html.Li([html.Strong("Absent: "), "Absent in ALL selected subjects."]),
                ], className="small"),
                html.Hr(),

                html.H6("📊 Charts & Visualizations", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Pass vs Fail Distribution (Pie): "), "Shows the proportion of students passing vs failing across selected subjects."]),
                    html.Li([html.Strong("Average Marks & Performance (Bar): "), "Compares average marks per subject side by side."]),
                ], className="small"),
                html.Hr(),

                html.H6("📋 Subject Performance Table", className="fw-bold text-dark"),
                html.Ul([
                    html.Li("Shows per-subject summary: Total Students, Appeared, Absent, Passed, Failed, and Pass %."),
                    html.Li("Detailed breakdown table shows each student's Internal, External, Total, and Result per subject."),
                    html.Li("Tables support sorting and filtering for quick lookups."),
                ], className="small"),
                html.Hr(),

                html.H6("🖱️ Interactive Features", className="fw-bold text-dark"),
                html.Ul([
                    html.Li([html.Strong("Click KPI cards: "), "Opens a modal with a detailed student list for that category. Download as Excel."]),
                    html.Li([html.Strong("Export: "), "Use the CSV, Excel, or PDF buttons to download the full report."]),
                ], className="small"),

                dbc.Alert("Tip: For SGPA, rankings, and VTU class categories (FCD, FC, etc.), visit the Ranking page.", color="info", className="small py-2 mb-0")
            ])
        ),
        dbc.ModalFooter(dbc.Button("Got it!", id="sa-close-legend", className="ms-auto", color="primary"))
    ], id="sa-legend-modal", is_open=False, size="lg", style={"zIndex": 10000}),

    # --- KPI Popup Modal ---
    dbc.Modal([
        dbc.ModalHeader([
            dbc.ModalTitle(id="sa-kpi-modal-title", className="fw-bold text-primary"),
            html.Div([
                dbc.Button("Download List (Excel)", id="sa-kpi-modal-excel-top", color="success", outline=True, size="sm", className="me-2"),
                dbc.Button("Close", id="sa-kpi-modal-close-top", color="secondary", size="sm")
            ], className="ms-auto d-flex")
        ], close_button=False),
        dbc.ModalBody([
            dash_table.DataTable(
                id="sa-kpi-modal-table",
                columns=[], data=[],
                style_table={"overflowX": "auto", "borderRadius": "8px", "border": "1px solid #d1d5db"},
                style_cell={"textAlign": "center", "padding": "12px", "fontSize": "13px"},
                style_header={"backgroundColor": "#1f2937", "color": "#ffffff", "fontWeight": "700"},
                page_action='none',
                style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': '#f3f4f6'}],
            )
        ]),
        dbc.ModalFooter([
            dbc.Button("Download List (Excel)", id="sa-kpi-modal-excel", color="success", outline=True),
            dbc.Button("Close", id="sa-kpi-modal-close", className="ms-auto", color="secondary")
        ])
    ], id="sa-kpi-modal", is_open=False, size="xl", style={"zIndex": 10000}),
    
    dcc.Download(id="sa-kpi-excel-download"),

    # --- Tabs for Charts ---
    # Wrapped in its own Loading component
    dcc.Loading(type="default", children=[
        dbc.Card(dbc.CardBody([
            dcc.Tabs(id="sa-chart-tabs", value="pie", children=[
                dcc.Tab(label="🎯 Pass vs Fail Distribution", value="pie"),
                dcc.Tab(label="📈 Average Marks & Performance", value="bar"),
            ]),
            html.Div(id="sa-subject-analysis-chart", className="mt-3"),
        ]), className="sa-card mb-4"),
    ]),


    # Hidden Download components
    dcc.Download(id="sa-download-csv"),
    dcc.Download(id="sa-download-xlsx"),
    dcc.Download(id="sa-download-summary-xlsx"),
    dcc.Download(id="sa-download-all-kpis"),
    html.Div(id="sa-pdf-download-trigger", style={"display": "none"}),

    # Use session stores to match global app.py stores
], fluid=True, className="pb-4")

# ==================== CALLBACKS ====================

# 1️⃣ Dropdown Control
@callback(
    Output("sa-section-filter", "options"),
    Input("section-data", "data"),
    Input("usn-mapping-store", "data")  # <-- Add this input
)
def update_section_dropdown(section_data, usn_mapping):
    options = [{"label": "All Sections", "value": "ALL"}]
    sections = set()
    
    # Add sections from Range Config
    if section_data:
        for sec in section_data.keys():
            sections.add(sec)
            
    # Add sections from Excel Mapping Config
    if usn_mapping:
        for sec in usn_mapping.values():
            sections.add(sec)
            
    # Sort and append unique sections
    for sec in sorted(list(sections)):
        options.append({"label": sec, "value": sec})
        
    # ONLY show "Not Assigned" if no sections are mapped at all
    if not sections:
        options.append({"label": "Not Assigned", "value": "Not Assigned"})
        
    return options

@callback(
    Output("sa-subject-checklist", "options"),
    Output("sa-subject-checklist", "value"),
    Input("overview-selected-subjects", "data"),
    Input("sa-subject-checklist", "value"),
    State("stored-data", "data"),
    prevent_initial_call=False
)
def update_subject_dropdown(overview_subjects, current_value, session_id):
    if not overview_subjects:
        return [], []
        
    name_mapping = {}
    if session_id:
        df = cache.get(session_id)
        if df is not None:
            for subj in overview_subjects:
                display_name = subj
                subj_cols = [c for c in df.columns if c.startswith(subj)]
                for col in subj_cols:
                    if " - " in col:
                        try:
                            rest = col.split(" - ", 1)[1]
                            for suffix in ["Result", "Total", "Internal", "External"]:
                                if rest.strip().endswith(suffix):
                                    possible_name = rest.rsplit(suffix, 1)[0].strip()
                                    if possible_name:
                                        display_name = f"{subj} - {possible_name}"
                                    break
                            if display_name != subj:
                                break
                        except: pass
                name_mapping[subj] = display_name

    # Add "Select All" and "Remove All" options at the top
    options = [
        {"label": "✓ Select All", "value": "__SELECT_ALL__"},
        {"label": "✕ Remove All", "value": "__REMOVE_ALL__"}
    ] 
    
    for s in overview_subjects:
        full_name = name_mapping.get(s, s)
        
        # Extract code part to display
        if " - " in full_name:
            c_part = full_name.split(" - ", 1)[0]
        else:
            c_part = full_name
            
        options.append({
            "label": html.Span(c_part, className="fw-bold"),
            "value": s,
            "title": full_name  # Keeps full name for standard HTML tooltips
        })

    all_subject_values = [opt["value"] for opt in options[2:]]  # Exclude the special markers
    
    ctx = dash.callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else "INITIAL_LOAD"

    if trigger == "sa-subject-checklist":
        # If "Select All" is clicked, select all subjects
        if current_value and "__SELECT_ALL__" in current_value:
            return options, all_subject_values
        # If "Remove All" is clicked, clear all
        elif current_value and "__REMOVE_ALL__" in current_value:
            return options, []
        # Remove special markers if user manually selects other subjects
        filtered_value = [v for v in (current_value or []) if not v.startswith("__")]
        return options, filtered_value
    
    # Return ALL values on initial load so that the charts and tables populate immediately
    return options, all_subject_values


# 2️⃣ Main Analysis
@callback(
    Output("sa-selected-count", "children"),
    Output("sa-kpi-cards", "children"),
    Output("sa-subject-table", "columns"),
    Output("sa-subject-table", "data"),
    Output("sa-subject-analysis-chart", "children"),
    Input("sa-subject-checklist", "value"),
    Input("sa-result-filter", "value"),
    Input("sa-section-filter", "value"),
    Input("sa-chart-tabs", "value"),
    State("stored-data", "data"),
    State("section-data", "data"),
    State("usn-mapping-store", "data"),
    prevent_initial_call=False
)
def update_analysis(selected_subjects, result_filter, section_filter, chart_tab, session_id, section_ranges, usn_mapping):
    if not session_id:
        raise PreventUpdate
    
    # Retrieve from Server Cache
    df = cache.get(session_id)
    if df is None:
        return "Session expired", html.P("Please return to Overview and upload data.", className="text-danger"), [], [], html.Div()

    # Remove the special markers if present
    selected_subjects = [s for s in (selected_subjects or []) if not s.startswith("__")]
    
    if not selected_subjects:
        return "0 subjects selected", html.P("Please select at least one subject.", className="text-muted text-center"), [], [], html.Div()

    first_col = df.columns[0]

    if "Name" not in df.columns:
        df["Name"] = ""

    selected_cols = []
    for subj in selected_subjects:
        selected_cols.extend([c for c in df.columns if c.startswith(f"{subj} ")])
    selected_cols = list(dict.fromkeys(selected_cols))

    df_sel = df[[first_col, "Name"] + selected_cols].copy()
    
    # Apply Section Assignment and Filtering
    df_sel['Section'] = df_sel[first_col].apply(lambda x: sa_assign_section(x, section_ranges, usn_mapping))
    if section_filter and section_filter != "ALL":
        df_sel = df_sel[df_sel['Section'] == section_filter]
        if df_sel.empty:
            return f"0 students in section {section_filter}", html.P(f"No data for section {section_filter}."), [], [], html.Div()

    num_cols = [c for c in df_sel.columns if any(k in c for k in ["Internal", "External", "Total"])]
    for c in num_cols:
        df_sel[c] = pd.to_numeric(df_sel[c], errors="coerce")

    result_cols = [c for c in df_sel.columns if "Result" in c]
    if result_cols:
        # Detect Absent ("A"), Pass ("P"), and Fail ("F")
        # Logic: 
        # - Fail if ANY 'F'
        # - Fail if 'A' exists but not ALL are 'A' (Absent in 1 = Fail)
        # - Absent only if ALL are 'A'
        # - Pass otherwise
        def determine_result(x):
            vals = [str(v).strip().upper() for v in x if pd.notna(v) and str(v).strip() != ""]
            
            # If student has NO data for any selected subject, mark as NA (to filter out)
            if not vals:
                return "NA"

            has_fail = any(v in ["F", "FAIL"] for v in vals)
            has_absent = any(v in ["A", "ABSENT"] for v in vals)
            all_absent = all(v in ["A", "ABSENT"] for v in vals)

            if has_fail:
                return "Fail"
            elif has_absent:
                # If absent in all, then Absent. If absent in some (and passed others), then Fail.
                return "Absent" if all_absent else "Fail"
            else:
                return "Pass"
        
        df_sel["Overall_Result"] = df_sel[result_cols].apply(determine_result, axis=1)
        
        # Filter out students who aren't taking ANY of the selected subjects (Result = NA)
        df_sel = df_sel[df_sel["Overall_Result"] != "NA"]
        
        # Calculate percentage for full FCD/FC breakdown based on all subjects in original df
        all_subj_cols = [c for c in df.columns if c.strip().endswith(' Total') and c not in ['Total_Marks', 'Grand Total'] and 'grand total' not in c.lower()]
        
        def assign_fcd(row_idx):
            res = df_sel.at[row_idx, "Overall_Result"]
            if res != "Pass":
                return res
                
            orig_row = df.loc[row_idx]
            attempted = 0
            t_marks = 0
            for c in all_subj_cols:
                v = pd.to_numeric(orig_row.get(c, 0), errors='coerce')
                if pd.notna(v) and v > 0:
                    attempted += 1
                    t_marks += v
            if attempted == 0: return "Pass"
            pct = (t_marks / (attempted * 100)) * 100
            
            if pct >= 70: return "FCD"
            elif pct >= 60: return "FC"
            elif pct >= 50: return "SC"
            else: return "Pass"
            
        df_sel["Overall_Result"] = [assign_fcd(idx) for idx in df_sel.index]

    else:
        df_sel["Overall_Result"] = "Pass"

    pass_cats = ["Pass", "FCD", "FC", "SC"]

    # Count Absent, Appeared, Passed, Failed (before filtering)
    total_students = len(df_sel)
    absent = (df_sel["Overall_Result"] == "Absent").sum()
    appeared = total_students - absent
    passed = df_sel["Overall_Result"].isin(pass_cats).sum()
    failed = (df_sel["Overall_Result"] == "Fail").sum()
    pass_pct_appeared = round((passed / appeared) * 100, 2) if appeared > 0 else 0

    # Apply result filter
    if result_filter == "PASS":
        df_sel = df_sel[df_sel["Overall_Result"].isin(pass_cats)]
    elif result_filter == "FAIL":
        df_sel = df_sel[df_sel["Overall_Result"] == "Fail"]
    elif result_filter == "ABSENT":
        df_sel = df_sel[df_sel["Overall_Result"] == "Absent"]
    # For "ALL" filter, keep all rows including Absent

    total = len(df_sel)
    passed_filtered = df_sel["Overall_Result"].isin(pass_cats).sum()
    failed_filtered = (df_sel["Overall_Result"] == "Fail").sum()
    absent_filtered = (df_sel["Overall_Result"] == "Absent").sum()

    # Dynamically build KPI list based on filter
    if result_filter == "ALL":
        kpis = [
            {"id": "total", "label": "TOTAL", "value": total, "color": "#3b82f6", "bg": "#eff6ff", "icon": "bi-people-fill"},
            {"id": "appeared", "label": "APPEARED", "value": appeared, "color": "#10b981", "bg": "#ecfdf5", "icon": "bi-person-check-fill"},
            {"id": "pass", "label": "PASSED", "value": passed, "color": "#0ea5e9", "bg": "#f0f9ff", "icon": "bi-check-circle-fill"},
            {"id": "fail", "label": "FAILED", "value": failed, "color": "#ef4444", "bg": "#fef2f2", "icon": "bi-x-circle-fill"},
            {"id": "absent", "label": "ABSENT", "value": absent, "color": "#f59e0b", "bg": "#fffbeb", "icon": "bi-person-x-fill"},
            {"id": "rate", "label": "PASS %", "value": f"{pass_pct_appeared}%", "color": "#8b5cf6", "bg": "#f5f3ff", "icon": "bi-graph-up"},
        ]
        col_class = "row-cols-2 row-cols-md-3 row-cols-lg-6 g-3"
    elif result_filter == "PASS":
        kpis = [
            {"id": "total", "label": "TOTAL", "value": total, "color": "#3b82f6", "bg": "#eff6ff", "icon": "bi-people-fill"},
            {"id": "pass", "label": "PASSED", "value": passed_filtered, "color": "#10b981", "bg": "#ecfdf5", "icon": "bi-check-circle-fill"},
        ]
        col_class = "row-cols-2 row-cols-md-6 g-3"
    elif result_filter == "FAIL":
        kpis = [
            {"id": "total", "label": "TOTAL", "value": total, "color": "#3b82f6", "bg": "#eff6ff", "icon": "bi-people-fill"},
            {"id": "fail", "label": "FAILED", "value": failed_filtered, "color": "#ef4444", "bg": "#fef2f2", "icon": "bi-x-circle-fill"},
        ]
        col_class = "row-cols-2 row-cols-md-6 g-3"
    else:  # ABSENT
        kpis = [
            {"id": "total", "label": "TOTAL", "value": total, "color": "#3b82f6", "bg": "#eff6ff", "icon": "bi-people-fill"},
            {"id": "absent", "label": "ABSENT", "value": absent_filtered, "color": "#f59e0b", "bg": "#fffbeb", "icon": "bi-person-x-fill"},
        ]
        col_class = "row-cols-2 row-cols-md-6 g-3"


    # =========================================================================
    # SUBJECT-WISE BREAKDOWN (Handle Absent Logic Correctly)
    # =========================================================================
    
    # 1. Subject-wise Analysis Data Structure
    subject_stats = []
    
    for subj in selected_subjects:
        # Robust column lookup: Find the actual column names in df_sel
        # This prevents issues where 'BNSK559 Result' (constructed) doesn't match 'BNSK559  Result' (actual with double space)
        # resulting in fallback or missing data. Use df_sel to ensure consistency with KPIs.
        subj_cols = [c for c in df_sel.columns if c.startswith(subj)]
        
        # Try to extract full subject name if available in columns
        display_name = subj
        for col in subj_cols:
            if " - " in col:
                # Expected format: "Code - Name Component"
                # e.g. "18CS51 - DATA STRUCTURES Total"
                try:
                    # Split by " - " to get "Name Component" part
                    rest = col.split(" - ", 1)[1]
                    # Remove the component suffix
                    for suffix in ["Result", "Total", "Internal", "External"]:
                        if rest.strip().endswith(suffix):
                            possible_name = rest.rsplit(suffix, 1)[0].strip()
                            if possible_name:
                                display_name = f"{subj} - {possible_name}"
                            break
                    if display_name != subj:
                        break
                except:
                    continue

        res_col = next((c for c in subj_cols if "Result" in c), None)
        int_col = next((c for c in subj_cols if "Internal" in c), None)
        ext_col = next((c for c in subj_cols if "External" in c), None)
        tot_col = next((c for c in subj_cols if "Total" in c), None)

        if not res_col:
            continue
            
        cols_to_fetch = [first_col, "Name", res_col]
        if int_col: cols_to_fetch.append(int_col)
        if ext_col: cols_to_fetch.append(ext_col)
        if tot_col: cols_to_fetch.append(tot_col)
        
        # Use filtered dataset
        subj_df = df_sel[cols_to_fetch].copy()

        # --- LOGIC: Validate entries for this subject ---
        # Ensure we only count students who have a valid entry for this subject
        # Drop rows where Result is NaN/None/Empty (Student didn't take this subject)
        subj_df = subj_df[subj_df[res_col].notna()]
        subj_df = subj_df[subj_df[res_col].astype(str).str.strip() != ""]

        if subj_df.empty:
            subject_stats.append({
                "Subject": display_name,
                "Total Students": 0, "Appeared": 0, "Absent": 0, "Passed": 0, "Failed": 0, "Pass %": 0
            })
            continue

        # Standardize Result
        subj_df[res_col] = subj_df[res_col].astype(str).str.strip().str.upper()
        
        # Identify Status
        def get_subj_status(row):
            r = row[res_col]
            e = row[ext_col] if ext_col else 0 
            
            try:
                e_val = float(e)
            except:
                e_val = 0
            
            if r in ['A', 'ABSENT'] and e_val == 0:
                return 'Absent'
            elif r in ['F', 'FAIL']:
                return 'Fail'
            elif r in ['P', 'PASS']:
                return 'Pass'
            else:
                if r in ['A', 'ABSENT']: return 'Absent'
                return 'Ignore' 
        
        if ext_col:
            subj_df[ext_col] = pd.to_numeric(subj_df[ext_col], errors='coerce').fillna(0)
        
        subj_df['Status'] = subj_df.apply(get_subj_status, axis=1)
        
        # Filter invalid statuses
        subj_df = subj_df[subj_df['Status'] != 'Ignore']

        # Stats
        s_total = len(subj_df)
        s_absent = (subj_df['Status'] == 'Absent').sum()
        s_appeared = s_total - s_absent
        s_passed = (subj_df['Status'] == 'Pass').sum()
        s_failed = (subj_df['Status'] == 'Fail').sum()
        s_pass_pct = round((s_passed / s_appeared) * 100, 2) if s_appeared > 0 else 0
        
        subject_stats.append({
            "Subject": display_name,
            "Total Students": s_total,
            "Appeared": s_appeared,
            "Absent": s_absent,
            "Passed": s_passed,
            "Failed": s_failed,
            "Pass %": s_pass_pct
        })
        
    subject_summary_df = pd.DataFrame(subject_stats)
    
    # If no subjects selected or found
    if subject_summary_df.empty:
        summary_card = html.Div(html.P("No subject data found.", className="text-muted"))
    else:
        # Create a Summary Table for Subject-wise stats
        summary_card = dbc.Card(dbc.CardBody([
            html.Div([
                html.H5("📚 Subject Level Performance", className="fw-bold mb-0 text-primary"),
                dbc.Button("Download Excel", id="sa-export-summary-xlsx", color="success", outline=True, size="sm")
            ], className="d-flex justify-content-between align-items-center mb-3"),
            dash_table.DataTable(
                id="sa-summary-table",
                data=subject_summary_df.to_dict('records'),
                columns=[{"name": i, "id": i} for i in subject_summary_df.columns],
                style_table={"overflowX": "auto", "borderRadius": "8px", "boxShadow": "0 4px 6px -1px rgba(0, 0, 0, 0.1)"},
                style_header={
                    "backgroundColor": "#1f2937",
                    "fontWeight": "700",
                    "color": "#ffffff",
                    "borderBottom": "2px solid #111827",
                    "padding": "12px",
                    "textTransform": "uppercase",
                    "fontSize": "12px",
                    "letterSpacing": "0.5px"
                },
                style_cell={
                    "textAlign": "center", 
                    "padding": "12px", 
                    "fontFamily": "Inter, sans-serif",
                    "fontSize": "14px",
                    "border": "1px solid #e2e8f0",
                    "color": "#1e293b"
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': '#f3f4f6'
                    },
                    {
                        "if": {"state": "selected"},
                        "backgroundColor": "rgba(59, 130, 246, 0.1)",
                        "border": "1px solid #3b82f6"
                    },
                    # Column-level color formatting
                    {
                        "if": {"column_id": "Total Students"},
                        "backgroundColor": "#eff6ff",
                        "color": "#1e3a8a",
                        "fontWeight": "600"
                    },
                    {
                        "if": {"column_id": "Appeared"},
                        "backgroundColor": "#e0f2fe",
                        "color": "#0369a1",
                        "fontWeight": "600"
                    },
                    {
                        "if": {"column_id": "Absent"},
                        "backgroundColor": "#fffbeb",
                        "color": "#b45309",
                        "fontWeight": "600"
                    },
                    {
                        "if": {"column_id": "Passed"},
                        "backgroundColor": "#ecfdf5",
                        "color": "#047857",
                        "fontWeight": "600"
                    },
                    {
                        "if": {"column_id": "Failed"},
                        "backgroundColor": "#fef2f2",
                        "color": "#b91c1c",
                        "fontWeight": "600"
                    },
                    # Conditional formatting for Pass %
                    {
                        "if": {
                            "filter_query": "{Pass %} >= 50",
                            "column_id": "Pass %"
                        },
                        "backgroundColor": "#d1fae5",
                        "color": "#065f46",
                        "fontWeight": "bold"
                    },
                    {
                        "if": {
                            "filter_query": "{Pass %} < 50",
                            "column_id": "Pass %"
                        },
                        "backgroundColor": "#fee2e2",
                        "color": "#991b1b",
                        "fontWeight": "bold"
                    }
                ],
                sort_action="native"
            )
        ]), className="sa-card mb-4")

    # Updated KPI Card style to match Overview page exactly

    cards = html.Div([
        dbc.Row([
            dbc.Col(
                # Changed from dbc.Card to html.Div to natively support n_clicks
                html.Div(
                    dbc.CardBody([
                        html.Div([
                            # Icon on the Left
                            html.Div(
                                html.I(className=f"bi {k['icon']} subject-kpi-icon", style={"color": k["color"]}),
                                className="subject-kpi-icon-box",
                                style={"backgroundColor": k["bg"]}
                            ),
                            
                            # Text on the Right
                            html.Div([
                                html.H6(k["label"], className="subject-kpi-label"),
                                html.H3(str(k["value"]), className="kpi-val subject-kpi-value", style={"color": k["color"]})
                            ], className="subject-kpi-text-box"),
                        ], className="subject-kpi-content-wrapper"),
                        
                        # Added Text Cue for Clickability (Hidden by default, shown on hover via CSS)
                        html.Div("👆 Click for details", className="kpi-hover-hint text-muted text-end mt-1", style={"fontSize": "0.6rem", "opacity": "0.8", "position": "absolute", "bottom": "8px", "right": "12px", "display": "none", "transition": "opacity 0.2s ease"}),

                    ], className="subject-kpi-body position-relative"),
                    className="card subject-kpi-card box-shadow-sm",
                    title=f"Click to view {k['label'].lower()} students list",
                    id={"type": "sa-kpi-card", "index": k["id"]},
                    n_clicks=0,
                    style={"--kpi-color": k['color'], "border": "none"}
                )
            )
            for k in kpis
        ], className=col_class + " mb-4"),  # added margin bottom to separate from summary table
        
        # INSERT SUMMARY CARD HERE
        summary_card
    ])



    # Table
    columns_for_table = []
    
    # Add Identity Columns first
    columns_for_table.append({"name": ["Student", "ID"], "id": first_col})
    columns_for_table.append({"name": ["Student", "Name"], "id": "Name"})
    columns_for_table.append({"name": ["Student", "Section"], "id": "Section"})

    # robust column grouping logic
    # Group columns by subject to ensure they appear together in the table
    # Sort columns by subject code/name first, then bu component order (Int, Ext, Tot, Res)
    
    def get_col_sort_key(col_name):
        # Extract subject prefix and component
        for s in ["Internal", "External", "Total", "Result"]:
            if col_name.endswith(f" {s}"):
                base = col_name[:-len(s)].strip()
                # Order: Internal=0, External=1, Total=2, Result=3
                order = {"Internal": 0, "External": 1, "Total": 2, "Result": 3}.get(s, 9)
                return (base, order)
        return (col_name, 99)

    selected_cols.sort(key=get_col_sort_key)

    for c in selected_cols:
        col_header = ["", c] # Default fallback
        
        for s in ["Internal", "External", "Total", "Result"]:
            if c.endswith(f" {s}"):
                base = c[:-len(s)].strip() # This extracts the full Subject Name e.g. "18CS51 - MATH"
                col_header = [base, s]
                break
            
        columns_for_table.append({"name": col_header, "id": c})

    columns_for_table.append({"name": ["Overall", "Result"], "id": "Overall_Result"})
    data = df_sel.to_dict("records")

    # Charts
    if chart_tab == "pie":
        # Use df_sel for pie counts (already filtered)
        pie_pass = df_sel["Overall_Result"].isin(pass_cats).sum()
        pie_fail = (df_sel["Overall_Result"] == "Fail").sum()
        pie_absent = (df_sel["Overall_Result"] == "Absent").sum()
        
        # Include Absent in pie chart when showing ALL filter
        if result_filter == "ALL":
            chart_values = [pie_pass, pie_fail, pie_absent]
            chart_labels = ["Pass", "Fail", "Absent"]
            chart_colors = ["#10b981", "#ef4444", "#f59e0b"]
        elif result_filter == "ABSENT":
            # When filtering for Absent only, show a pie with just absent count
            chart_values = [pie_absent]
            chart_labels = ["Absent"]
            chart_colors = ["#f59e0b"]
        else:
            chart_values = [pie_pass, pie_fail] if result_filter == "PASS" else [pie_fail] if result_filter == "FAIL" else [pie_pass]
            chart_labels = ["Pass", "Fail"] if result_filter == "PASS" else ["Fail"] if result_filter == "FAIL" else ["Pass"]
            chart_colors = ["#10b981", "#ef4444"] if result_filter == "PASS" else ["#ef4444"] if result_filter == "FAIL" else ["#10b981"]
        
        fig = px.pie(
            values=chart_values,
            names=chart_labels,
            color=chart_labels,
            color_discrete_map={label: color for label, color in zip(chart_labels, chart_colors)},
            hole=0.4
        )
        # Fix pie chart tooltip (clean style, no label=value=color)
        fig.update_traces(
            hovertemplate="<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>"
        )
        fig.update_layout(title="Pass vs Fail Distribution", title_x=0.5, template="plotly_white")
        chart = dcc.Graph(id="sa-pie-chart", figure=fig, style={"cursor": "pointer"})
    else:
        # Exclude Absent students from bar averages
        df_for_avg = df_sel[df_sel["Overall_Result"] != "Absent"].copy()
        
        # If all students are absent, show a message
        if df_for_avg.empty:
            chart = html.P("No students with marks to display. All selected students are absent.", className="text-muted text-center")
        else:
            # Build data for the bar chart with Full Names
            avg_marks_data = []
            
            for subj in selected_subjects:
                # 1. Identify Full Name (Same logic as tables)
                display_name = subj
                subj_cols = [c for c in df_sel.columns if c.startswith(subj)]
                for col in subj_cols:
                    if " - " in col:
                        try:
                            rest = col.split(" - ", 1)[1]
                            for suffix in ["Result", "Total", "Internal", "External"]:
                                if rest.strip().endswith(suffix):
                                    possible_name = rest.rsplit(suffix, 1)[0].strip()
                                    if possible_name:
                                        display_name = f"{subj} - {possible_name}"
                                    break
                            if display_name != subj: break
                        except: continue

                # 2. Calculate Average
                # Filter related Total columns
                total_cols = [c for c in selected_cols if subj in c and "Total" in c]
                if not total_cols:
                    continue
                
                # Calculate mean of totals for this subject across all students
                # Note: This averages the student's average if multiple totals exist (rare), or just the single total
                val = df_for_avg[total_cols].mean(axis=1).mean()
                
                avg_marks_data.append({"Subject_Code": subj, "Full_Name": display_name, "Average": val})

            if not avg_marks_data:
                 chart = html.P("No data available for chart.", className="text-muted text-center")
            else:
                df_chart = pd.DataFrame(avg_marks_data)
                
                bar_fig = px.bar(
                    df_chart, 
                    x="Subject_Code", 
                    y="Average",
                    text=df_chart["Average"].apply(lambda x: f"{x:.1f}"),
                    color="Subject_Code", 
                    color_discrete_sequence=px.colors.qualitative.Plotly,
                    custom_data=["Full_Name"] # Store full name for tooltip
                )
                
                # --- CUSTOM TOOLTIP & LAYOUT FIXES ---
                bar_fig.update_traces(
                    textposition="outside",
                    hovertemplate="<b>%{customdata[0]}</b><br>Avg Marks: %{y:.2f}<extra></extra>"
                )
                
                bar_fig.update_layout(
                    title="Average Total Marks per Subject", 
                    title_x=0.5, 
                    template="plotly_white",
                    yaxis_title="Average Marks", 
                    xaxis_title="Subject Code",
                    showlegend=False      # Hide legend as x-axis shows codes
                )
                
                # --- NEW PERFORMANCE BAR CHART (Pass %) ---
                if not subject_summary_df.empty:
                    # Create a copy so we don't accidentally mutate the underlying table data format inappropriately
                    perf_df = subject_summary_df.copy()
                    perf_df["Subject_Code"] = perf_df["Subject"].apply(lambda x: x.split(" - ")[0] if isinstance(x, str) and " - " in x else x)
                    
                    perf_fig = px.bar(
                        perf_df,
                        x="Subject_Code",
                        y="Pass %",
                        text=perf_df["Pass %"].apply(lambda x: f"{x}%"),
                        color="Subject_Code",
                        color_discrete_sequence=px.colors.qualitative.Pastel,
                        custom_data=["Subject"] # original display name
                    )
                    perf_fig.update_traces(
                        textposition="outside",
                        hovertemplate="<b>%{customdata[0]}</b><br>Pass Percentage: %{y}%<extra></extra>"
                    )
                    perf_fig.update_layout(
                        title="Subject-wise Pass Percentage",
                        title_x=0.5,
                        template="plotly_white",
                        yaxis_title="Pass Percentage (%)",
                        xaxis_title="Subject Code",
                        showlegend=False,
                        yaxis=dict(range=[0, min(100 + 15, 110)])  # keep headroom for label
                    )
                    
                    chart = dbc.Row([
                        dbc.Col(dcc.Graph(figure=bar_fig), xs=12, lg=6),
                        dbc.Col(dcc.Graph(figure=perf_fig), xs=12, lg=6)
                    ], className="g-3")
                else:
                    chart = dcc.Graph(figure=bar_fig)

    return f"{len(selected_subjects)} subjects selected", cards, columns_for_table, data, chart


# 3️⃣ Export Callbacks
@callback(
    Output("sa-download-csv", "data"),
    Input("sa-export-csv", "n_clicks"),
    Input("sa-export-detailed-csv", "n_clicks"),
    State('sa-subject-table', 'data'),
    State('sa-subject-table', 'columns'),
    prevent_initial_call=True
)
def export_csv(n_clicks_top, n_clicks_inline, table_data, table_columns):
    """Export the visible table to CSV."""
    # Check which button triggered the callback
    ctx = dash.callback_context
    if not ctx.triggered:
        return no_update
        
    trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]
    
    # If the triggered button has no clicks (e.g. initial render weirdness), prevent update
    if trigger_id == "sa-export-csv" and not n_clicks_top:
        return no_update
    if trigger_id == "sa-export-detailed-csv" and not n_clicks_inline:
        return no_update
        
    if not table_data:
        return no_update
    
    df = pd.DataFrame(table_data)
    
    # 🌟 CRITICAL FIX: Ensure mapping of columns exactly matches what displays on screen
    col_ids = [c['id'] for c in table_columns]
    valid_cols = [c for c in col_ids if c in df.columns]
    df = df[valid_cols]
    
    # Check if we have multi-level headers (which are represented as lists)
    is_multiindex = any(isinstance(col['name'], list) for col in table_columns)
    
    if is_multiindex:
        # Create MultiIndex for columns
        tuples = []
        for col in table_columns:
            if col['id'] not in valid_cols: continue
            
            if isinstance(col['name'], list):
                # Ensure all parts of the header are strings and pad if necessary to max depth
                # Most of our headers here are depth 2 (Category, Subcategory)
                tuples.append(tuple(str(x) for x in col['name']))
            else:
                # If it's not a list, copy the same name for both levels
                name = str(col['name'])
                tuples.append((name, name))
                
        df.columns = pd.MultiIndex.from_tuples(tuples)
    else:
        df.columns = [col['name'] for col in table_columns if col['id'] in valid_cols]
    
    return dcc.send_data_frame(df.to_csv, "subject_analysis.csv", index=False)

@callback(
    Output("sa-download-xlsx", "data"),
    Input("sa-export-xlsx", "n_clicks"),
    Input("sa-export-detailed-xlsx", "n_clicks"),
    State('sa-subject-table', 'data'),
    State('sa-subject-table', 'columns'),
    prevent_initial_call=True
)
def export_xlsx(n_clicks_top, n_clicks_inline, table_data, table_columns):
    """Export the visible table to Excel."""
    # Check which button triggered the callback
    ctx = dash.callback_context
    if not ctx.triggered:
        return no_update
        
    trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]
    
    # If the triggered button has no clicks, prevent update
    if trigger_id == "sa-export-xlsx" and not n_clicks_top:
        return no_update
    if trigger_id == "sa-export-detailed-xlsx" and not n_clicks_inline:
        return no_update
        
    if not table_data:
        return no_update
        
    df = pd.DataFrame(table_data)
    
    # 🌟 CRITICAL FIX: Ensure mapping of columns exactly matches what displays on screen
    col_ids = [c['id'] for c in table_columns]
    valid_cols = [c for c in col_ids if c in df.columns]
    df = df[valid_cols]
    
    # Check if we have multi-level headers (which are represented as lists)
    is_multiindex = any(isinstance(col['name'], list) for col in table_columns)
    
    if is_multiindex:
        # Create MultiIndex for columns
        tuples = []
        for col in table_columns:
            if col['id'] not in valid_cols: continue
            
            if isinstance(col['name'], list):
                tuples.append(tuple(str(x) for x in col['name']))
            else:
                name = str(col['name'])
                tuples.append((name, name))
                
        df.columns = pd.MultiIndex.from_tuples(tuples)
        
        # When using MultiIndex columns in Excel, pandas requires index=True.
        # So we create a blank hidden index so it aligns nicely below the multi-row header.
        df.index = [""] * len(df)
        df.index.name = None
        return dcc.send_data_frame(df.to_excel, "subject_analysis.xlsx", sheet_name="Subject Analysis", index=True)
    else:
        df.columns = [col['name'] for col in table_columns if col['id'] in valid_cols]
        return dcc.send_data_frame(df.to_excel, "subject_analysis.xlsx", sheet_name="Subject Analysis", index=False)

@callback(
    Output("sa-download-summary-xlsx", "data"),
    Input("sa-export-summary-xlsx", "n_clicks"),
    State('sa-summary-table', 'data'),
    State('sa-summary-table', 'columns'),
    prevent_initial_call=True
)
def export_summary_xlsx(n_clicks, table_data, table_columns):
    """Export the subject summary table to Excel."""
    if not n_clicks or not table_data:
        return no_update
        
    df = pd.DataFrame(table_data)
    
    # Simple headers for summary table, as they are not grouped
    flat_headers = [col['name'] for col in table_columns]
    df.columns = flat_headers

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name="Performance", index=False)
    
    workbook = writer.book
    worksheet = writer.sheets["Performance"]
    
    # Define styles matching dashboard UI (ARGB hex)
    styles = {
        "Total Students": {"bg": "FFEFF6FF", "font": "FF1E3A8A"},
        "Appeared": {"bg": "FFE0F2FE", "font": "FF0369A1"},
        "Absent": {"bg": "FFFFFBEB", "font": "FFB45309"},
        "Passed": {"bg": "FFECFDF5", "font": "FF047857"},
        "Failed": {"bg": "FFFEF2F2", "font": "FFB91C1C"}
    }
    pass_high_style = {"bg": "FFD1FAE5", "font": "FF065F46"}
    pass_low_style = {"bg": "FFFEE2E2", "font": "FF991B1B"}
    
    # Apply column styles
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Header style
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = Font(bold=True, color="FFFFFFFF")
        header_cell.fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column width
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = max(len(str(col_name)) + 5, 15)
        if col_name == "Subject":
            worksheet.column_dimensions[col_letter].width = 50
            
        # Data cells
        for row_idx, val in enumerate(df[col_name], start=2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply our column-based colors
            if col_name in styles:
                cell.fill = PatternFill(start_color=styles[col_name]["bg"], end_color=styles[col_name]["bg"], fill_type="solid")
                cell.font = Font(color=styles[col_name]["font"], bold=True)
            elif col_name == "Pass %":
                try:
                    num_val = float(val)
                    if num_val >= 50:
                        cell.fill = PatternFill(start_color=pass_high_style["bg"], end_color=pass_high_style["bg"], fill_type="solid")
                        cell.font = Font(color=pass_high_style["font"], bold=True)
                    else:
                        cell.fill = PatternFill(start_color=pass_low_style["bg"], end_color=pass_low_style["bg"], fill_type="solid")
                        cell.font = Font(color=pass_low_style["font"], bold=True)
                except ValueError:
                    pass
            elif row_idx % 2 == 1:
                # Generic odd row coloring (matching dashboard table)
                cell.fill = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")

    writer.close()
    output.seek(0)
    
    return dcc.send_bytes(output.getvalue(), "subject_level_performance.xlsx")

@callback(
    Output("sa-legend-modal", "is_open"),
    [Input("sa-open-legend", "n_clicks"), Input("sa-close-legend", "n_clicks")],
    [State("sa-legend-modal", "is_open")],
)
def sa_toggle_legend(n1, n2, is_open):
    if n1 or n2:
        return not is_open
    return is_open

dash.clientside_callback(
    """
    function(n_clicks) {
        if (!n_clicks) {
            return window.dash_clientside.no_update;
        }
        setTimeout(function () { window.print(); }, 150);
        return "";
    }
    """,
    Output("sa-pdf-download-trigger", "children"),
    Input("sa-export-pdf", "n_clicks"),
    prevent_initial_call=True
)

# 4️⃣ KPI Target Click Callback
@callback(
    Output("sa-kpi-modal", "is_open"),
    Output("sa-kpi-modal-title", "children"),
    Output("sa-kpi-modal-table", "data"),
    Output("sa-kpi-modal-table", "columns"),
    Input({"type": "sa-kpi-card", "index": ALL}, "n_clicks"),
    Input("sa-pie-chart", "clickData"),
    Input("sa-kpi-modal-close", "n_clicks"),
    Input("sa-kpi-modal-close-top", "n_clicks"),
    State("sa-subject-table", "data"),
    State("sa-subject-table", "columns"),
    prevent_initial_call=True
)
def handle_kpi_click(kpi_clicks, pie_click, close_click, close_click_top, table_data, table_cols):
    ctx = dash.callback_context
    if not ctx.triggered:
        raise PreventUpdate
    
    trigger_id = ctx.triggered[0]["prop_id"]
    
    # Check if close button triggered (top or bottom)
    if "sa-kpi-modal-close" in trigger_id:
        return False, dash.no_update, dash.no_update, dash.no_update
        
    kpi_type = None

    # Handle Pie Chart Click
    if "sa-pie-chart" in trigger_id:
        if not pie_click:
            raise PreventUpdate
        point_data = pie_click['points'][0]
        clicked_label = point_data.get('label', '').lower()
        if clicked_label == "pass":
            kpi_type = "pass"
        elif clicked_label == "fail":
            kpi_type = "fail"
        elif clicked_label == "absent":
            kpi_type = "absent"
        else:
            raise PreventUpdate

    # Handle KPI Card Click
    else:
        # Ignore initial setup triggers where all clicks might be 0/None
        if all(c == 0 or c is None for c in kpi_clicks):
            raise PreventUpdate
            
        # Extract which specific card was clicked
        try:
            prop_dict = json.loads(trigger_id.split(".")[0])
            kpi_type = prop_dict["index"]
        except Exception:
            raise PreventUpdate
        
    if not table_data:
        return True, f"Student List: {kpi_type.upper() if kpi_type else 'UNKNOWN'}", [], table_cols
        
    # Filter the exact data subset for the clicked KPI
    filtered_data = []
    if kpi_type == "total":
        filtered_data = table_data
    elif kpi_type == "appeared":
        filtered_data = [row for row in table_data if row.get("Overall_Result") != "Absent"]
    elif kpi_type == "pass":
        filtered_data = [row for row in table_data if row.get("Overall_Result") in ["Pass", "FCD", "FC", "SC"]]
    elif kpi_type == "fail":
        filtered_data = [row for row in table_data if row.get("Overall_Result") == "Fail"]
    elif kpi_type == "absent":
        filtered_data = [row for row in table_data if row.get("Overall_Result") == "Absent"]
    elif kpi_type == "rate":
        # Treating Rate % click as looking at passed students
        filtered_data = [row for row in table_data if row.get("Overall_Result") in ["Pass", "FCD", "FC", "SC"]]
        kpi_type = "pass"
        
    title = f"📃 Detail List: {kpi_type.title()} ({len(filtered_data)} Students)"
    
    return True, title, filtered_data, table_cols

# Excel Download directly inside Modal
@callback(
    Output("sa-kpi-excel-download", "data"),
    Input("sa-kpi-modal-excel", "n_clicks"),
    Input("sa-kpi-modal-excel-top", "n_clicks"),
    State("sa-kpi-modal-table", "data"),
    State("sa-kpi-modal-table", "columns"),
    State("sa-kpi-modal-title", "children"),
    prevent_initial_call=True
)
def download_modal_excel(n_clicks_bottom, n_clicks_top, table_data, table_cols, title):
    if not dash.ctx.triggered or not table_data:
        raise PreventUpdate
        
    df = pd.DataFrame(table_data)
    
    # Check if headers are configured as a multi-index list (as in the screenshot)
    if table_cols and isinstance(table_cols[0].get('name'), list):
        # We need to construct a proper multi-index dataframe headers for excel export
        cols_to_keep = [col['id'] for col in table_cols if col['id'] in df.columns]
        if cols_to_keep:
            df = df[cols_to_keep]
            
        # Extract the multi-level names
        multi_names = [tuple(col['name']) for col in table_cols if col['id'] in cols_to_keep]
        
        # Assign multi-index column to pandas DataFrame
        df.columns = pd.MultiIndex.from_tuples(multi_names)
    elif table_cols:
        col_map = {col['id']: col['name'] for col in table_cols if 'name' in col and 'id' in col}
        # Only keep columns that are in table_cols and rename them
        cols_to_keep = [col['id'] for col in table_cols if col['id'] in df.columns]
        if cols_to_keep:
            df = df[cols_to_keep]
        df.rename(columns=col_map, inplace=True)
    
    # Create a clean filename from the modal title
    safe_title = "Student_List"
    if title and isinstance(title, str):
        # Extract just the category part before the parentheses
        clean_str = title.split('(')[0].replace('📃', '').replace('Detail List:', '').strip()
        safe_title = re.sub(r'[^A-Za-z0-9_]', '_', clean_str)
        
    # Apply the MultiIndex Excel index workaround
    if isinstance(df.columns, pd.MultiIndex):
        df.index = [""] * len(df)
        df.index.name = None
        return dcc.send_data_frame(df.to_excel, f"{safe_title}_Report.xlsx", index=True)
    else:
        return dcc.send_data_frame(df.to_excel, f"{safe_title}_Report.xlsx", index=False)

@callback(
    Output("sa-download-all-kpis", "data"),
    Input("sa-export-all-kpis", "n_clicks"),
    State("sa-kpi-modal-table", "data"), # It's better to fetch from the master dataset
    State("stored-data", "data"),
    State("sa-subject-selector", "value"),
    State("sa-section-filter", "value"),
    State("sa-result-filter", "value"),
    prevent_initial_call=True
)
def download_all_kpis_subject_analysis(n_clicks, kpi_table_data, stored_data, subject, section, result_filter):
    if not n_clicks or not stored_data or not subject:
        raise PreventUpdate
        
    df = pd.DataFrame(stored_data)
    
    # Check if this subject really exists in data
    subject_cols = [c for c in df.columns if c.endswith(' Total')]
    subj_base_names = [c.replace(' Total', '') for c in subject_cols]
    
    # Use exact subject name matching
    target_base = subject
    if target_base not in subj_base_names:
        # User might have passed full code + name combo. The data has "Code Components"
        target_base = subject

    # Filter section if selected
    if section and section != "ALL" and "Section" in df.columns:
        df = df[df["Section"] == section]
        
    int_col = f"{target_base} Internal"
    ext_col = f"{target_base} External"
    tot_col = f"{target_base} Total"
    res_col = f"{target_base} Result"
    
    req_cols = [c for c in [int_col, ext_col, tot_col, res_col] if c in df.columns]
    if not req_cols:
        raise PreventUpdate

    # Filter by result 
    if res_col in df.columns and result_filter != "ALL":
        if result_filter == "PASS":
            df = df[df[res_col].astype(str).str.upper().isin(["P", "PASS"])]
        elif result_filter == "FAIL":
            df = df[df[res_col].astype(str).str.upper().isin(["F", "FAIL"])]
        elif result_filter == "ABSENT":
            df = df[df[res_col].astype(str).str.upper().isin(["A", "ABSENT"])]
            
    # Standardize result text
    df['Overall_Result'] = df[res_col] if res_col in df.columns else "Unknown"

    kpi_definitions = [
        ('Total Students', df),
        ('Appeared', df[df['Overall_Result'] != 'Absent']),
        ('Passed', df[df['Overall_Result'].isin(['Pass', 'FCD', 'FC', 'SC'])]),
        ('Failed', df[df['Overall_Result'] == 'Fail']),
        ('Absent', df[df['Overall_Result'] == 'Absent'])
    ]
    
    from io import BytesIO
    out = BytesIO()
    writer = pd.ExcelWriter(out, engine='openpyxl')
    
    has_sheets = False
    for sheet_name, scoped_df in kpi_definitions:
        if scoped_df.empty: continue
            
        display_cols = ['Student_ID', 'Name', 'Section']
        for c in [int_col, ext_col, tot_col, res_col]:
            if c in scoped_df.columns:
                display_cols.append(c)
                
        export_df = scoped_df[display_cols].copy()
        
        # Friendly Headers
        col_map = {
            'Student_ID': 'Student ID',
            'Name': 'Student Name',
            'Section': 'Section',
            int_col: 'Internal Marks',
            ext_col: 'External Marks',
            tot_col: 'Total Marks',
            res_col: 'Overall Result'
        }
        export_df.rename(columns=col_map, inplace=True)
        
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)
        has_sheets = True
        
    if not has_sheets:
        empty_df = pd.DataFrame({"Message": ["No data available"]})
        empty_df.to_excel(writer, sheet_name="No Data", index=False)
        
    writer.close()
    out.seek(0)
    
    return dcc.send_bytes(out.read(), "Consolidated_Subject_KPIs.xlsx")